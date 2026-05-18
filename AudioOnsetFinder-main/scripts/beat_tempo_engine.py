"""Reusable beat and tempo processing helpers for the rhythm pipeline."""

from __future__ import annotations

from dataclasses import dataclass
import os
import warnings

import librosa
import numpy as np


AUDIO_EXTENSIONS = {".wav", ".mp3", ".flac", ".ogg", ".aiff", ".aif", ".m4a", ".mp4"}

BEAT_SUMMARY_HEADERS = [
    "File Name",
    "Beat Tempo Estimate (BPM)",
    "Beat Times Used (s)",
    "Beat Count",
    "Beat Hop Length",
    "Beat Start BPM",
    "Beat Tightness",
    "Beat Trim Enabled",
    "Beat BPM Override",
    "PLP Hop Length",
    "PLP Win Length",
    "PLP Tempo Min BPM",
    "PLP Tempo Max BPM",
    "PLP Peak Count",
    "PLP Peak Times (s)",
]

DETAIL_HEADERS = [
    "File Name",
    "Frame Index",
    "Time (s)",
    "Pulse Value",
    "Is Pulse Local Max",
]


@dataclass(frozen=True)
class BeatTempoConfig:
    """Runtime configuration for the beat and tempo step."""

    audio_folder: str
    output_excel_path: str
    output_sheet_name: str = "Beat-Tempo"
    beat_hop_length: int = 512
    beat_start_bpm: float = 120.0
    beat_tightness: float = 100.0
    beat_trim: bool = True
    beat_bpm_override: float = 0.0
    plp_hop_length: int = 512
    plp_win_length: int = 384
    plp_tempo_min: float = 30.0
    plp_tempo_max: float = 300.0
    export_pulse_detail: bool = False
    derive_plp_peaks: bool = True
    plp_peak_min_spacing_ms: int = 100


def find_audio_files(folder: str) -> list[str]:
    """Return sorted list of supported audio basenames in *folder*."""
    if not os.path.isdir(folder):
        return []
    return sorted(
        entry
        for entry in os.listdir(folder)
        if os.path.splitext(entry)[1].lower() in AUDIO_EXTENSIONS
    )


def plp_local_maxima(
    pulse_curve: np.ndarray,
    pulse_times: np.ndarray,
    min_spacing_ms: float,
) -> np.ndarray:
    """Return PLP local-max times using a greedy minimum-spacing rule."""
    if len(pulse_curve) < 3:
        return np.array([], dtype=float)

    is_max = np.zeros(len(pulse_curve), dtype=bool)
    is_max[1:-1] = (
        (pulse_curve[1:-1] > pulse_curve[:-2])
        & (pulse_curve[1:-1] > pulse_curve[2:])
    )
    candidate_idx = np.where(is_max)[0]
    if len(candidate_idx) == 0:
        return np.array([], dtype=float)

    hop_duration_sec = (pulse_times[1] - pulse_times[0]) if len(pulse_times) > 1 else 1.0
    min_spacing_frames = max(1, int(round((min_spacing_ms / 1000.0) / hop_duration_sec)))

    kept_idx = [candidate_idx[0]]
    for idx in candidate_idx[1:]:
        if idx - kept_idx[-1] >= min_spacing_frames:
            kept_idx.append(idx)
        elif pulse_curve[idx] > pulse_curve[kept_idx[-1]]:
            kept_idx[-1] = idx

    return pulse_times[np.array(kept_idx, dtype=int)]


def safe_scalar(value) -> float:
    """Convert a numpy scalar or 0-d array to a Python float safely."""
    if isinstance(value, np.ndarray):
        return float(value.flat[0])
    return float(value)


def format_times(times_sec, decimals: int = 6) -> str:
    """Serialize a sequence of times to a comma-separated string."""
    return ", ".join(f"{t:.{decimals}f}" for t in sorted(times_sec))


def load_or_create_workbook(path: str):
    """Load an existing workbook or create a new one with no default sheet."""
    import openpyxl

    if os.path.isfile(path):
        return openpyxl.load_workbook(path)

    output_dir = os.path.dirname(path)
    if output_dir:
        os.makedirs(output_dir, exist_ok=True)

    workbook = openpyxl.Workbook()
    if "Sheet" in workbook.sheetnames:
        del workbook["Sheet"]
    return workbook


def ws_header_map(ws) -> dict[str, int]:
    """Return {column title: 1-based column index} for the first worksheet row."""
    return {
        str(cell.value).strip(): cell.column
        for cell in ws[1]
        if cell.value is not None
    }


def ensure_column(ws, header_map: dict[str, int], col_title: str) -> int:
    """Add *col_title* to the header row if absent and return its index."""
    if col_title in header_map:
        return header_map[col_title]
    next_col = max(header_map.values(), default=0) + 1
    ws.cell(row=1, column=next_col, value=col_title)
    header_map[col_title] = next_col
    return next_col


def find_summary_row(ws, header_map: dict[str, int], basename: str) -> int:
    """Return the 1-based summary row whose File Name matches *basename*."""
    if "File Name" not in header_map:
        return 0
    name_col = header_map["File Name"]
    target = basename.strip().lower()
    for row_idx in range(2, ws.max_row + 1):
        cell_value = ws.cell(row=row_idx, column=name_col).value
        if cell_value is not None and str(cell_value).strip().lower() == target:
            return row_idx
    return 0


def append_summary_row(ws, header_map: dict[str, int], basename: str) -> int:
    """Append a new summary row for *basename* and return its row index."""
    if "File Name" not in header_map:
        raise ValueError("Beat summary sheet has no 'File Name' column.")
    next_row = ws.max_row + 1
    ws.cell(row=next_row, column=header_map["File Name"], value=basename)
    return next_row


def write_summary_columns(ws, header_map: dict[str, int], row: int, data: dict) -> None:
    """Write *data* {column_title: value} into *row* of *ws*."""
    for col_title, value in data.items():
        col_idx = ensure_column(ws, header_map, col_title)
        ws.cell(row=row, column=col_idx, value=value)


def ensure_summary_sheet(workbook, sheet_name: str):
    """Return the beat summary worksheet plus its header map."""
    if sheet_name not in workbook.sheetnames:
        ws_summary = workbook.create_sheet(sheet_name)
    else:
        ws_summary = workbook[sheet_name]

    for col_idx, header in enumerate(BEAT_SUMMARY_HEADERS, start=1):
        if ws_summary.cell(row=1, column=col_idx).value is None:
            ws_summary.cell(row=1, column=col_idx, value=header)

    header_map = ws_header_map(ws_summary)
    ensure_column(ws_summary, header_map, "File Name")
    return ws_summary, header_map


def ensure_detail_sheet(workbook, detail_sheet_name: str):
    """Return the optional pulse-detail worksheet."""
    if detail_sheet_name not in workbook.sheetnames:
        ws_detail = workbook.create_sheet(detail_sheet_name)
    else:
        ws_detail = workbook[detail_sheet_name]

    for col_idx, header in enumerate(DETAIL_HEADERS, start=1):
        if ws_detail.cell(row=1, column=col_idx).value is None:
            ws_detail.cell(row=1, column=col_idx, value=header)
    return ws_detail


def clear_detail_rows(ws_detail, basename: str) -> None:
    """Remove existing detail rows for *basename* before rewriting them."""
    rows_to_delete = []
    for row_idx in range(2, ws_detail.max_row + 1):
        if ws_detail.cell(row=row_idx, column=1).value == basename:
            rows_to_delete.append(row_idx)
    for row_idx in reversed(rows_to_delete):
        ws_detail.delete_rows(row_idx)


def write_pulse_detail_rows(
    ws_detail,
    basename: str,
    pulse_times: np.ndarray,
    pulse_curve: np.ndarray,
    plp_peak_times: np.ndarray,
) -> None:
    """Write optional frame-level PLP detail rows for one file."""
    clear_detail_rows(ws_detail, basename)

    plp_peak_set = set()
    if len(plp_peak_times) > 0 and len(pulse_times) > 1:
        hop_sec = pulse_times[1] - pulse_times[0]
        for peak_time in plp_peak_times:
            frame_index = int(round((peak_time - pulse_times[0]) / hop_sec))
            if 0 <= frame_index < len(pulse_curve):
                plp_peak_set.add(frame_index)

    for frame_idx, (time_sec, pulse_value) in enumerate(zip(pulse_times, pulse_curve)):
        ws_detail.append(
            [
                basename,
                frame_idx,
                round(float(time_sec), 6),
                round(float(pulse_value), 8),
                1 if frame_idx in plp_peak_set else 0,
            ]
        )


def process_file(file_path: str, config: BeatTempoConfig) -> dict:
    """Run beat tracking and PLP on one audio file."""
    y, sr = librosa.load(file_path, sr=None, mono=True)

    beat_kwargs = {
        "y": y,
        "sr": sr,
        "hop_length": config.beat_hop_length,
        "start_bpm": config.beat_start_bpm,
        "tightness": config.beat_tightness,
        "trim": config.beat_trim,
        "units": "time",
    }
    if config.beat_bpm_override > 0.0:
        beat_kwargs["bpm"] = config.beat_bpm_override

    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        tempo_raw, beat_times = librosa.beat.beat_track(**beat_kwargs)

    beat_tempo_bpm = safe_scalar(tempo_raw)
    beat_times = np.sort(np.asarray(beat_times, dtype=float))

    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        pulse_curve = librosa.beat.plp(
            y=y,
            sr=sr,
            hop_length=config.plp_hop_length,
            win_length=config.plp_win_length,
            tempo_min=config.plp_tempo_min,
            tempo_max=config.plp_tempo_max,
        )

    pulse_times = librosa.frames_to_time(
        np.arange(len(pulse_curve)),
        sr=sr,
        hop_length=config.plp_hop_length,
    )

    plp_peak_times = np.array([], dtype=float)
    if config.derive_plp_peaks and len(pulse_curve) >= 3:
        plp_peak_times = plp_local_maxima(
            pulse_curve,
            pulse_times,
            config.plp_peak_min_spacing_ms,
        )

    return {
        "beat_tempo_bpm": beat_tempo_bpm,
        "beat_times": beat_times,
        "pulse_curve": np.asarray(pulse_curve, dtype=float),
        "pulse_times": np.asarray(pulse_times, dtype=float),
        "plp_peak_times": np.asarray(plp_peak_times, dtype=float),
    }


def build_summary_data(result: dict, config: BeatTempoConfig) -> dict:
    """Build the per-file Excel summary payload from a process result."""
    beat_times = result["beat_times"]
    plp_peak_times = result["plp_peak_times"]
    return {
        "Beat Tempo Estimate (BPM)": round(result["beat_tempo_bpm"], 4),
        "Beat Times Used (s)": format_times(beat_times),
        "Beat Count": len(beat_times),
        "Beat Hop Length": config.beat_hop_length,
        "Beat Start BPM": config.beat_start_bpm,
        "Beat Tightness": config.beat_tightness,
        "Beat Trim Enabled": config.beat_trim,
        "Beat BPM Override": config.beat_bpm_override if config.beat_bpm_override > 0 else "",
        "PLP Hop Length": config.plp_hop_length,
        "PLP Win Length": config.plp_win_length,
        "PLP Tempo Min BPM": config.plp_tempo_min,
        "PLP Tempo Max BPM": config.plp_tempo_max,
        "PLP Peak Count": len(plp_peak_times),
        "PLP Peak Times (s)": format_times(plp_peak_times) if len(plp_peak_times) > 0 else "",
    }


def run_beat_tempo_step(config: BeatTempoConfig, *, print_fn=print) -> int:
    """Execute the beat and tempo step using the supplied configuration."""
    print_fn("=" * 60)
    print_fn("  BEAT AND TEMPO STEP")
    print_fn("=" * 60)
    print_fn(f"  Audio folder  : {config.audio_folder}")
    print_fn(f"  Output Excel  : {config.output_excel_path}")
    print_fn(f"  Sheet name    : {config.output_sheet_name}")
    print_fn("")

    if not config.output_excel_path or not config.output_excel_path.strip():
        print_fn("  ERROR: output_excel_path is empty.")
        print_fn("  Please set the output Excel path in pipeline_config.json")
        print_fn("  or via the GUI's 'Output Excel file' field.")
        return 1

    audio_files = find_audio_files(config.audio_folder)
    if not audio_files:
        print_fn(f"  WARNING: No audio files found in: {config.audio_folder}")
        print_fn("  Beat and Tempo step skipped.")
        return 0

    print_fn(f"  Found {len(audio_files)} audio file(s) to process.\n")

    try:
        workbook = load_or_create_workbook(config.output_excel_path)
    except ImportError:
        print_fn("  ERROR: openpyxl not installed. Install with: pip install openpyxl")
        return 1
    except Exception as exc:
        print_fn(f"  ERROR: Cannot open Excel file: {exc}")
        return 1

    ws_summary, header_map = ensure_summary_sheet(workbook, config.output_sheet_name)
    ws_detail = None
    if config.export_pulse_detail:
        ws_detail = ensure_detail_sheet(
            workbook,
            f"{config.output_sheet_name} (Detail)",
        )

    n_ok = 0
    n_fail = 0
    for basename in audio_files:
        file_path = os.path.join(config.audio_folder, basename)
        print_fn(f"  Processing: {basename} ...", end=" ", flush=True)
        try:
            result = process_file(file_path, config)
        except Exception as exc:
            print_fn(f"FAILED ({type(exc).__name__}: {exc})")
            n_fail += 1
            continue

        row = find_summary_row(ws_summary, header_map, basename)
        if row == 0:
            row = append_summary_row(ws_summary, header_map, basename)

        write_summary_columns(ws_summary, header_map, row, build_summary_data(result, config))

        if ws_detail is not None:
            write_pulse_detail_rows(
                ws_detail,
                basename,
                result["pulse_times"],
                result["pulse_curve"],
                result["plp_peak_times"],
            )

        beats_str = f"{len(result['beat_times'])} beat(s)"
        peaks_str = f"{len(result['plp_peak_times'])} PLP peak(s)"
        tempo_str = f"{result['beat_tempo_bpm']:.1f} BPM"
        print_fn(f"OK  ({tempo_str}, {beats_str}, {peaks_str})")
        n_ok += 1

    try:
        workbook.save(config.output_excel_path)
    except Exception as exc:
        print_fn(f"  ERROR: Failed to save Excel workbook: {exc}")
        return 1

    print_fn("")
    print_fn(f"  Saved beat/tempo results to: {config.output_excel_path}")
    print_fn(f"  Completed: {n_ok} succeeded, {n_fail} failed.")
    return 0 if n_ok > 0 or n_fail == 0 else 1


__all__ = [
    "AUDIO_EXTENSIONS",
    "BEAT_SUMMARY_HEADERS",
    "DETAIL_HEADERS",
    "BeatTempoConfig",
    "append_summary_row",
    "build_summary_data",
    "ensure_column",
    "ensure_detail_sheet",
    "ensure_summary_sheet",
    "find_audio_files",
    "find_summary_row",
    "format_times",
    "load_or_create_workbook",
    "plp_local_maxima",
    "process_file",
    "run_beat_tempo_step",
    "safe_scalar",
    "write_pulse_detail_rows",
    "write_summary_columns",
    "ws_header_map",
]