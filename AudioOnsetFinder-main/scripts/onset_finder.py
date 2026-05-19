"""Extract rhythmic timing features from audio files and export analysis artifacts.

This is the core pipeline script. It scans the ``audioFiles/`` directory, detects
onsets in each recording, computes interval-based rhythm metrics, writes Audacity
label tracks and spectrogram review images, and exports both file-level summaries
and dyadic event tables to a multi-sheet Excel workbook.

Current role in the pipeline:
- Input: audio recordings in ``audioFiles/``
- Output: ``Cross_Species_Rhythm_Data.xlsx`` plus optional labels and spectrograms

Implemented next-step goals from ``docs/ReadMe.md`` and ``docs/nextSteps.md``:
- Cluster onsets that occur within 25 ms to reduce false dyads from overlapping signals.
- Export a stable-rhythm subset alongside the full dyadic dataset.
- Pre-filter audio with a high-pass filter and gate onsets by local amplitude to reduce
  false positives from background noise (configurable toggles in section 1).
- Preserve manual-review artifacts so noisy files can still be checked visually.
"""

import os  # directory listing and file path joining
import sys  # exit with status code on fatal errors

import librosa  # core audio analysis (onsets, RMS, beat tracking)
import librosa.display  # spectrogram plotting utilities for QA images
import matplotlib.pyplot as plt  # generating plots for visual outputs
import numpy as np  # array math for intervals, normalization, stats
import pandas as pd  # dataframes for summaries and Excel export
from scipy.signal import butter, sosfilt  # filter implementation used for high-pass pre-filter

from onset_detectors import (detect_onsets, refine_onsets_to_sample, available_methods,
                             last_f0_metrics, last_whisper_transcript,
                             last_whisperx_phonemes, last_ramus_metrics,
                             last_madmom_tempo, last_madmom_downbeats,
                             extract_pitch_metrics)
try:
    from .onset_batching import (
        ONSET_SETTING_NAMES,
        load_batch_routing_inputs,
        resolve_audio_files,
        snapshot_module_settings,
        temporary_module_settings,
    )
except ImportError:
    from onset_batching import (
        ONSET_SETTING_NAMES,
        load_batch_routing_inputs,
        resolve_audio_files,
        snapshot_module_settings,
        temporary_module_settings,
    )
try:
    from .onset_routing import build_detector_call
except ImportError:
    from onset_routing import build_detector_call
try:
    from .onset_postprocessing import (
        apply_highpass,
        apply_focus_regions,
        build_spectral_profile,
        cluster_onsets,
        enforce_min_interval,
        gate_onsets_by_broadband,
        gate_onsets_by_amplitude,
        gate_onsets_by_sharpness,
        gate_onsets_by_spectral_match,
        gate_onsets_by_variable_bounds,
    )
except ImportError:
    from onset_postprocessing import (
        apply_highpass,
        apply_focus_regions,
        build_spectral_profile,
        cluster_onsets,
        enforce_min_interval,
        gate_onsets_by_broadband,
        gate_onsets_by_amplitude,
        gate_onsets_by_sharpness,
        gate_onsets_by_spectral_match,
        gate_onsets_by_variable_bounds,
    )
try:
    from .onset_metrics import (
        build_dyad_records as _build_dyad_records_impl,
        calculate_rhythm_metrics as _calculate_rhythm_metrics_impl,
        calculate_speech_rhythm_metrics as _calculate_speech_rhythm_metrics_impl,
        calculate_stable_subset_metrics as _calculate_stable_subset_metrics_impl,
        get_stable_dyad_flags as _get_stable_dyad_flags_impl,
        is_stable_match as _is_stable_match_impl,
    )
    from .onset_metadata import (
        DEMOGRAPHIC_FIELDS,
        build_file_demographics_rows,
        load_user_metadata,
        merge_user_metadata_into_summary_rows,
    )
    from .onset_exports import (
        write_audacity_labels,
        write_praat_textgrid_export,
        write_whisper_transcript_exports,
    )
    from .shared_output_writers import (
        save_matplotlib_figure,
        save_openpyxl_workbook,
        write_workbook_sheets,
    )
except ImportError:
    from onset_metrics import (
        build_dyad_records as _build_dyad_records_impl,
        calculate_rhythm_metrics as _calculate_rhythm_metrics_impl,
        calculate_speech_rhythm_metrics as _calculate_speech_rhythm_metrics_impl,
        calculate_stable_subset_metrics as _calculate_stable_subset_metrics_impl,
        get_stable_dyad_flags as _get_stable_dyad_flags_impl,
        is_stable_match as _is_stable_match_impl,
    )
    from onset_metadata import (
        DEMOGRAPHIC_FIELDS,
        build_file_demographics_rows,
        load_user_metadata,
        merge_user_metadata_into_summary_rows,
    )
    from onset_exports import (
        write_audacity_labels,
        write_praat_textgrid_export,
        write_whisper_transcript_exports,
    )
    from shared_output_writers import (
        save_matplotlib_figure,
        save_openpyxl_workbook,
        write_workbook_sheets,
    )

# ==========================================
# 1. MAIN CONFIGURATION & TOGGLES
# ==========================================
# --- Named Presets ---
# Choose a preset to load tuned defaults for a particular dataset type, or set
# ACTIVE_PRESET to None to use the manual values below. Each preset overrides
# only the onset-detection and noise-filter parameters; paths and output toggles
# are always controlled by the manual values below the preset dictionary.
PRESETS = {
    "birdsong": {
        "description": "Small passerines with rapid, high-frequency calls",
        "HIGHPASS_CUTOFF_HZ": 500,
        "ONSET_AMPLITUDE_GATE": 0.08,
        "ONSET_AMPLITUDE_WINDOW_MS": 30,
        "ONSET_SHARPNESS_GATE": 0.0,
        "ONSET_SHARPNESS_WINDOW_MS": 20,
        "MIN_INTER_ONSET_MS": 12,
        "ONSET_CLUSTER_WINDOW_MS": 15,
        "STABLE_RHYTHM_TOLERANCE": 0.25,
        "ONSET_DELTA": 0.05,
        "ONSET_HOP_LENGTH": 256,
        "ONSET_BACKTRACK": False,
        "ONSET_REFINE_ENABLED": True,
        "ONSET_REFINE_WINDOW_MS": 10,
        "ONSET_REFINE_ENERGY_GATE": 0.0,
        "ONSET_METHOD": "adaptive_hp",
    },
    "percussion_clean": {
        "description": "Clean percussion recordings — sensitive to quiet hits and rapid sequences",
        "HIGHPASS_CUTOFF_HZ": 80,
        "ONSET_AMPLITUDE_GATE": 0.005,
        "ONSET_AMPLITUDE_WINDOW_MS": 50,
        "ONSET_SHARPNESS_GATE": 0.0,
        "ONSET_SHARPNESS_WINDOW_MS": 20,
        "ONSET_BROADBAND_MIN_BANDS": 3,
        "ONSET_BROADBAND_N_BANDS": 6,
        "ONSET_BROADBAND_THRESHOLD": 0.15,
        "MIN_INTER_ONSET_MS": 15,
        "ONSET_CLUSTER_WINDOW_MS": 12,
        "STABLE_RHYTHM_TOLERANCE": 0.25,
        "ONSET_DELTA": 0.04,
        "ONSET_HOP_LENGTH": 128,
        "ONSET_BACKTRACK": False,
        "ONSET_REFINE_ENABLED": True,
        "ONSET_REFINE_WINDOW_MS": 8,
        "ONSET_REFINE_ENERGY_GATE": 0.005,
        "ONSET_METHOD": "adaptive_hp",
    },
    "percussion_messy": {
        "description": "Noisy percussion and drumming with sharp transient attacks",
        "HIGHPASS_CUTOFF_HZ": 80,
        "ONSET_AMPLITUDE_GATE": 0.03,
        "ONSET_AMPLITUDE_WINDOW_MS": 50,
        "ONSET_SHARPNESS_GATE": 0.10,
        "ONSET_SHARPNESS_WINDOW_MS": 20,
        "MIN_INTER_ONSET_MS": 30,
        "ONSET_CLUSTER_WINDOW_MS": 25,
        "STABLE_RHYTHM_TOLERANCE": 0.25,
        "ONSET_DELTA": 0.07,
        "ONSET_HOP_LENGTH": 256,
        "ONSET_BACKTRACK": False,
        "ONSET_REFINE_ENABLED": True,
        "ONSET_REFINE_WINDOW_MS": 10,
        "ONSET_REFINE_ENERGY_GATE": 0.0,
        "ONSET_METHOD": "adaptive_hp",
    },
    "primate": {
        "description": "Primate calls and drumming — mid-frequency with moderate attack",
        "HIGHPASS_CUTOFF_HZ": 150,
        "ONSET_AMPLITUDE_GATE": 0.06,
        "ONSET_AMPLITUDE_WINDOW_MS": 60,
        "ONSET_SHARPNESS_GATE": 0.15,
        "ONSET_SHARPNESS_WINDOW_MS": 25,
        "MIN_INTER_ONSET_MS": 40,
        "ONSET_CLUSTER_WINDOW_MS": 30,
        "STABLE_RHYTHM_TOLERANCE": 0.30,
        "ONSET_DELTA": 0.06,
        "ONSET_HOP_LENGTH": 256,
        "ONSET_BACKTRACK": False,
        "ONSET_REFINE_ENABLED": True,
        "ONSET_REFINE_WINDOW_MS": 10,
        "ONSET_REFINE_ENERGY_GATE": 0.0,
        "ONSET_METHOD": "adaptive_hp",
    },
    "insect": {
        "description": "High-frequency percussive clicks (e.g. crickets, cicadas)",
        "HIGHPASS_CUTOFF_HZ": 1000,
        "ONSET_AMPLITUDE_GATE": 0.10,
        "ONSET_AMPLITUDE_WINDOW_MS": 20,
        "ONSET_SHARPNESS_GATE": 0.0,
        "ONSET_SHARPNESS_WINDOW_MS": 15,
        "MIN_INTER_ONSET_MS": 10,
        "ONSET_CLUSTER_WINDOW_MS": 10,
        "STABLE_RHYTHM_TOLERANCE": 0.20,
        "ONSET_DELTA": 0.04,
        "ONSET_HOP_LENGTH": 128,
        "ONSET_BACKTRACK": False,
        "ONSET_REFINE_ENABLED": True,
        "ONSET_REFINE_WINDOW_MS": 10,
        "ONSET_REFINE_ENERGY_GATE": 0.0,
        "ONSET_METHOD": "adaptive_hp",
    },
    "whale / marine mammal": {
        "description": "Whale song, dolphin clicks, and marine mammal calls",
        "HIGHPASS_CUTOFF_HZ": 50,
        "ONSET_AMPLITUDE_GATE": 0.04,
        "ONSET_AMPLITUDE_WINDOW_MS": 100,
        "ONSET_SHARPNESS_GATE": 0.0,
        "ONSET_SHARPNESS_WINDOW_MS": 40,
        "MIN_INTER_ONSET_MS": 80,
        "ONSET_CLUSTER_WINDOW_MS": 50,
        "STABLE_RHYTHM_TOLERANCE": 0.30,
        "ONSET_DELTA": 0.06,
        "ONSET_HOP_LENGTH": 512,
        "ONSET_BACKTRACK": False,
        "ONSET_REFINE_ENABLED": True,
        "ONSET_REFINE_WINDOW_MS": 20,
        "ONSET_REFINE_ENERGY_GATE": 0.0,
        "ONSET_METHOD": "adaptive_hp",
    },
    "amphibian": {
        "description": "Frog and toad calls — rhythmic, tonal, mid-frequency",
        "HIGHPASS_CUTOFF_HZ": 300,
        "ONSET_AMPLITUDE_GATE": 0.05,
        "ONSET_AMPLITUDE_WINDOW_MS": 40,
        "ONSET_SHARPNESS_GATE": 0.05,
        "ONSET_SHARPNESS_WINDOW_MS": 25,
        "MIN_INTER_ONSET_MS": 20,
        "ONSET_CLUSTER_WINDOW_MS": 20,
        "STABLE_RHYTHM_TOLERANCE": 0.25,
        "ONSET_DELTA": 0.06,
        "ONSET_HOP_LENGTH": 256,
        "ONSET_BACKTRACK": False,
        "ONSET_REFINE_ENABLED": True,
        "ONSET_REFINE_WINDOW_MS": 10,
        "ONSET_REFINE_ENERGY_GATE": 0.0,
        "ONSET_METHOD": "adaptive_hp",
    },
    "general / mixed": {
        "description": "Balanced defaults for unknown or varied recordings",
        "HIGHPASS_CUTOFF_HZ": 200,
        "ONSET_AMPLITUDE_GATE": 0.05,
        "ONSET_AMPLITUDE_WINDOW_MS": 50,
        "ONSET_SHARPNESS_GATE": 0.05,
        "ONSET_SHARPNESS_WINDOW_MS": 20,
        "MIN_INTER_ONSET_MS": 25,
        "ONSET_CLUSTER_WINDOW_MS": 20,
        "STABLE_RHYTHM_TOLERANCE": 0.25,
        "ONSET_DELTA": 0.06,
        "ONSET_HOP_LENGTH": 256,
        "ONSET_BACKTRACK": False,
        "ONSET_REFINE_ENABLED": True,
        "ONSET_REFINE_WINDOW_MS": 10,
        "ONSET_REFINE_ENERGY_GATE": 0.0,
        "ONSET_METHOD": "adaptive_hp",
    },
    "ensemble": {
        "description": "Group performance — multiple simultaneous sources (drumming circles, choruses)",
        "HIGHPASS_CUTOFF_HZ": 60,
        "ONSET_AMPLITUDE_GATE": 0.02,
        "ONSET_AMPLITUDE_WINDOW_MS": 30,
        "ONSET_SHARPNESS_GATE": 0.0,
        "ONSET_SHARPNESS_WINDOW_MS": 20,
        "MIN_INTER_ONSET_MS": 5,
        "ONSET_CLUSTER_WINDOW_MS": 0,
        "STABLE_RHYTHM_TOLERANCE": 0.35,
        "ONSET_DELTA": 0.04,
        "ONSET_HOP_LENGTH": 128,
        "ONSET_BACKTRACK": False,
        "ONSET_REFINE_ENABLED": True,
        "ONSET_REFINE_WINDOW_MS": 5,
        "ONSET_REFINE_ENERGY_GATE": 0.0,
        "ONSET_METHOD": "adaptive_hp",
    },
    # ----- Language / Speech presets -----
    "speech_syllable": {
        "description": "Human speech — syllable-level onsets via Praat intensity peaks (de Jong & Wempe 2009). "
                       "Requires parselmouth. Best for speech rhythm metrics (nPVI, %V, ΔC/ΔV).",
        "HIGHPASS_CUTOFF_HZ": 60,
        "ONSET_AMPLITUDE_GATE": 0.0,
        "ONSET_AMPLITUDE_WINDOW_MS": 50,
        "ONSET_SHARPNESS_GATE": 0.0,
        "ONSET_SHARPNESS_WINDOW_MS": 20,
        "ONSET_BROADBAND_MIN_BANDS": 0,
        "ONSET_BROADBAND_N_BANDS": 6,
        "ONSET_BROADBAND_THRESHOLD": 0.15,
        "MIN_INTER_ONSET_MS": 50,
        "ONSET_CLUSTER_WINDOW_MS": 30,
        "STABLE_RHYTHM_TOLERANCE": 0.30,
        "ONSET_DELTA": 0.05,
        "ONSET_HOP_LENGTH": 256,
        "ONSET_BACKTRACK": False,
        "ONSET_REFINE_ENABLED": False,
        "ONSET_REFINE_WINDOW_MS": 10,
        "ONSET_REFINE_ENERGY_GATE": 0.0,
        "ONSET_METHOD": "syllable_nuclei",
        "SYLLABLE_INTENSITY_THRESHOLD": -25.0,
        "SYLLABLE_MIN_DIP_DB": 2.0,
        "SYLLABLE_MIN_PAUSE_MS": 30.0,
        "SYLLABLE_VOICING_THRESHOLD": 0.3,
        "SYLLABLE_TIME_STEP": 0.01,
    },
    "speech_word": {
        "description": "Human speech — word-level onsets via Whisper automatic speech recognition. "
                       "Requires openai-whisper. Best for word-rate and pause analysis.",
        "HIGHPASS_CUTOFF_HZ": 60,
        "ONSET_AMPLITUDE_GATE": 0.0,
        "ONSET_AMPLITUDE_WINDOW_MS": 50,
        "ONSET_SHARPNESS_GATE": 0.0,
        "ONSET_SHARPNESS_WINDOW_MS": 20,
        "ONSET_BROADBAND_MIN_BANDS": 0,
        "ONSET_BROADBAND_N_BANDS": 6,
        "ONSET_BROADBAND_THRESHOLD": 0.15,
        "MIN_INTER_ONSET_MS": 80,
        "ONSET_CLUSTER_WINDOW_MS": 50,
        "STABLE_RHYTHM_TOLERANCE": 0.35,
        "ONSET_DELTA": 0.06,
        "ONSET_HOP_LENGTH": 256,
        "ONSET_BACKTRACK": False,
        "ONSET_REFINE_ENABLED": False,
        "ONSET_REFINE_WINDOW_MS": 10,
        "ONSET_REFINE_ENERGY_GATE": 0.0,
        "ONSET_METHOD": "whisper_words",
        "WHISPER_MODEL_SIZE": "base",
        "WHISPER_LANGUAGE": None,
        "WHISPER_WORD_TIMESTAMPS": True,
    },
    "speech_acoustic": {
        "description": "Human speech — acoustic onset detection using existing energy-based methods. "
                       "No extra dependencies needed. Detects syllable-level amplitude onsets.",
        "HIGHPASS_CUTOFF_HZ": 60,
        "ONSET_AMPLITUDE_GATE": 0.02,
        "ONSET_AMPLITUDE_WINDOW_MS": 40,
        "ONSET_SHARPNESS_GATE": 0.0,
        "ONSET_SHARPNESS_WINDOW_MS": 20,
        "ONSET_BROADBAND_MIN_BANDS": 0,
        "ONSET_BROADBAND_N_BANDS": 6,
        "ONSET_BROADBAND_THRESHOLD": 0.15,
        "MIN_INTER_ONSET_MS": 40,
        "ONSET_CLUSTER_WINDOW_MS": 20,
        "STABLE_RHYTHM_TOLERANCE": 0.30,
        "ONSET_DELTA": 0.04,
        "ONSET_HOP_LENGTH": 256,
        "ONSET_BACKTRACK": False,
        "ONSET_REFINE_ENABLED": True,
        "ONSET_REFINE_WINDOW_MS": 10,
        "ONSET_REFINE_ENERGY_GATE": 0.0,
        "ONSET_METHOD": "adaptive_hp",
        "HP_SMOOTH_LAMBDA": 30,
        "HP_THRESHOLD_LAMBDA": 1e7,
        "HP_ENVELOPE_WINDOW_MS": 10,
        "HP_ENVELOPE_HOP_MS": 1,
    },
    "speech_phoneme": {
        "description": "Human speech — phoneme-level alignment via WhisperX forced alignment. "
                       "Requires whisperx. Computes Ramus et al. (1999) metrics: %V, ΔV, ΔC, rPVI-C, nPVI-V.",
        "HIGHPASS_CUTOFF_HZ": 60,
        "ONSET_AMPLITUDE_GATE": 0.0,
        "ONSET_AMPLITUDE_WINDOW_MS": 50,
        "ONSET_SHARPNESS_GATE": 0.0,
        "ONSET_SHARPNESS_WINDOW_MS": 20,
        "ONSET_BROADBAND_MIN_BANDS": 0,
        "ONSET_BROADBAND_N_BANDS": 6,
        "ONSET_BROADBAND_THRESHOLD": 0.15,
        "MIN_INTER_ONSET_MS": 30,
        "ONSET_CLUSTER_WINDOW_MS": 20,
        "STABLE_RHYTHM_TOLERANCE": 0.30,
        "ONSET_DELTA": 0.05,
        "ONSET_HOP_LENGTH": 256,
        "ONSET_BACKTRACK": False,
        "ONSET_REFINE_ENABLED": False,
        "ONSET_REFINE_WINDOW_MS": 10,
        "ONSET_REFINE_ENERGY_GATE": 0.0,
        "ONSET_METHOD": "whisperx_phonemes",
        "WHISPERX_MODEL_SIZE": "base",
        "WHISPERX_LANGUAGE": None,
        "WHISPERX_DEVICE": "cpu",
    },
    # ----- Music / Song presets -----
    "music_beat_tracking": {
        "description": "Music / songs — deep-learning beat tracker (Böck et al. 2016) via madmom. "
                       "Finds the metrical pulse (beats & tempo) rather than individual acoustic onsets. "
                       "Works well across genres including non-Western music.",
        "HIGHPASS_CUTOFF_HZ": 30,
        "ONSET_AMPLITUDE_GATE": 0.0,
        "ONSET_AMPLITUDE_WINDOW_MS": 50,
        "ONSET_SHARPNESS_GATE": 0.0,
        "ONSET_SHARPNESS_WINDOW_MS": 20,
        "MIN_INTER_ONSET_MS": 200,
        "ONSET_CLUSTER_WINDOW_MS": 30,
        "STABLE_RHYTHM_TOLERANCE": 0.35,
        "ONSET_DELTA": 0.05,
        "ONSET_HOP_LENGTH": 256,
        "ONSET_BACKTRACK": False,
        "ONSET_REFINE_ENABLED": False,
        "ONSET_REFINE_WINDOW_MS": 10,
        "ONSET_REFINE_ENERGY_GATE": 0.0,
        "ONSET_METHOD": "madmom_beats",
        "MADMOM_MIN_BPM": 40,
        "MADMOM_MAX_BPM": 240,
        "MADMOM_DOWNBEATS": True,
        "MADMOM_FPS": 100,
        "MADMOM_TRANSITION_LAMBDA": 100,
        "PITCH_TRACKER": "pyin",
        "PITCH_FMIN": 65.0,
        "PITCH_FMAX": 1047.0,
        "TEMPO_ADAPTIVE_MIN_IOI": True,
        "TEMPO_ADAPTIVE_FRACTION": 0.5,
    },
}

# Set this to a preset name (e.g. "birdsong") or None to use manual values.
ACTIVE_PRESET = None

# These toggles control which review artifacts and filtered exports are created.
CREATE_SPECTROGRAMS = True
SPECTROGRAM_CHUNK_ENABLED = False
SPECTROGRAM_CHUNK_SECONDS = 30
CREATE_AUDACITY_LABELS = True
ADD_COLUMN_COMMENTS = True
ADD_FORMULA_SHEET = True
COLUMN_COMMENT_AUTHOR = "Bioacoustics Rhythm Pipeline"
CLUSTER_OVERLAPPING_ONSETS = True
ONSET_CLUSTER_WINDOW_MS = 25
FILTER_STABLE_RHYTHMS = True
STABLE_RHYTHM_TOLERANCE = 0.25

# --- Noise-Handling Pre-Filters (nextSteps.md item 2) ---
# NOTE: High-pass filtering and dB-based noise muting are now handled by
# audio_editor.py as a pre-processing step. These toggles are kept here as a
# fallback but are DISABLED by default to avoid double-filtering, which can
# over-aggressively remove quiet but legitimate animal calls.
#
# Only re-enable these if you are running the extractor on raw (un-muted) audio.
APPLY_HIGHPASS_FILTER = False
HIGHPASS_CUTOFF_HZ = 200

# Onset amplitude gate: discard onsets whose local RMS energy is below a fraction
# of the file's peak RMS. Set to a positive value to filter out weak onsets that
# may be residual noise (e.g. cicada remnants after spectral denoising).
# Recommended: 0.05-0.10 for field recordings, 0.0 to disable.
ONSET_AMPLITUDE_GATE = 0.05
ONSET_AMPLITUDE_WINDOW_MS = 50

# Onset sharpness gate: discard onsets whose attack slope (rate of energy rise)
# is below a fraction of the file's steepest attack. Distinguishes genuine sharp
# transients (e.g. primate drumming) from softer percussive sounds (branch snaps,
# footfalls) that have similar loudness but a less explosive attack.
# Recommended: 0.10-0.25 for primate drumming, 0.0 to disable.
ONSET_SHARPNESS_GATE = 0.0
ONSET_SHARPNESS_WINDOW_MS = 20

# Broadband gate: keep onsets only if significant energy appears across enough
# frequency bands simultaneously. Real percussion hits light up a vertical stripe
# in the spectrogram (many bands at once); noise artifacts tend to be narrowband.
# ONSET_BROADBAND_MIN_BANDS = minimum number of active bands (0 to disable).
# ONSET_BROADBAND_N_BANDS   = how many equal-width log bands to divide the spectrum into.
# ONSET_BROADBAND_THRESHOLD = fraction of per-band peak energy to consider "active".
# Recommended: min_bands=3, n_bands=6, threshold=0.15 for clean percussion.
ONSET_BROADBAND_MIN_BANDS = 0
ONSET_BROADBAND_N_BANDS = 6
ONSET_BROADBAND_THRESHOLD = 0.15

# Minimum inter-onset interval: any onset that follows the previous one by fewer
# than this many milliseconds is dropped (applied after clustering). Guards against
# impossibly fast detections caused by transient noise.
MIN_INTER_ONSET_MS = 30

# Onset detection sensitivity (passed to librosa as the `delta` parameter).
# Higher values = fewer, more prominent onsets detected. Lower = more sensitive.
# Recommended: 0.07 (general), 0.10-0.15 (noisy field recordings with cicadas).
ONSET_DELTA = 0.10

# Hop length for onset detection (samples). Controls temporal resolution:
#   512 → ~11.6 ms (librosa default — coarse, onsets can be off by up to ~6 ms)
#   256 → ~5.8 ms  (recommended — good balance of precision and speed)
#   128 → ~2.9 ms  (very fine — slower, may add noise-triggered false onsets)
# Reducing hop_length lets onsets snap to a finer time grid.
ONSET_HOP_LENGTH = 256

# Backtracking: when True, librosa rolls each detected onset *backward* to the
# nearest preceding local energy minimum. Intended to find the true attack start,
# but on noisy field recordings it often overshoots, placing onsets too early.
# Set False to keep onsets at the energy *peak* instead.
ONSET_BACKTRACK = False

# --- Sample-Level Onset Refinement ---
# When enabled, each coarse frame-level onset is refined to the nearest sample
# that shows the steepest energy rise (via the Hilbert amplitude envelope
# derivative) within a ±ONSET_REFINE_WINDOW_MS search window. This achieves
# sub-millisecond precision (~0.023 ms at 44.1 kHz) while keeping the coarse
# detector's robustness against cicada and other broadband noise.
#
# ONSET_REFINE_ENABLED  — True to activate sample-level refinement.
# ONSET_REFINE_WINDOW_MS — half-width of the search window around each coarse
#                          onset (±ms). 10 ms is a safe default; increase if the
#                          coarse detector systematically lands far from the true
#                          attack. Larger windows cost more CPU but stay fast.
# ONSET_REFINE_ENERGY_GATE — optional local energy ratio threshold (0.0–1.0).
#                            If the peak envelope in the refinement window is
#                            below this fraction of the file's peak envelope,
#                            the coarse time is kept unchanged (avoids snapping
#                            to noise). Set 0.0 to disable.
ONSET_REFINE_ENABLED = True
ONSET_REFINE_WINDOW_MS = 10
ONSET_REFINE_ENERGY_GATE = 0.0

# --- Onset Detection Method ---
# Controls which algorithm finds onset candidates before refinement/clustering.
# See onset_detectors.py for full documentation of each method.
#
# "adaptive_hp" (default) — HP adaptive baseline (Roeske et al. 2020).
# "librosa"      — Standard spectral-flux onset detection.
# "moving_median" — Moving-median adaptive baseline (simpler/faster than HP).
# "superflux"    — Spectral flux with vibrato suppression (Böck & Widmer 2013).
# "cfar"         — Constant False Alarm Rate detector.
# "per_band"     — Per-frequency-band adaptive threshold.
ONSET_METHOD = "adaptive_hp"

# HP filter parameters (used when ONSET_METHOD == "adaptive_hp"):
HP_SMOOTH_LAMBDA = 50
HP_THRESHOLD_LAMBDA = 5e7
HP_ENVELOPE_WINDOW_MS = 10
HP_ENVELOPE_HOP_MS = 1

# Moving-median parameters (used when ONSET_METHOD == "moving_median"):
MEDIAN_WINDOW_MS = 200
MEDIAN_THRESHOLD_SCALE = 1.5

# Superflux parameters (used when ONSET_METHOD == "superflux"):
SUPERFLUX_LAG = 2
SUPERFLUX_MAX_SIZE = 3

# CFAR parameters (used when ONSET_METHOD == "cfar"):
CFAR_GUARD_MS = 20
CFAR_TRAINING_MS = 200
CFAR_THRESHOLD_FACTOR = 4.0

# Per-band parameters (used when ONSET_METHOD == "per_band"):
PER_BAND_N_BANDS = 6
PER_BAND_FREQ_MIN = 200
PER_BAND_FREQ_MAX = None   # None = sr/2
PER_BAND_MEDIAN_MS = 200
PER_BAND_THRESHOLD_SCALE = 1.5
PER_BAND_MIN_BANDS = 2

# Syllable nuclei parameters (used when ONSET_METHOD == "syllable_nuclei"):
SYLLABLE_INTENSITY_THRESHOLD = -25.0
SYLLABLE_MIN_DIP_DB = 2.0
SYLLABLE_MIN_PAUSE_MS = 30.0
SYLLABLE_VOICING_THRESHOLD = 0.3
SYLLABLE_TIME_STEP = 0.01

# Whisper word-onset parameters (used when ONSET_METHOD == "whisper_words"):
WHISPER_MODEL_SIZE = "base"
WHISPER_LANGUAGE = None      # None = auto-detect
WHISPER_WORD_TIMESTAMPS = True

# Speech analysis parameters:
PAUSE_THRESHOLD_MS = 250.0          # IOIs longer than this = pauses (Dellwo 2006)
EXPORT_TEXTGRID = True              # Export Praat TextGrid files alongside labels
EXPORT_TRANSCRIPT = True            # Export Whisper transcript .txt when using whisper_words

# WhisperX phoneme alignment parameters (ONSET_METHOD == "whisperx_phonemes"):
WHISPERX_MODEL_SIZE = "base"
WHISPERX_LANGUAGE = None            # None = auto-detect (but forced alignment requires known language)
WHISPERX_DEVICE = "cpu"             # "cpu" or "cuda"

# madmom beat tracker parameters (ONSET_METHOD == "madmom_beats"):
MADMOM_MIN_BPM = 40
MADMOM_MAX_BPM = 240
MADMOM_FPS = 100
MADMOM_DOWNBEATS = False             # Also detect downbeats (bar-level "1"s)
MADMOM_TRANSITION_LAMBDA = 100       # Higher = stricter tempo continuity

# Standalone pitch tracker (runs after onset detection, any method):
PITCH_TRACKER = "none"               # "none", "pyin", "crepe", "praat"
PITCH_FMIN = 65.0                    # Min expected F0 in Hz (C2)
PITCH_FMAX = 1047.0                  # Max expected F0 in Hz (C6)

# Tempo-adaptive MIN_INTER_ONSET_MS:
TEMPO_ADAPTIVE_MIN_IOI = False       # Adapt MIN_INTER_ONSET_MS from detected tempo
TEMPO_ADAPTIVE_FRACTION = 0.5        # Use this fraction of beat interval as minimum

# --- Apply preset overrides ---
if ACTIVE_PRESET is not None:
    if ACTIVE_PRESET not in PRESETS:
        raise ValueError(f"Unknown preset '{ACTIVE_PRESET}'. Choose from: {list(PRESETS.keys())}")
    _preset = PRESETS[ACTIVE_PRESET]
    print(f"Using preset '{ACTIVE_PRESET}': {_preset['description']}")
    HIGHPASS_CUTOFF_HZ = _preset["HIGHPASS_CUTOFF_HZ"]
    ONSET_AMPLITUDE_GATE = _preset["ONSET_AMPLITUDE_GATE"]
    ONSET_AMPLITUDE_WINDOW_MS = _preset["ONSET_AMPLITUDE_WINDOW_MS"]
    ONSET_SHARPNESS_GATE = _preset["ONSET_SHARPNESS_GATE"]
    ONSET_SHARPNESS_WINDOW_MS = _preset["ONSET_SHARPNESS_WINDOW_MS"]
    ONSET_BROADBAND_MIN_BANDS = _preset.get("ONSET_BROADBAND_MIN_BANDS", ONSET_BROADBAND_MIN_BANDS)
    ONSET_BROADBAND_N_BANDS = _preset.get("ONSET_BROADBAND_N_BANDS", ONSET_BROADBAND_N_BANDS)
    ONSET_BROADBAND_THRESHOLD = _preset.get("ONSET_BROADBAND_THRESHOLD", ONSET_BROADBAND_THRESHOLD)
    MIN_INTER_ONSET_MS = _preset["MIN_INTER_ONSET_MS"]
    ONSET_CLUSTER_WINDOW_MS = _preset["ONSET_CLUSTER_WINDOW_MS"]
    STABLE_RHYTHM_TOLERANCE = _preset["STABLE_RHYTHM_TOLERANCE"]
    ONSET_DELTA = _preset["ONSET_DELTA"]
    ONSET_HOP_LENGTH = _preset["ONSET_HOP_LENGTH"]
    ONSET_BACKTRACK = _preset["ONSET_BACKTRACK"]
    ONSET_REFINE_ENABLED = _preset["ONSET_REFINE_ENABLED"]
    ONSET_REFINE_WINDOW_MS = _preset["ONSET_REFINE_WINDOW_MS"]
    ONSET_REFINE_ENERGY_GATE = _preset["ONSET_REFINE_ENERGY_GATE"]
    ONSET_METHOD = _preset["ONSET_METHOD"]
    # Speech-specific overrides (only present in speech presets)
    SYLLABLE_INTENSITY_THRESHOLD = _preset.get("SYLLABLE_INTENSITY_THRESHOLD", SYLLABLE_INTENSITY_THRESHOLD)
    SYLLABLE_MIN_DIP_DB = _preset.get("SYLLABLE_MIN_DIP_DB", SYLLABLE_MIN_DIP_DB)
    SYLLABLE_MIN_PAUSE_MS = _preset.get("SYLLABLE_MIN_PAUSE_MS", SYLLABLE_MIN_PAUSE_MS)
    SYLLABLE_VOICING_THRESHOLD = _preset.get("SYLLABLE_VOICING_THRESHOLD", SYLLABLE_VOICING_THRESHOLD)
    SYLLABLE_TIME_STEP = _preset.get("SYLLABLE_TIME_STEP", SYLLABLE_TIME_STEP)
    WHISPER_MODEL_SIZE = _preset.get("WHISPER_MODEL_SIZE", WHISPER_MODEL_SIZE)
    WHISPER_LANGUAGE = _preset.get("WHISPER_LANGUAGE", WHISPER_LANGUAGE)
    WHISPER_WORD_TIMESTAMPS = _preset.get("WHISPER_WORD_TIMESTAMPS", WHISPER_WORD_TIMESTAMPS)
    PAUSE_THRESHOLD_MS = _preset.get("PAUSE_THRESHOLD_MS", PAUSE_THRESHOLD_MS)
    EXPORT_TEXTGRID = _preset.get("EXPORT_TEXTGRID", EXPORT_TEXTGRID)
    EXPORT_TRANSCRIPT = _preset.get("EXPORT_TRANSCRIPT", EXPORT_TRANSCRIPT)
    WHISPERX_MODEL_SIZE = _preset.get("WHISPERX_MODEL_SIZE", WHISPERX_MODEL_SIZE)
    WHISPERX_LANGUAGE = _preset.get("WHISPERX_LANGUAGE", WHISPERX_LANGUAGE)
    WHISPERX_DEVICE = _preset.get("WHISPERX_DEVICE", WHISPERX_DEVICE)
    MADMOM_MIN_BPM = _preset.get("MADMOM_MIN_BPM", MADMOM_MIN_BPM)
    MADMOM_MAX_BPM = _preset.get("MADMOM_MAX_BPM", MADMOM_MAX_BPM)
    MADMOM_FPS = _preset.get("MADMOM_FPS", MADMOM_FPS)
    MADMOM_DOWNBEATS = _preset.get("MADMOM_DOWNBEATS", MADMOM_DOWNBEATS)
    MADMOM_TRANSITION_LAMBDA = _preset.get("MADMOM_TRANSITION_LAMBDA", MADMOM_TRANSITION_LAMBDA)
    PITCH_TRACKER = _preset.get("PITCH_TRACKER", PITCH_TRACKER)
    PITCH_FMIN = _preset.get("PITCH_FMIN", PITCH_FMIN)
    PITCH_FMAX = _preset.get("PITCH_FMAX", PITCH_FMAX)
    TEMPO_ADAPTIVE_MIN_IOI = _preset.get("TEMPO_ADAPTIVE_MIN_IOI", TEMPO_ADAPTIVE_MIN_IOI)
    TEMPO_ADAPTIVE_FRACTION = _preset.get("TEMPO_ADAPTIVE_FRACTION", TEMPO_ADAPTIVE_FRACTION)

# --- Default paths (relative to the project root, portable across machines) ---
_PROJECT_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
audio_folder = os.path.join(_PROJECT_DIR, "audioFiles_muted_clean")
output_excel_path = os.path.join(_PROJECT_DIR, "Cross_Species_Rhythm_Data.xlsx")
EXTRACTOR_SPECIFY_FILES = False
EXTRACTOR_SELECTED_FILES = []

# ------------------------------------------------------------------
# GUI CONFIG OVERRIDE — Read settings from pipeline_config.json if present.
# ------------------------------------------------------------------
_config_path = os.path.join(_PROJECT_DIR, "pipeline_config.json")
if os.path.isfile(_config_path):
    import json as _json
    with open(_config_path) as _f:
        _cfg = _json.load(_f).get("extractor", {})
    audio_folder = _cfg.get("audio_folder", audio_folder)
    output_excel_path = _cfg.get("output_excel_path", output_excel_path)
    EXTRACTOR_SPECIFY_FILES = _cfg.get("EXTRACTOR_SPECIFY_FILES", EXTRACTOR_SPECIFY_FILES)
    EXTRACTOR_SELECTED_FILES = _cfg.get("EXTRACTOR_SELECTED_FILES", EXTRACTOR_SELECTED_FILES)
    ACTIVE_PRESET = _cfg.get("ACTIVE_PRESET", ACTIVE_PRESET)
    CREATE_SPECTROGRAMS = _cfg.get("CREATE_SPECTROGRAMS", CREATE_SPECTROGRAMS)
    SPECTROGRAM_CHUNK_ENABLED = _cfg.get("SPECTROGRAM_CHUNK_ENABLED", SPECTROGRAM_CHUNK_ENABLED)
    SPECTROGRAM_CHUNK_SECONDS = _cfg.get("SPECTROGRAM_CHUNK_SECONDS", SPECTROGRAM_CHUNK_SECONDS)
    CREATE_AUDACITY_LABELS = _cfg.get("CREATE_AUDACITY_LABELS", CREATE_AUDACITY_LABELS)
    ADD_COLUMN_COMMENTS = _cfg.get("ADD_COLUMN_COMMENTS", True)
    ADD_FORMULA_SHEET = _cfg.get("ADD_FORMULA_SHEET", True)
    COLUMN_COMMENT_AUTHOR = _cfg.get("COLUMN_COMMENT_AUTHOR", COLUMN_COMMENT_AUTHOR)
    CLUSTER_OVERLAPPING_ONSETS = _cfg.get("CLUSTER_OVERLAPPING_ONSETS", CLUSTER_OVERLAPPING_ONSETS)
    ONSET_CLUSTER_WINDOW_MS = _cfg.get("ONSET_CLUSTER_WINDOW_MS", ONSET_CLUSTER_WINDOW_MS)
    FILTER_STABLE_RHYTHMS = _cfg.get("FILTER_STABLE_RHYTHMS", FILTER_STABLE_RHYTHMS)
    STABLE_RHYTHM_TOLERANCE = _cfg.get("STABLE_RHYTHM_TOLERANCE", STABLE_RHYTHM_TOLERANCE)
    APPLY_HIGHPASS_FILTER = _cfg.get("APPLY_HIGHPASS_FILTER", APPLY_HIGHPASS_FILTER)
    HIGHPASS_CUTOFF_HZ = _cfg.get("HIGHPASS_CUTOFF_HZ", HIGHPASS_CUTOFF_HZ)
    ONSET_AMPLITUDE_GATE = _cfg.get("ONSET_AMPLITUDE_GATE", ONSET_AMPLITUDE_GATE)
    ONSET_AMPLITUDE_WINDOW_MS = _cfg.get("ONSET_AMPLITUDE_WINDOW_MS", ONSET_AMPLITUDE_WINDOW_MS)
    ONSET_SHARPNESS_GATE = _cfg.get("ONSET_SHARPNESS_GATE", ONSET_SHARPNESS_GATE)
    ONSET_SHARPNESS_WINDOW_MS = _cfg.get("ONSET_SHARPNESS_WINDOW_MS", ONSET_SHARPNESS_WINDOW_MS)
    ONSET_BROADBAND_MIN_BANDS = _cfg.get("ONSET_BROADBAND_MIN_BANDS", ONSET_BROADBAND_MIN_BANDS)
    ONSET_BROADBAND_N_BANDS = _cfg.get("ONSET_BROADBAND_N_BANDS", ONSET_BROADBAND_N_BANDS)
    ONSET_BROADBAND_THRESHOLD = _cfg.get("ONSET_BROADBAND_THRESHOLD", ONSET_BROADBAND_THRESHOLD)
    MIN_INTER_ONSET_MS = _cfg.get("MIN_INTER_ONSET_MS", MIN_INTER_ONSET_MS)
    ONSET_METHOD = _cfg.get("ONSET_METHOD", ONSET_METHOD)
    ONSET_DELTA = _cfg.get("ONSET_DELTA", ONSET_DELTA)
    ONSET_HOP_LENGTH = _cfg.get("ONSET_HOP_LENGTH", ONSET_HOP_LENGTH)
    ONSET_BACKTRACK = _cfg.get("ONSET_BACKTRACK", ONSET_BACKTRACK)
    ONSET_REFINE_ENABLED = _cfg.get("ONSET_REFINE_ENABLED", ONSET_REFINE_ENABLED)
    ONSET_REFINE_WINDOW_MS = _cfg.get("ONSET_REFINE_WINDOW_MS", ONSET_REFINE_WINDOW_MS)
    ONSET_REFINE_ENERGY_GATE = _cfg.get("ONSET_REFINE_ENERGY_GATE", ONSET_REFINE_ENERGY_GATE)
    HP_SMOOTH_LAMBDA = _cfg.get("HP_SMOOTH_LAMBDA", HP_SMOOTH_LAMBDA)
    HP_THRESHOLD_LAMBDA = _cfg.get("HP_THRESHOLD_LAMBDA", HP_THRESHOLD_LAMBDA)
    HP_ENVELOPE_WINDOW_MS = _cfg.get("HP_ENVELOPE_WINDOW_MS", HP_ENVELOPE_WINDOW_MS)
    HP_ENVELOPE_HOP_MS = _cfg.get("HP_ENVELOPE_HOP_MS", HP_ENVELOPE_HOP_MS)
    MEDIAN_WINDOW_MS = _cfg.get("MEDIAN_WINDOW_MS", MEDIAN_WINDOW_MS)
    MEDIAN_THRESHOLD_SCALE = _cfg.get("MEDIAN_THRESHOLD_SCALE", MEDIAN_THRESHOLD_SCALE)
    SUPERFLUX_LAG = _cfg.get("SUPERFLUX_LAG", SUPERFLUX_LAG)
    SUPERFLUX_MAX_SIZE = _cfg.get("SUPERFLUX_MAX_SIZE", SUPERFLUX_MAX_SIZE)
    CFAR_GUARD_MS = _cfg.get("CFAR_GUARD_MS", CFAR_GUARD_MS)
    CFAR_TRAINING_MS = _cfg.get("CFAR_TRAINING_MS", CFAR_TRAINING_MS)
    CFAR_THRESHOLD_FACTOR = _cfg.get("CFAR_THRESHOLD_FACTOR", CFAR_THRESHOLD_FACTOR)
    PER_BAND_N_BANDS = _cfg.get("PER_BAND_N_BANDS", PER_BAND_N_BANDS)
    PER_BAND_FREQ_MIN = _cfg.get("PER_BAND_FREQ_MIN", PER_BAND_FREQ_MIN)
    PER_BAND_FREQ_MAX = _cfg.get("PER_BAND_FREQ_MAX", PER_BAND_FREQ_MAX)
    PER_BAND_MEDIAN_MS = _cfg.get("PER_BAND_MEDIAN_MS", PER_BAND_MEDIAN_MS)
    PER_BAND_THRESHOLD_SCALE = _cfg.get("PER_BAND_THRESHOLD_SCALE", PER_BAND_THRESHOLD_SCALE)
    PER_BAND_MIN_BANDS = _cfg.get("PER_BAND_MIN_BANDS", PER_BAND_MIN_BANDS)
    # Syllable nuclei parameters
    SYLLABLE_INTENSITY_THRESHOLD = float(_cfg.get("SYLLABLE_INTENSITY_THRESHOLD", SYLLABLE_INTENSITY_THRESHOLD))
    SYLLABLE_MIN_DIP_DB = float(_cfg.get("SYLLABLE_MIN_DIP_DB", SYLLABLE_MIN_DIP_DB))
    SYLLABLE_MIN_PAUSE_MS = float(_cfg.get("SYLLABLE_MIN_PAUSE_MS", SYLLABLE_MIN_PAUSE_MS))
    SYLLABLE_VOICING_THRESHOLD = float(_cfg.get("SYLLABLE_VOICING_THRESHOLD", SYLLABLE_VOICING_THRESHOLD))
    SYLLABLE_TIME_STEP = float(_cfg.get("SYLLABLE_TIME_STEP", SYLLABLE_TIME_STEP))
    # Whisper word-onset parameters
    WHISPER_MODEL_SIZE = _cfg.get("WHISPER_MODEL_SIZE", WHISPER_MODEL_SIZE)
    WHISPER_LANGUAGE = _cfg.get("WHISPER_LANGUAGE", WHISPER_LANGUAGE)
    WHISPER_WORD_TIMESTAMPS = bool(_cfg.get("WHISPER_WORD_TIMESTAMPS", WHISPER_WORD_TIMESTAMPS))
    # Speech analysis parameters
    PAUSE_THRESHOLD_MS = float(_cfg.get("PAUSE_THRESHOLD_MS", PAUSE_THRESHOLD_MS))
    EXPORT_TEXTGRID = bool(_cfg.get("EXPORT_TEXTGRID", EXPORT_TEXTGRID))
    EXPORT_TRANSCRIPT = bool(_cfg.get("EXPORT_TRANSCRIPT", EXPORT_TRANSCRIPT))
    # WhisperX phoneme alignment parameters
    WHISPERX_MODEL_SIZE = _cfg.get("WHISPERX_MODEL_SIZE", WHISPERX_MODEL_SIZE)
    WHISPERX_LANGUAGE = _cfg.get("WHISPERX_LANGUAGE", WHISPERX_LANGUAGE)
    WHISPERX_DEVICE = _cfg.get("WHISPERX_DEVICE", WHISPERX_DEVICE)
    # madmom beat tracker parameters
    MADMOM_MIN_BPM = _cfg.get("MADMOM_MIN_BPM", MADMOM_MIN_BPM)
    MADMOM_MAX_BPM = _cfg.get("MADMOM_MAX_BPM", MADMOM_MAX_BPM)
    MADMOM_FPS = _cfg.get("MADMOM_FPS", MADMOM_FPS)
    MADMOM_DOWNBEATS = bool(_cfg.get("MADMOM_DOWNBEATS", MADMOM_DOWNBEATS))
    MADMOM_TRANSITION_LAMBDA = _cfg.get("MADMOM_TRANSITION_LAMBDA", MADMOM_TRANSITION_LAMBDA)
    # Standalone pitch tracker
    PITCH_TRACKER = _cfg.get("PITCH_TRACKER", PITCH_TRACKER)
    PITCH_FMIN = float(_cfg.get("PITCH_FMIN", PITCH_FMIN))
    PITCH_FMAX = float(_cfg.get("PITCH_FMAX", PITCH_FMAX))
    # Tempo-adaptive MIN_INTER_ONSET_MS
    TEMPO_ADAPTIVE_MIN_IOI = bool(_cfg.get("TEMPO_ADAPTIVE_MIN_IOI", TEMPO_ADAPTIVE_MIN_IOI))
    TEMPO_ADAPTIVE_FRACTION = float(_cfg.get("TEMPO_ADAPTIVE_FRACTION", TEMPO_ADAPTIVE_FRACTION))
    # Per-file settings path (if the GUI exported per-file onset configs)
    _PER_FILE_SETTINGS_PATH = _cfg.get("per_file_settings_path", "")
    # Focus regions path (JSON exported by Onset Editor for time/freq masking)
    _FOCUS_REGIONS_PATH = _cfg.get("focus_regions_path", "")
    # Layer specification (from pipeline "Specify Layers" UI)
    EXTRACTOR_SPECIFY_LAYERS = _cfg.get("EXTRACTOR_SPECIFY_LAYERS", False)
    EXTRACTOR_SELECTED_LAYERS = _cfg.get("EXTRACTOR_SELECTED_LAYERS", [])
    # Spectral match threshold (0–1): minimum cosine similarity to the
    # positive-region spectral profile for an onset to be kept.
    SPECTRAL_MATCH_THRESHOLD = float(_cfg.get("spectral_match_threshold", 0.3))
    # Re-apply preset if config specifies one (GUI handles preset field population,
    # so individual values from config take precedence above).
    del _json, _f, _cfg
else:
    _PER_FILE_SETTINGS_PATH = ""
    _FOCUS_REGIONS_PATH = ""
    EXTRACTOR_SPECIFY_LAYERS = False
    EXTRACTOR_SELECTED_LAYERS = []
    SPECTRAL_MATCH_THRESHOLD = 0.3
del _PROJECT_DIR, _config_path


is_stable_match = _is_stable_match_impl
get_stable_dyad_flags = _get_stable_dyad_flags_impl
build_dyad_records = _build_dyad_records_impl
calculate_rhythm_metrics = _calculate_rhythm_metrics_impl
calculate_stable_subset_metrics = _calculate_stable_subset_metrics_impl
calculate_speech_rhythm_metrics = _calculate_speech_rhythm_metrics_impl


if __name__ == "__main__":

    # --- Validate audio folder before starting ---
    if not os.path.isdir(audio_folder):
        print(f"ERROR: Audio folder not found: {audio_folder}")
        print("Run the Audio Editor (Step 1) first, or update the audio_folder path.")
        sys.exit(1)

    # These lists accumulate rows before they are written to separate Excel sheets.
    file_summary_data = []
    dyadic_events_data = []
    stable_dyadic_events_data = []

    # ==========================================
    # 2. AUDIO PROCESSING LOOP
    # ==========================================
    print("Starting extraction process... This might take a moment depending on file sizes.")

    _VALID_EXTENSIONS = (".wav", ".mp3", ".flac", ".ogg")
    _selected_files = EXTRACTOR_SELECTED_FILES if EXTRACTOR_SPECIFY_FILES else []
    _audio_files, _missing_audio_files = resolve_audio_files(
        audio_folder, _selected_files, _VALID_EXTENSIONS)

    for _missing_name in _missing_audio_files:
        print(f"WARNING: Requested file not found or unsupported: {_missing_name}")

    if _selected_files:
        print(f"Selected files: {len(_audio_files)} of {len(_selected_files)} requested file(s)")

    if not _audio_files:
        print(f"WARNING: No audio files found in {audio_folder}")
        print(f"Supported formats: {', '.join(_VALID_EXTENSIONS)}")

    _batch_inputs = load_batch_routing_inputs(
        audio_folder,
        _audio_files,
        per_file_settings_path=_PER_FILE_SETTINGS_PATH,
        focus_regions_path=_FOCUS_REGIONS_PATH,
        specify_layers=EXTRACTOR_SPECIFY_LAYERS,
        selected_layers=EXTRACTOR_SELECTED_LAYERS,
    )
    _per_file_cfg = _batch_inputs.per_file_cfg
    _per_file_focus = _batch_inputs.per_file_focus
    _layer_configs = _batch_inputs.layer_configs

    if _per_file_cfg:
        print(f"Loaded per-file onset settings for {len(_per_file_cfg)} file(s)")

    # --- Per-file user-supplied metadata (Group, Species, Latitude, ...) ---
    # These are fields the experimental analyses (pDFA, Mantel, GLMM, PGLS)
    # need but the Onset Finder cannot compute. Filled in by the user via
    # the Metadata tab of the GUI's Per-File Settings dialog and stored in
    # the same JSON under each file's "metadata" section.
    _per_file_metadata, _general_metadata = load_user_metadata(_PER_FILE_SETTINGS_PATH)
    if _per_file_metadata:
        print(f"Loaded user-supplied metadata for "
              f"{len(_per_file_metadata)} file(s)")
    if _general_metadata:
        print(f"Loaded general metadata defaults: "
              f"{', '.join(_general_metadata.keys())}")

    if _per_file_focus:
        print(f"Loaded focus regions for {len(_per_file_focus)} file(s)")

    if _layer_configs:
        _total_layers = sum(len(v) for v in _layer_configs.values())
        print(f"Loaded onset layers: {_total_layers} layer(s) across "
              f"{len(_layer_configs)} file(s)")

    for filename in _audio_files:
        file_path = os.path.join(audio_folder, filename)
        print(f"Processing: {filename}...")

        with temporary_module_settings(
            sys.modules[__name__],
            _per_file_cfg.get(filename),
            ONSET_SETTING_NAMES,
        ):
            try:
                # Load the original sample rate so timing values stay faithful to the recording.
                y, sr = librosa.load(file_path, sr=None)

                # Apply high-pass filter to remove low-frequency rumble before onset detection.
                if APPLY_HIGHPASS_FILTER and HIGHPASS_CUTOFF_HZ > 0:
                    y = apply_highpass(y, sr, HIGHPASS_CUTOFF_HZ)

                total_duration = librosa.get_duration(y=y, sr=sr)

                # These coarse acoustic summaries provide quick descriptive context per file.
                tempo_bpm, _ = librosa.beat.beat_track(y=y, sr=sr)
                estimated_bpm = float(tempo_bpm[0]) if isinstance(tempo_bpm, np.ndarray) else float(tempo_bpm)

                rms = librosa.feature.rms(y=y)
                avg_loudness = np.mean(rms)

                centroid = librosa.feature.spectral_centroid(y=y, sr=sr)
                avg_brightness = np.mean(centroid)

                # Onset detection is the core dependency for all downstream rhythm metrics.
                _onset_method, _onset_kwargs = build_detector_call(
                    snapshot_module_settings(sys.modules[__name__], ONSET_SETTING_NAMES)
                )
                onset_times = detect_onsets(_onset_method, y, sr, **_onset_kwargs)

                # Standalone pitch tracker (runs independently of onset method)
                if PITCH_TRACKER and PITCH_TRACKER != "none":
                    try:
                        extract_pitch_metrics(
                            y, sr, method=PITCH_TRACKER,
                            fmin=PITCH_FMIN, fmax=PITCH_FMAX,
                        )
                    except Exception as _pitch_err:
                        print(f"  -> Pitch tracker ({PITCH_TRACKER}) failed: {_pitch_err}")

                # Sample-level refinement: snap each coarse onset to the steepest
                # energy rise in a ±ONSET_REFINE_WINDOW_MS neighbourhood, giving
                # ~0.023 ms precision at 44.1 kHz instead of the ~5.8 ms frame grid.
                if ONSET_REFINE_ENABLED and len(onset_times) > 0:
                    onset_times = refine_onsets_to_sample(
                        onset_times, y, sr,
                        window_ms=ONSET_REFINE_WINDOW_MS,
                        energy_gate=ONSET_REFINE_ENERGY_GATE,
                    )

                # Merge near-simultaneous onsets before the interval and dyad calculations.
                clustered_onset_times = (
                    cluster_onsets(onset_times, ONSET_CLUSTER_WINDOW_MS)
                    if CLUSTER_OVERLAPPING_ONSETS
                    else np.array(onset_times, dtype=float)
                )

                # Apply focus regions (positive = keep only inside, negative = exclude)
                if filename in _per_file_focus:
                    _file_regions = _per_file_focus[filename]
                    if isinstance(_file_regions, list) and _file_regions:
                        clustered_onset_times = apply_focus_regions(
                            clustered_onset_times, _file_regions
                        )
                        # Spectral similarity gate: keep onsets matching the
                        # positive-region spectral profile drawn by the user.
                        _profile = build_spectral_profile(y, sr, _file_regions)
                        if _profile is not None:
                            clustered_onset_times = gate_onsets_by_spectral_match(
                                clustered_onset_times, y, sr, _profile,
                                threshold=SPECTRAL_MATCH_THRESHOLD,
                            )

                # Amplitude gate: reject onsets in quiet regions that are likely noise artifacts.
                if ONSET_AMPLITUDE_GATE > 0:
                    clustered_onset_times = gate_onsets_by_amplitude(
                        clustered_onset_times, y, sr, ONSET_AMPLITUDE_GATE, ONSET_AMPLITUDE_WINDOW_MS
                    )

                # Sharpness gate: reject onsets with a weak attack slope.
                if ONSET_SHARPNESS_GATE > 0:
                    clustered_onset_times = gate_onsets_by_sharpness(
                        clustered_onset_times, y, sr, ONSET_SHARPNESS_GATE, ONSET_SHARPNESS_WINDOW_MS
                    )

                # Broadband gate: reject onsets that only activate a few frequency bands.
                if ONSET_BROADBAND_MIN_BANDS > 0:
                    clustered_onset_times = gate_onsets_by_broadband(
                        clustered_onset_times, y, sr,
                        ONSET_BROADBAND_MIN_BANDS,
                        ONSET_BROADBAND_N_BANDS,
                        ONSET_BROADBAND_THRESHOLD,
                    )

                # Minimum inter-onset spacing: drop impossibly fast repeated detections.
                # Tempo-adaptive override: compute MIN_INTER_ONSET_MS from BPM.
                _effective_min_ioi = MIN_INTER_ONSET_MS
                if TEMPO_ADAPTIVE_MIN_IOI:
                    _est_bpm = None
                    # Prefer madmom tempo if available
                    if last_madmom_tempo and last_madmom_tempo.get("BPM", 0) > 0:
                        _est_bpm = last_madmom_tempo["BPM"]
                    # Fallback: estimate from detected onsets (median IOI)
                    elif len(clustered_onset_times) >= 3:
                        _iois = np.diff(clustered_onset_times)
                        _med_ioi = float(np.median(_iois))
                        if _med_ioi > 0:
                            _est_bpm = 60.0 / _med_ioi
                    if _est_bpm and _est_bpm > 0:
                        _beat_ms = 60000.0 / _est_bpm
                        _effective_min_ioi = max(1, int(_beat_ms * TEMPO_ADAPTIVE_FRACTION))
                        print(f"  -> Tempo-adaptive min IOI: {_est_bpm:.1f} BPM "
                              f"× {TEMPO_ADAPTIVE_FRACTION:.0%} = {_effective_min_ioi} ms")
                if _effective_min_ioi > 0:
                    clustered_onset_times = enforce_min_interval(clustered_onset_times, _effective_min_ioi)

                # Per-layer onset filtering: when layers are specified, apply
                # each layer's focus regions to produce per-layer onset sets.
                _per_layer_onsets = {}
                if filename in _layer_configs:
                    _file_layer_cfgs = _layer_configs[filename]
                    for _lc in _file_layer_cfgs:
                        _lname = _lc.get("name", "?")
                        _lregs = _lc.get("focus_regions", [])
                        if _lregs:
                            _layer_times = apply_focus_regions(
                                clustered_onset_times, _lregs
                            )
                        else:
                            _layer_times = np.array(clustered_onset_times, dtype=float)
                        _per_layer_onsets[_lname] = _layer_times
                    # If layers produced results, combine all layer onsets as the
                    # final onset set (union, deduplicated, sorted).
                    if _per_layer_onsets:
                        _all_layer_times = np.concatenate(
                            list(_per_layer_onsets.values()))
                        _all_layer_times = np.unique(_all_layer_times)
                        clustered_onset_times = _all_layer_times

                intervals_seconds = np.diff(clustered_onset_times)
                intervals_ms = intervals_seconds * 1000
                stable_dyad_flags = get_stable_dyad_flags(intervals_ms, STABLE_RHYTHM_TOLERANCE)

                file_dyad_records = build_dyad_records(filename, intervals_ms, stable_dyad_flags)
                file_stable_dyad_records = [record for record in file_dyad_records if record["Stable Rhythm"]]

                dyadic_events_data.extend(file_dyad_records)
                if FILTER_STABLE_RHYTHMS:
                    stable_dyadic_events_data.extend(file_stable_dyad_records)

                metrics = calculate_rhythm_metrics(intervals_seconds, file_dyad_records)
                stable_metrics = calculate_stable_subset_metrics(file_stable_dyad_records)

                # Speech-specific rhythm metrics (computed for all methods but most
                # meaningful when using syllable_nuclei or whisper_words).
                _intervals_ms = [i * 1000.0 for i in intervals_seconds]
                speech_metrics = calculate_speech_rhythm_metrics(
                    _intervals_ms, pause_threshold_ms=PAUSE_THRESHOLD_MS)

                # --- AUDACITY LABEL EXPORT ---
                # These labels support the manual verification step recommended for noisy recordings.
                if CREATE_AUDACITY_LABELS and len(clustered_onset_times) > 0:
                    write_audacity_labels(
                        audio_folder,
                        filename,
                        clustered_onset_times,
                        refine_enabled=ONSET_REFINE_ENABLED,
                    )

                # --- WHISPER TRANSCRIPT EXPORT ---
                if EXPORT_TRANSCRIPT and ONSET_METHOD == "whisper_words" and last_whisper_transcript:
                    _transcript_export = write_whisper_transcript_exports(
                        audio_folder,
                        filename,
                        last_whisper_transcript,
                    )
                    print(f"  Transcript saved: {_transcript_export['txt_path']}")
                    _summary_row = file_summary_data[-1]
                    _summary_row["Whisper Transcript"] = _transcript_export["full_text"]

                # --- PRAAT TEXTGRID EXPORT ---
                if EXPORT_TEXTGRID and len(clustered_onset_times) > 0:
                    _transcript_entries = (
                        last_whisper_transcript
                        if ONSET_METHOD == "whisper_words" and EXPORT_TRANSCRIPT and last_whisper_transcript
                        else None
                    )
                    _tg_path = write_praat_textgrid_export(
                        audio_folder,
                        filename,
                        clustered_onset_times,
                        total_duration,
                        transcript_entries=_transcript_entries,
                    )
                    print(f"  TextGrid saved: {_tg_path}")

                # --- SPECTROGRAM VISUALIZER ---
                # The spectrogram overlay is intended for visual QA, not for downstream computation.
                if CREATE_SPECTROGRAMS and len(clustered_onset_times) > 0:
                    base_stem = os.path.splitext(filename)[0]
                    onset_arr = np.array(clustered_onset_times)

                    if SPECTROGRAM_CHUNK_ENABLED and SPECTROGRAM_CHUNK_SECONDS > 0:
                        chunk_dur = float(SPECTROGRAM_CHUNK_SECONDS)
                        n_chunks = max(1, int(np.ceil(total_duration / chunk_dur)))
                    else:
                        chunk_dur = total_duration
                        n_chunks = 1

                    for chunk_idx in range(n_chunks):
                        t_start = chunk_idx * chunk_dur
                        t_end = min(t_start + chunk_dur, total_duration)
                        s_start = int(t_start * sr)
                        s_end = int(t_end * sr)
                        y_chunk = y[s_start:s_end]

                        chunk_onsets = onset_arr[(onset_arr >= t_start) & (onset_arr < t_end)]

                        spectrogram = librosa.amplitude_to_db(np.abs(librosa.stft(y_chunk)), ref=np.max)
                        plt.figure(figsize=(10, 4))
                        librosa.display.specshow(spectrogram, x_axis="time", y_axis="log",
                                                 sr=sr, x_coords=np.linspace(t_start, t_end,
                                                 spectrogram.shape[1]))
                        plt.colorbar(format="%+2.0f dB")
                        if n_chunks == 1:
                            plt.title(f"Detected Onsets: {filename}")
                            image_filename = f"{base_stem}_plot.png"
                        else:
                            plt.title(f"Detected Onsets: {filename}  [{t_start:.1f}–{t_end:.1f} s]")
                            image_filename = f"{base_stem}_{chunk_idx + 1}.png"
                        plt.vlines(
                            chunk_onsets,
                            ymin=0,
                            ymax=sr / 2,
                            color="r",
                            alpha=0.8,
                            linestyle="--",
                            label="Onsets",
                        )
                        image_path = os.path.join(audio_folder, image_filename)
                        save_matplotlib_figure(plt, image_path)
                        plt.close()

                # --- SAVE FILE SUMMARY DATA ---
                # The summary sheet is the bridge between raw extraction and quick exploratory review.
                file_summary_data.append(
                    {
                        "File Name": filename,
                        "Total Duration (s)": round(total_duration, 3),
                        "Estimated Overall BPM": round(estimated_bpm, 1),
                        "Average Cycle Duration (ms)": round(metrics["Average Cycle Duration (ms)"], 2)
                        if metrics["Average Cycle Duration (ms)"] is not None
                        else "N/A",
                        "Avg Loudness (Energy)": round(avg_loudness, 4),
                        "Avg Brightness (Centroid Hz)": round(avg_brightness, 2),
                        "Total Onsets Found (Raw)": len(onset_times),
                        "Total Onsets Used": len(clustered_onset_times),
                        "Onsets Merged by Clustering": max(len(onset_times) - len(clustered_onset_times), 0),
                        "High-Pass Filter Hz": HIGHPASS_CUTOFF_HZ if APPLY_HIGHPASS_FILTER else "Off",
                        "Amplitude Gate Threshold": ONSET_AMPLITUDE_GATE if ONSET_AMPLITUDE_GATE > 0 else "Off",
                        "Sharpness Gate Threshold": ONSET_SHARPNESS_GATE if ONSET_SHARPNESS_GATE > 0 else "Off",
                        "Min Inter-Onset (ms)": MIN_INTER_ONSET_MS if MIN_INTER_ONSET_MS > 0 else "Off",
                        "Onset Delta": ONSET_DELTA if ONSET_METHOD in ("librosa", "superflux") else "N/A",
                        "Onset Method": ONSET_METHOD,
                        "Active Preset": ACTIVE_PRESET if ACTIVE_PRESET else "None",
                        "Onset Refinement": f"±{ONSET_REFINE_WINDOW_MS}ms (sample-level)" if ONSET_REFINE_ENABLED else "Off",
                        "Stable Dyads Retained": len(file_stable_dyad_records),
                        "Stable Rhythm Filter Enabled": FILTER_STABLE_RHYTHMS,
                        "nPVI (Isochrony)": round(metrics["nPVI (Isochrony)"], 2)
                        if metrics["nPVI (Isochrony)"] is not None
                        else "N/A",
                        "CV of Intervals": round(metrics["CV of Intervals"], 4)
                        if metrics["CV of Intervals"] is not None
                        else "N/A",
                        "r_k Std Dev": round(metrics["r_k Std Dev"], 4)
                        if metrics["r_k Std Dev"] is not None
                        else "N/A",
                        "r_k Entropy (Categorical Measure)": round(metrics["r_k Entropy (Categorical Measure)"], 4)
                        if metrics["r_k Entropy (Categorical Measure)"] is not None
                        else "N/A",
                        "Stable Rhythm nPVI": round(stable_metrics["Stable Rhythm nPVI"], 2)
                        if stable_metrics["Stable Rhythm nPVI"] is not None
                        else "N/A",
                        "Stable Rhythm CV": round(stable_metrics["Stable Rhythm CV"], 4)
                        if stable_metrics["Stable Rhythm CV"] is not None
                        else "N/A",
                        "Stable Rhythm r_k Std Dev": round(stable_metrics["Stable Rhythm r_k Std Dev"], 4)
                        if stable_metrics["Stable Rhythm r_k Std Dev"] is not None
                        else "N/A",
                        "Stable Rhythm Entropy": round(stable_metrics["Stable Rhythm Entropy"], 4)
                        if stable_metrics["Stable Rhythm Entropy"] is not None
                        else "N/A",
                        # Speech rhythm metrics (Grabe & Low 2002, Dellwo 2006)
                        "Speech Rate (onsets/sec)": speech_metrics["Speech Rate (onsets/sec)"]
                        if speech_metrics["Speech Rate (onsets/sec)"] is not None
                        else "N/A",
                        "Mean IOI (ms)": speech_metrics["Mean IOI (ms)"]
                        if speech_metrics["Mean IOI (ms)"] is not None
                        else "N/A",
                        "Std IOI (ms)": speech_metrics["Std IOI (ms)"]
                        if speech_metrics["Std IOI (ms)"] is not None
                        else "N/A",
                        "PVI-raw (ms)": speech_metrics["PVI-raw (ms)"]
                        if speech_metrics["PVI-raw (ms)"] is not None
                        else "N/A",
                        "nPVI-speech (%)": speech_metrics["nPVI (%)"]
                        if speech_metrics["nPVI (%)"] is not None
                        else "N/A",
                        "VarcoIOI (%)": speech_metrics["VarcoIOI (%)"]
                        if speech_metrics["VarcoIOI (%)"] is not None
                        else "N/A",
                        "Articulation Rate (onsets/sec)": speech_metrics["Articulation Rate (onsets/sec)"]
                        if speech_metrics["Articulation Rate (onsets/sec)"] is not None
                        else "N/A",
                        "Pause Count": speech_metrics["Pause Count"]
                        if speech_metrics["Pause Count"] is not None
                        else "N/A",
                        "Mean Pause Duration (ms)": speech_metrics["Mean Pause Duration (ms)"]
                        if speech_metrics["Mean Pause Duration (ms)"] is not None
                        else "N/A",
                        "Phonation Time Ratio": speech_metrics["Phonation Time Ratio"]
                        if speech_metrics["Phonation Time Ratio"] is not None
                        else "N/A",
                        "Exact Onset Times Used (s)": ", ".join(
                            [str(round(onset_time, 6 if ONSET_REFINE_ENABLED else 3)) for onset_time in clustered_onset_times]
                        ),
                    }
                )
                # Append per-layer onset columns when layers are active
                if _per_layer_onsets:
                    _prec = 6 if ONSET_REFINE_ENABLED else 3
                    _summary_row = file_summary_data[-1]
                    for _lname, _ltimes in _per_layer_onsets.items():
                        _col = f"Onset Times — {_lname} (s)"
                        _summary_row[_col] = ", ".join(
                            str(round(t, _prec)) for t in _ltimes)

                # F0/prosody metrics (populated by syllable_nuclei via side-channel)
                _summary_row = file_summary_data[-1]
                _f0_keys = [
                    "F0 Mean (Hz)", "F0 Std (Hz)", "F0 Min (Hz)",
                    "F0 Max (Hz)", "F0 Range (Hz)", "Jitter (local)",
                    "Intensity Mean (dB)", "Intensity Std (dB)",
                ]
                for _k in _f0_keys:
                    _summary_row[_k] = last_f0_metrics.get(_k, "N/A")
                if last_f0_metrics.get("Pitch Tracker"):
                    _summary_row["Pitch Tracker"] = last_f0_metrics["Pitch Tracker"]

                # Ramus et al. (1999) metrics (populated by whisperx_phonemes)
                _ramus_keys = [
                    "%V", "DeltaV (ms)", "DeltaC (ms)",
                    "rPVI-C (ms)", "nPVI-V (%)",
                ]
                for _k in _ramus_keys:
                    _summary_row[_k] = last_ramus_metrics.get(_k, "N/A")

                # madmom tempo data (populated by madmom_beats)
                if last_madmom_tempo:
                    _summary_row["Tempo BPM"] = last_madmom_tempo.get("BPM", "N/A")
                    _summary_row["Tempo BPM (alt)"] = last_madmom_tempo.get("BPM_alt", "N/A")
                    if len(last_madmom_downbeats) > 0:
                        _summary_row["N Downbeats"] = len(last_madmom_downbeats)

            except Exception as e:
                print(f"  -> ERROR processing {filename}: {e} — skipping this file")

    # ==========================================
    # 3. EXPORT EXCEL FILE WITH MULTIPLE TABS
    # ==========================================
    # Merge user-supplied metadata (Group, Species, Latitude, ...) into each
    # file-summary row. We NEVER overwrite a key that already exists in the
    # row — the Onset Finder's own measurements take priority over any
    # metadata value the user may have entered for the same key.
    # Per-file overrides win over general defaults for the same key.
    _all_meta_fields = merge_user_metadata_into_summary_rows(
        file_summary_data,
        _per_file_metadata,
        _general_metadata,
    )
    if _all_meta_fields:
        print(f"Merged user metadata columns into File Summaries: "
              f"{', '.join(_all_meta_fields)}")

    # --- Build a dedicated "File Demographics" sheet -----------------
    # Holds just File Name + every user-supplied metadata field (Group,
    # Species, Latitude, Longitude, Modality, Function, Tempo_BPM,
    # BodyMass_kg, plus any custom fields the user defined).  Downstream
    # analyses can merge this on File Name to pick up demographic info
    # without having to parse it out of File Summaries.
    _demographics_rows = build_file_demographics_rows(
        file_summary_data,
        _per_file_metadata,
        _general_metadata,
        demographic_fields=DEMOGRAPHIC_FIELDS,
    )

    # Keeping summaries and dyads on separate sheets makes the plotting scripts simpler.
    df_summary = pd.DataFrame(file_summary_data)
    df_dyads = pd.DataFrame(dyadic_events_data)
    df_stable_dyads = pd.DataFrame(stable_dyadic_events_data)
    df_demographics = (pd.DataFrame(_demographics_rows)
                       if _demographics_rows else None)

    # Ensure the output directory exists (handles any custom path from the GUI).
    _output_dir = os.path.dirname(output_excel_path)
    if _output_dir and not os.path.isdir(_output_dir):
        os.makedirs(_output_dir, exist_ok=True)

    workbook_sheets = {
        "File Summaries": df_summary,
        "Dyadic Events (For Plots)": df_dyads,
    }
    if FILTER_STABLE_RHYTHMS:
        workbook_sheets["Dyadic Events (Stable Rhythms)"] = df_stable_dyads
    if df_demographics is not None and not df_demographics.empty:
        workbook_sheets["File Demographics"] = df_demographics
        print(f"Wrote File Demographics sheet: "
              f"{len(df_demographics)} row(s), "
              f"{len(df_demographics.columns) - 1} field(s).")

    write_workbook_sheets(output_excel_path, workbook_sheets)

    # --- Add explanatory comments to column headers ---
    _SUMMARY_COMMENTS = {
        "File Name": "The name of the audio file that was analysed.",
        "Total Duration (s)": "Total length of the audio file in seconds.",
        "Estimated Overall BPM": "Estimated beats per minute, calculated from the median inter-onset interval across the entire file.",
        "Average Cycle Duration (ms)": "Mean duration of one full rhythmic cycle (two consecutive intervals combined) in milliseconds.",
        "Avg Loudness (Energy)": "Average RMS energy of the audio signal across all frames — a measure of overall volume/loudness.",
        "Avg Brightness (Centroid Hz)": "Average spectral centroid in Hz — higher values mean the sound energy is concentrated at higher frequencies (brighter timbre).",
        "Total Onsets Found (Raw)": "Number of onset events initially detected before any clustering or merging was applied.",
        "Total Onsets Used": "Number of onsets remaining after clustering and all gating filters. These are the onsets used for rhythm analysis.",
        "Onsets Merged by Clustering": "Number of raw onsets that were merged into neighbouring onsets because they fell within the cluster window.",
        "High-Pass Filter Hz": "High-pass filter cutoff frequency used during onset detection (if enabled). 'Off' means no high-pass was applied at this stage.",
        "Amplitude Gate Threshold": "Fraction of peak RMS energy an onset must exceed to be kept. 'Off' means no amplitude gating was applied.",
        "Sharpness Gate Threshold": "Fraction of the steepest attack slope an onset must exceed to be kept. 'Off' means no sharpness gating was applied.",
        "Min Inter-Onset (ms)": "Minimum allowed gap in milliseconds between consecutive onsets. Closer onsets are removed. 'Off' means no minimum was enforced.",
        "Onset Delta": "Peak-picking sensitivity threshold used by the librosa/superflux onset methods. 'N/A' for methods that don't use this parameter.",
        "Onset Method": "Which onset detection algorithm was used: adaptive_hp, librosa, moving_median, superflux, cfar, or per_band.",
        "Active Preset": "Name of the detection preset applied (e.g. 'birdsong', 'percussion_clean'), or 'None' if parameters were set manually.",
        "Extractor Engine": "Which extraction engine was used: 'standard' (full-featured) or 'thebeat' (thebeat package).",
        "Onset Refinement": "Whether sample-level Hilbert-envelope refinement was applied to improve onset timing precision, and the search window used.",
        "Stable Dyads Retained": "Number of dyadic rhythm events that passed the stable-rhythm consistency filter for this file.",
        "Stable Rhythm Filter Enabled": "Whether the stable-rhythm filter was turned on (True) or off (False) for this analysis run.",
        "nPVI (Isochrony)": "Normalised Pairwise Variability Index — measures how much consecutive intervals differ. 0 = perfectly isochronous (metronomic), higher = more variable timing.",
        "CV of Intervals": "Coefficient of Variation of inter-onset intervals (standard deviation / mean). Lower values indicate more regular timing.",
        "r_k Std Dev": "Standard deviation of the rhythm ratio (r_k) values across all dyads. Lower values mean more consistent rhythm ratios.",
        "r_k Entropy (Categorical Measure)": "Shannon entropy of the rhythm ratio distribution. Higher values indicate a wider spread of rhythm types; lower values indicate clustering around specific ratios.",
        "Stable Rhythm nPVI": "nPVI computed using only the stable-rhythm dyads (those passing the consistency filter). 'N/A' if no stable dyads exist.",
        "Stable Rhythm CV": "CV of intervals computed using only the stable-rhythm subset. 'N/A' if no stable dyads exist.",
        "Stable Rhythm r_k Std Dev": "Standard deviation of r_k for stable dyads only. 'N/A' if no stable dyads exist.",
        "Stable Rhythm Entropy": "Shannon entropy of r_k for stable dyads only. 'N/A' if no stable dyads exist.",
        "Exact Onset Times Used (s)": "Comma-separated list of every onset time (in seconds from file start) that was used for rhythm analysis, after all filtering and clustering.",
    }

    _DYAD_COMMENTS = {
        "File Name": "The audio file this dyadic event came from.",
        "Dyad Index": "Sequential index of this dyad within the file (1 = first pair of consecutive intervals).",
        "Interval 1 (ms)": "Duration in milliseconds of the first interval in this consecutive pair.",
        "Interval 2 (ms)": "Duration in milliseconds of the second interval in this consecutive pair.",
        "Cycle Duration [cd] (ms)": "Sum of both intervals (Interval 1 + Interval 2) — the total duration of one rhythmic cycle in milliseconds.",
        "Short Interval [i_s] (ms)": "The shorter of the two intervals in this dyad, in milliseconds.",
        "Long Interval [i_l] (ms)": "The longer of the two intervals in this dyad, in milliseconds.",
        "Rhythm Ratio [r_k]": "Interval 1 divided by the Cycle Duration (i\u2081 / cd). Ranges from 0 to 1. A value of 0.5 means both intervals are equal (isochronous). Values below 0.5 mean the first interval is shorter; above 0.5 means it is longer.",
        "Stable Rhythm": "Whether this dyad passed the stable-rhythm consistency filter (True/False). True means this rhythm pattern was repeated consistently across neighbouring cycles.",
    }

    # --- Add explanatory comments to column headers ---
    if ADD_COLUMN_COMMENTS:
        from openpyxl import load_workbook
        from openpyxl.comments import Comment
        wb = load_workbook(output_excel_path)
        for sheet_name, col_comments in [
            ("File Summaries", _SUMMARY_COMMENTS),
            ("Dyadic Events (For Plots)", _DYAD_COMMENTS),
            ("Dyadic Events (Stable Rhythms)", _DYAD_COMMENTS),
        ]:
            if sheet_name not in wb.sheetnames:
                continue
            ws = wb[sheet_name]
            for col_idx, cell in enumerate(ws[1], start=1):
                header_text = cell.value
                if header_text and header_text in col_comments:
                    cell.comment = Comment(col_comments[header_text], COLUMN_COMMENT_AUTHOR)
        save_openpyxl_workbook(wb, output_excel_path)
        print("Column header comments added to Excel file.")

    # --- Add "Formulas Used" reference sheet ---
    if ADD_FORMULA_SHEET:
        from openpyxl import load_workbook as _load_wb
        from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

        # Each entry: (Column Name, Python code, Excel equivalent)
        # Order matches the exact column order in the respective sheets.
        _SUMMARY_FORMULAS = [
            (
                "File Name",
                "filename  # The audio file name (string, no formula)",
                '(Text value — no formula.  e.g. "drum01.wav")',
            ),
            (
                "Total Duration (s)",
                "round(librosa.get_duration(y=y, sr=sr), 3)",
                "ROUND(total_samples / sample_rate, 3)",
            ),
            (
                "Estimated Overall BPM",
                "tempo_bpm, _ = librosa.beat.beat_track(y=y, sr=sr)\n"
                "round(float(tempo_bpm[0] if isinstance(tempo_bpm, np.ndarray) else tempo_bpm), 1)",
                "ROUND(60 / AVERAGE(inter_onset_intervals_s), 1)\n"
                "(Note: librosa uses autocorrelation, not a simple average.)",
            ),
            (
                "Average Cycle Duration (ms)",
                "cycle_durations = [d['Cycle Duration [cd] (ms)'] for d in dyad_records]\n"
                "float(np.mean(cycle_durations))",
                "AVERAGE('Dyadic Events (For Plots)'!E:E)\n"
                "(Average of Cycle Duration [cd] (ms) for this file.)",
            ),
            (
                "Avg Loudness (Energy)",
                "rms = librosa.feature.rms(y=y)\n"
                "round(np.mean(rms), 4)",
                "ROUND(AVERAGE(frame_rms_values), 4)\n"
                "(RMS energy per frame, averaged. Requires raw audio.)",
            ),
            (
                "Avg Brightness (Centroid Hz)",
                "centroid = librosa.feature.spectral_centroid(y=y, sr=sr)\n"
                "round(np.mean(centroid), 2)",
                "ROUND(AVERAGE(frame_centroid_values), 2)\n"
                "(Spectral centroid per frame, averaged. Requires raw audio.)",
            ),
            (
                "Total Onsets Found (Raw)",
                "len(onset_times)  # Count before clustering/gating",
                "(Integer count — no formula.)",
            ),
            (
                "Total Onsets Used",
                "len(clustered_onset_times)  # Count after all gates & clustering",
                "(Integer count — no formula.)",
            ),
            (
                "Onsets Merged by Clustering",
                "max(len(onset_times) - len(clustered_onset_times), 0)",
                "MAX([Total Onsets Found (Raw)] - [Total Onsets Used], 0)",
            ),
            (
                "High-Pass Filter Hz",
                'HIGHPASS_CUTOFF_HZ if APPLY_HIGHPASS_FILTER else "Off"',
                '(Config value. "Off" if disabled.)',
            ),
            (
                "Amplitude Gate Threshold",
                'ONSET_AMPLITUDE_GATE if ONSET_AMPLITUDE_GATE > 0 else "Off"',
                '(Config value. "Off" if disabled.)',
            ),
            (
                "Sharpness Gate Threshold",
                'ONSET_SHARPNESS_GATE if ONSET_SHARPNESS_GATE > 0 else "Off"',
                '(Config value. "Off" if disabled.)',
            ),
            (
                "Min Inter-Onset (ms)",
                'MIN_INTER_ONSET_MS if MIN_INTER_ONSET_MS > 0 else "Off"',
                '(Config value. "Off" if disabled.)',
            ),
            (
                "Onset Delta",
                'ONSET_DELTA if ONSET_METHOD in ("librosa", "superflux") else "N/A"',
                "(Config value — librosa/superflux only.)",
            ),
            (
                "Onset Method",
                "ONSET_METHOD  # e.g. 'adaptive_hp', 'librosa', etc.",
                "(Config value — no formula.)",
            ),
            (
                "Active Preset",
                'ACTIVE_PRESET if ACTIVE_PRESET else "None"',
                "(Config value — no formula.)",
            ),
            (
                "Onset Refinement",
                'f"±{ONSET_REFINE_WINDOW_MS}ms (sample-level)" if ONSET_REFINE_ENABLED else "Off"',
                "(Config value. Hilbert-envelope refinement window.)",
            ),
            (
                "Stable Dyads Retained",
                "len([d for d in dyad_records if d['Stable Rhythm'] == True])",
                "COUNTIF('Dyadic Events (For Plots)'!I:I, TRUE)\n"
                '(Count rows where Stable Rhythm = TRUE for this file.)',
            ),
            (
                "Stable Rhythm Filter Enabled",
                "FILTER_STABLE_RHYTHMS  # Boolean config flag",
                "(Config value — TRUE or FALSE.)",
            ),
            (
                "nPVI (Isochrony)",
                "rk_values = [d['Rhythm Ratio [r_k]'] for d in dyad_records]\n"
                "rk_count = len(rk_values)\n"
                "npvi = (400 / rk_count) * np.sum(np.abs(rk_values - 0.5))",
                "( 400 / COUNTA(rk_range) ) * SUMPRODUCT( ABS(rk_range - 0.5) )\n"
                "rk_range = 'Dyadic Events (For Plots)'!H:H for this file.",
            ),
            (
                "CV of Intervals",
                "intervals_seconds = np.diff(sorted_onset_times)\n"
                "mean_interval = np.mean(intervals_seconds)\n"
                "cv = np.std(intervals_seconds) / mean_interval",
                "STDEV(interval_range) / AVERAGE(interval_range)\n"
                "interval_range = consecutive differences of onset times.",
            ),
            (
                "r_k Std Dev",
                "rk_values = [d['Rhythm Ratio [r_k]'] for d in dyad_records]\n"
                "rk_std_dev = float(np.std(rk_values))",
                "STDEV.P(rk_range)\n"
                "rk_range = 'Dyadic Events (For Plots)'!H:H for this file.\n"
                "(Python np.std = population std dev = STDEV.P in Excel.)",
            ),
            (
                "r_k Entropy (Categorical Measure)",
                "counts, _ = np.histogram(rk_values, bins=10, range=(0, 1))\n"
                "rk_entropy = float(scipy.stats.entropy(counts))\n"
                "# Shannon entropy with natural log (base e)",
                "No single-cell Excel formula.\n"
                "(1) 10-bin histogram of r_k in [0,1).\n"
                "(2) p_i = count_i / total.\n"
                "(3) Entropy = -SUM(p_i * LN(p_i)) where p_i > 0.",
            ),
            (
                "Stable Rhythm nPVI",
                "stable_rk = [d['Rhythm Ratio [r_k]'] for d in stable_dyads]\n"
                "s_npvi = (400 / len(stable_rk)) * np.sum(np.abs(stable_rk - 0.5))",
                "( 400 / COUNTA(stable_rk) ) * SUMPRODUCT( ABS(stable_rk - 0.5) )\n"
                "stable_rk = r_k from Stable Rhythms sheet for this file.",
            ),
            (
                "Stable Rhythm CV",
                "stable_ints = [d['Interval 1 (ms)'], d['Interval 2 (ms)'] for d in stable_dyads]\n"
                "s_cv = np.std(stable_ints) / np.mean(stable_ints)",
                "STDEV(stable_intervals) / AVERAGE(stable_intervals)\n"
                "stable_intervals = Intervals 1 & 2 from stable dyads.",
            ),
            (
                "Stable Rhythm r_k Std Dev",
                "stable_rk = [d['Rhythm Ratio [r_k]'] for d in stable_dyads]\n"
                "s_std = float(np.std(stable_rk))",
                "STDEV.P(stable_rk)\n"
                "stable_rk = r_k from Stable Rhythms sheet for this file.",
            ),
            (
                "Stable Rhythm Entropy",
                "counts, _ = np.histogram(stable_rk, bins=10, range=(0, 1))\n"
                "s_entropy = float(scipy.stats.entropy(counts))",
                "No single-cell Excel formula.\n"
                "Same 10-bin histogram entropy as r_k Entropy,\n"
                "but using only stable dyad r_k values.",
            ),
            (
                "Exact Onset Times Used (s)",
                '", ".join([str(round(t, 6 if ONSET_REFINE_ENABLED else 3))\n'
                "            for t in clustered_onset_times])",
                "(Comma-separated text list — no formula.\n"
                "Onset times in seconds from file start.)",
            ),
        ]

        _DYAD_FORMULAS = [
            (
                "File Name",
                "filename  # Inherited from the parent audio file",
                "(Text value — same as File Summaries sheet.)",
            ),
            (
                "Dyad Index",
                "index + 1  # 1-based sequential counter per file",
                "ROW() - 1  (approximate — sequential per file)",
            ),
            (
                "Interval 1 (ms)",
                "intervals_ms[index]  # (onset[i+1] - onset[i]) * 1000",
                "(onset_time_{i+1} - onset_time_{i}) * 1000",
            ),
            (
                "Interval 2 (ms)",
                "intervals_ms[index + 1]  # (onset[i+2] - onset[i+1]) * 1000",
                "(onset_time_{i+2} - onset_time_{i+1}) * 1000",
            ),
            (
                "Cycle Duration [cd] (ms)",
                "round(interval_1 + interval_2, 2)",
                "[Interval 1 (ms)] + [Interval 2 (ms)]",
            ),
            (
                "Short Interval [i_s] (ms)",
                "round(min(interval_1, interval_2), 2)",
                "MIN([Interval 1 (ms)], [Interval 2 (ms)])",
            ),
            (
                "Long Interval [i_l] (ms)",
                "round(max(interval_1, interval_2), 2)",
                "MAX([Interval 1 (ms)], [Interval 2 (ms)])",
            ),
            (
                "Rhythm Ratio [r_k]",
                "round(interval_1 / cycle_duration, 4)\n"
                "# Range [0, 1].  0.5 = isochronous.",
                "[Interval 1 (ms)] / [Cycle Duration (ms)]",
            ),
            (
                "Stable Rhythm",
                "stable_flags[index]\n"
                "# True if both:\n"
                "#   |i1a-i1b| / max(i1a,i1b) <= tolerance\n"
                "#   |i2a-i2b| / max(i2a,i2b) <= tolerance\n"
                "# a = this dyad, b = next dyad",
                "TRUE if ABS(this_i1 - next_i1)/MAX(this_i1, next_i1) <= tol\n"
                "AND ABS(this_i2 - next_i2)/MAX(this_i2, next_i2) <= tol\n"
                "(Compares this dyad's intervals with the next.)",
            ),
        ]

        wb = _load_wb(output_excel_path)

        # ── Helper: write a transposed formulas section ──
        # Layout mirrors the original sheet: columns across, with rows for
        # the Python formula and Excel equivalent stacked below each header.
        _header_font = Font(bold=True, size=11)
        _header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2",
                                   fill_type="solid")
        _label_font = Font(bold=True, size=10, color="444444")
        _wrap = Alignment(wrap_text=True, vertical="top")
        _thin_border = Border(
            bottom=Side(style="thin", color="AAAAAA"))

        def _write_formula_section(ws, start_row, title, formulas, comments):
            """Write one transposed section starting at start_row.

            Returns the next free row number.
            """
            # Section title row
            ws.cell(row=start_row, column=1, value=title).font = Font(
                bold=True, size=12)
            start_row += 1

            # Row labels in column A
            ws.cell(row=start_row, column=1,
                    value="Column Name").font = _label_font
            ws.cell(row=start_row + 1, column=1,
                    value="Description").font = _label_font
            ws.cell(row=start_row + 2, column=1,
                    value="Python Code").font = _label_font
            ws.cell(row=start_row + 3, column=1,
                    value="Excel Formula").font = _label_font

            for col_idx, (col_name, py_code, xl_code) in enumerate(formulas,
                                                                     start=2):
                # Header cell (row 1 of section)
                hdr = ws.cell(row=start_row, column=col_idx, value=col_name)
                hdr.font = _header_font
                hdr.fill = _header_fill
                hdr.alignment = _wrap
                hdr.border = _thin_border

                # Description (row 2) — from the comments dict
                desc_text = comments.get(col_name, "")
                desc_cell = ws.cell(row=start_row + 1, column=col_idx,
                                    value=desc_text)
                desc_cell.alignment = _wrap

                # Python formula (row 3)
                py_cell = ws.cell(row=start_row + 2, column=col_idx,
                                  value=py_code)
                py_cell.alignment = _wrap

                # Excel formula (row 4) — prefix with apostrophe character
                # to prevent Excel from interpreting it as a live formula
                safe_xl = xl_code
                if safe_xl.lstrip().startswith("="):
                    safe_xl = "'" + safe_xl  # force text interpretation
                xl_cell = ws.cell(row=start_row + 3, column=col_idx,
                                  value=safe_xl)
                xl_cell.alignment = _wrap

            return start_row + 5  # 4 data rows + 1 blank separator

        ws = wb.create_sheet("Formulas Used")

        next_row = _write_formula_section(
            ws, 1, "File Summaries — Formulas Used",
            _SUMMARY_FORMULAS, _SUMMARY_COMMENTS)
        _write_formula_section(
            ws, next_row, "Dyadic Events — Formulas Used",
            _DYAD_FORMULAS, _DYAD_COMMENTS)

        # Auto-fit column widths (cap at 45 to keep it readable)
        for col_cells in ws.columns:
            max_len = 0
            col_letter = col_cells[0].column_letter
            for cell in col_cells:
                if cell.value:
                    lines = str(cell.value).split("\n")
                    max_len = max(max_len, max(len(line) for line in lines))
            ws.column_dimensions[col_letter].width = min(max_len + 3, 45)
        # Column A (row labels) can be narrower
        ws.column_dimensions["A"].width = 16

        save_openpyxl_workbook(wb, output_excel_path)
        print("Formulas reference sheet added to Excel file.")

    print(f"\nSUCCESS! Processed {len(file_summary_data)} files.")
    print(f"Extracted {len(dyadic_events_data)} total dyadic rhythms.")
    if FILTER_STABLE_RHYTHMS:
        print(f"Retained {len(stable_dyadic_events_data)} dyadic rhythms in the stable-rhythm sheet.")
    print(f"Spreadsheet saved to: {output_excel_path}")