"""Extract rhythmic timing features using thebeat as the analysis backend.

This is an alternative to ``onset_finder.py``. Instead of computing IOIs
and rhythm metrics manually, it feeds onset times into the ``thebeat``
package (an object-oriented sequence handler for bioacoustics) to derive
intervals and advanced metrics like nPVI and Entropy.

This script now supports the full feature set of the standard extractor:
- All 6 onset detection methods (via onset_detectors.py)
- Sample-level onset refinement (Hilbert envelope)
- Onset clustering, amplitude gating, min inter-onset enforcement
- High-pass pre-filtering
- Stable rhythm filtering (3-sheet Excel output)
- Audacity label export
- Spectrogram export with onset markers
- Named presets (birdsong, percussion, primate, insect)
- GUI JSON config override

Current role in the pipeline (when selected via main.py):
- Input:  muted audio in ``audioFiles_muted_clean/``
- Output: ``Cross_Species_Rhythm_Data.xlsx`` (same workbook format as the
          standard extractor so the downstream visualisers work unchanged)
"""

import os  # directory listing and file path joining
import sys  # exit with status code on fatal errors

import librosa  # onset detection and audio loading
import librosa.display  # spectrogram plotting utilities for QA images
import matplotlib.pyplot as plt  # generating plots for visual outputs
import numpy as np  # numerical array operations for interval math
import pandas as pd  # dataframe construction and Excel export
from scipy.signal import butter, sosfilt  # high-pass filter implementation
from scipy.stats import entropy  # Shannon entropy for rhythm ratio distribution
from thebeat import Sequence  # object-oriented IOI / rhythm sequence handler

from onset_detectors import (detect_onsets, refine_onsets_to_sample, available_methods,
                             last_f0_metrics, last_madmom_tempo,
                             extract_pitch_metrics)
try:
    from .onset_exports import write_audacity_labels
    from .shared_output_writers import (
        save_matplotlib_figure,
        save_openpyxl_workbook,
        write_workbook_sheets,
    )
except ImportError:
    from onset_exports import write_audacity_labels
    from shared_output_writers import (
        save_matplotlib_figure,
        save_openpyxl_workbook,
        write_workbook_sheets,
    )

# ==========================================
# 1. CONFIGURATION
# ==========================================
# --- Named Presets ---
PRESETS = {
    "birdsong": {
        "description": "Small passerines with rapid, high-frequency calls",
        "HIGHPASS_CUTOFF_HZ": 500,
        "ONSET_AMPLITUDE_GATE": 0.08,
        "ONSET_AMPLITUDE_WINDOW_MS": 30,
        "ONSET_SHARPNESS_GATE": 0.0,
        "ONSET_SHARPNESS_WINDOW_MS": 20,
        "MIN_INTER_ONSET_MS": 12,
        "ONSET_CLUSTER_WINDOW_MS": 15,
        "STABLE_RHYTHM_TOLERANCE": 0.25,
        "ONSET_DELTA": 0.05,
        "ONSET_HOP_LENGTH": 256,
        "ONSET_BACKTRACK": False,
        "ONSET_REFINE_ENABLED": True,
        "ONSET_REFINE_WINDOW_MS": 10,
        "ONSET_REFINE_ENERGY_GATE": 0.0,
        "ONSET_METHOD": "adaptive_hp",
    },
    "percussion_clean": {
        "description": "Clean percussion recordings — sensitive to quiet hits and rapid sequences",
        "HIGHPASS_CUTOFF_HZ": 80,
        "ONSET_AMPLITUDE_GATE": 0.005,
        "ONSET_AMPLITUDE_WINDOW_MS": 50,
        "ONSET_SHARPNESS_GATE": 0.0,
        "ONSET_SHARPNESS_WINDOW_MS": 20,
        "ONSET_BROADBAND_MIN_BANDS": 3,
        "ONSET_BROADBAND_N_BANDS": 6,
        "ONSET_BROADBAND_THRESHOLD": 0.15,
        "MIN_INTER_ONSET_MS": 15,
        "ONSET_CLUSTER_WINDOW_MS": 12,
        "STABLE_RHYTHM_TOLERANCE": 0.25,
        "ONSET_DELTA": 0.04,
        "ONSET_HOP_LENGTH": 128,
        "ONSET_BACKTRACK": False,
        "ONSET_REFINE_ENABLED": True,
        "ONSET_REFINE_WINDOW_MS": 8,
        "ONSET_REFINE_ENERGY_GATE": 0.005,
        "ONSET_METHOD": "adaptive_hp",
    },
    "percussion_messy": {
        "description": "Noisy percussion and drumming with sharp transient attacks",
        "HIGHPASS_CUTOFF_HZ": 80,
        "ONSET_AMPLITUDE_GATE": 0.03,
        "ONSET_AMPLITUDE_WINDOW_MS": 50,
        "ONSET_SHARPNESS_GATE": 0.10,
        "ONSET_SHARPNESS_WINDOW_MS": 20,
        "MIN_INTER_ONSET_MS": 30,
        "ONSET_CLUSTER_WINDOW_MS": 25,
        "STABLE_RHYTHM_TOLERANCE": 0.25,
        "ONSET_DELTA": 0.07,
        "ONSET_HOP_LENGTH": 256,
        "ONSET_BACKTRACK": False,
        "ONSET_REFINE_ENABLED": True,
        "ONSET_REFINE_WINDOW_MS": 10,
        "ONSET_REFINE_ENERGY_GATE": 0.0,
        "ONSET_METHOD": "adaptive_hp",
    },
    "primate": {
        "description": "Primate calls and drumming — mid-frequency with moderate attack",
        "HIGHPASS_CUTOFF_HZ": 150,
        "ONSET_AMPLITUDE_GATE": 0.06,
        "ONSET_AMPLITUDE_WINDOW_MS": 60,
        "ONSET_SHARPNESS_GATE": 0.15,
        "ONSET_SHARPNESS_WINDOW_MS": 25,
        "MIN_INTER_ONSET_MS": 40,
        "ONSET_CLUSTER_WINDOW_MS": 30,
        "STABLE_RHYTHM_TOLERANCE": 0.30,
        "ONSET_DELTA": 0.06,
        "ONSET_HOP_LENGTH": 256,
        "ONSET_BACKTRACK": False,
        "ONSET_REFINE_ENABLED": True,
        "ONSET_REFINE_WINDOW_MS": 10,
        "ONSET_REFINE_ENERGY_GATE": 0.0,
        "ONSET_METHOD": "adaptive_hp",
    },
    "insect": {
        "description": "High-frequency percussive clicks (e.g. crickets, cicadas)",
        "HIGHPASS_CUTOFF_HZ": 1000,
        "ONSET_AMPLITUDE_GATE": 0.10,
        "ONSET_AMPLITUDE_WINDOW_MS": 20,
        "ONSET_SHARPNESS_GATE": 0.0,
        "ONSET_SHARPNESS_WINDOW_MS": 15,
        "MIN_INTER_ONSET_MS": 10,
        "ONSET_CLUSTER_WINDOW_MS": 10,
        "STABLE_RHYTHM_TOLERANCE": 0.20,
        "ONSET_DELTA": 0.04,
        "ONSET_HOP_LENGTH": 128,
        "ONSET_BACKTRACK": False,
        "ONSET_REFINE_ENABLED": True,
        "ONSET_REFINE_WINDOW_MS": 10,
        "ONSET_REFINE_ENERGY_GATE": 0.0,
        "ONSET_METHOD": "adaptive_hp",
    },
    # ----- Language / Speech presets -----
    "speech_syllable": {
        "description": "Human speech — syllable-level onsets via Praat intensity peaks (de Jong & Wempe 2009). "
                       "Requires parselmouth.",
        "HIGHPASS_CUTOFF_HZ": 60,
        "ONSET_AMPLITUDE_GATE": 0.0,
        "ONSET_AMPLITUDE_WINDOW_MS": 50,
        "ONSET_SHARPNESS_GATE": 0.0,
        "ONSET_SHARPNESS_WINDOW_MS": 20,
        "MIN_INTER_ONSET_MS": 50,
        "ONSET_CLUSTER_WINDOW_MS": 30,
        "STABLE_RHYTHM_TOLERANCE": 0.30,
        "ONSET_DELTA": 0.05,
        "ONSET_HOP_LENGTH": 256,
        "ONSET_BACKTRACK": False,
        "ONSET_REFINE_ENABLED": False,
        "ONSET_REFINE_WINDOW_MS": 10,
        "ONSET_REFINE_ENERGY_GATE": 0.0,
        "ONSET_METHOD": "syllable_nuclei",
    },
    "speech_word": {
        "description": "Human speech — word-level onsets via Whisper ASR. Requires openai-whisper.",
        "HIGHPASS_CUTOFF_HZ": 60,
        "ONSET_AMPLITUDE_GATE": 0.0,
        "ONSET_AMPLITUDE_WINDOW_MS": 50,
        "ONSET_SHARPNESS_GATE": 0.0,
        "ONSET_SHARPNESS_WINDOW_MS": 20,
        "MIN_INTER_ONSET_MS": 80,
        "ONSET_CLUSTER_WINDOW_MS": 50,
        "STABLE_RHYTHM_TOLERANCE": 0.35,
        "ONSET_DELTA": 0.06,
        "ONSET_HOP_LENGTH": 256,
        "ONSET_BACKTRACK": False,
        "ONSET_REFINE_ENABLED": False,
        "ONSET_REFINE_WINDOW_MS": 10,
        "ONSET_REFINE_ENERGY_GATE": 0.0,
        "ONSET_METHOD": "whisper_words",
    },
    "speech_acoustic": {
        "description": "Human speech — acoustic onset detection using energy-based methods. No extra deps.",
        "HIGHPASS_CUTOFF_HZ": 60,
        "ONSET_AMPLITUDE_GATE": 0.02,
        "ONSET_AMPLITUDE_WINDOW_MS": 40,
        "ONSET_SHARPNESS_GATE": 0.0,
        "ONSET_SHARPNESS_WINDOW_MS": 20,
        "MIN_INTER_ONSET_MS": 40,
        "ONSET_CLUSTER_WINDOW_MS": 20,
        "STABLE_RHYTHM_TOLERANCE": 0.30,
        "ONSET_DELTA": 0.04,
        "ONSET_HOP_LENGTH": 256,
        "ONSET_BACKTRACK": False,
        "ONSET_REFINE_ENABLED": True,
        "ONSET_REFINE_WINDOW_MS": 10,
        "ONSET_REFINE_ENERGY_GATE": 0.0,
        "ONSET_METHOD": "adaptive_hp",
    },
    "speech_phoneme": {
        "description": "Human speech — phoneme-level alignment via WhisperX. Computes Ramus et al. (1999) metrics.",
        "HIGHPASS_CUTOFF_HZ": 60,
        "ONSET_AMPLITUDE_GATE": 0.0,
        "ONSET_AMPLITUDE_WINDOW_MS": 50,
        "ONSET_SHARPNESS_GATE": 0.0,
        "ONSET_SHARPNESS_WINDOW_MS": 20,
        "MIN_INTER_ONSET_MS": 30,
        "ONSET_CLUSTER_WINDOW_MS": 20,
        "STABLE_RHYTHM_TOLERANCE": 0.30,
        "ONSET_DELTA": 0.05,
        "ONSET_HOP_LENGTH": 256,
        "ONSET_BACKTRACK": False,
        "ONSET_REFINE_ENABLED": False,
        "ONSET_REFINE_WINDOW_MS": 10,
        "ONSET_REFINE_ENERGY_GATE": 0.0,
        "ONSET_METHOD": "whisperx_phonemes",
    },
    "music_beat_tracking": {
        "description": "Music / songs — deep-learning beat tracker (Böck et al. 2016) via madmom.",
        "HIGHPASS_CUTOFF_HZ": 30,
        "ONSET_AMPLITUDE_GATE": 0.0,
        "ONSET_AMPLITUDE_WINDOW_MS": 50,
        "ONSET_SHARPNESS_GATE": 0.0,
        "ONSET_SHARPNESS_WINDOW_MS": 20,
        "MIN_INTER_ONSET_MS": 200,
        "ONSET_CLUSTER_WINDOW_MS": 30,
        "STABLE_RHYTHM_TOLERANCE": 0.35,
        "ONSET_DELTA": 0.05,
        "ONSET_HOP_LENGTH": 256,
        "ONSET_BACKTRACK": False,
        "ONSET_REFINE_ENABLED": False,
        "ONSET_REFINE_WINDOW_MS": 10,
        "ONSET_REFINE_ENERGY_GATE": 0.0,
        "ONSET_METHOD": "madmom_beats",
        "MADMOM_MIN_BPM": 40,
        "MADMOM_MAX_BPM": 240,
        "MADMOM_DOWNBEATS": True,
        "MADMOM_FPS": 100,
        "MADMOM_TRANSITION_LAMBDA": 100,
        "PITCH_TRACKER": "pyin",
        "PITCH_FMIN": 65.0,
        "PITCH_FMAX": 1047.0,
        "TEMPO_ADAPTIVE_MIN_IOI": True,
        "TEMPO_ADAPTIVE_FRACTION": 0.5,
    },
}

ACTIVE_PRESET = None

# Output toggles
CREATE_SPECTROGRAMS = True
SPECTROGRAM_CHUNK_ENABLED = False
SPECTROGRAM_CHUNK_SECONDS = 30
CREATE_AUDACITY_LABELS = True
ADD_COLUMN_COMMENTS = True
CLUSTER_OVERLAPPING_ONSETS = True
ONSET_CLUSTER_WINDOW_MS = 25
FILTER_STABLE_RHYTHMS = True
STABLE_RHYTHM_TOLERANCE = 0.25

# Noise-handling fallbacks (disabled by default — the muter handles this)
APPLY_HIGHPASS_FILTER = False
HIGHPASS_CUTOFF_HZ = 200

ONSET_AMPLITUDE_GATE = 0.05
ONSET_AMPLITUDE_WINDOW_MS = 50
ONSET_SHARPNESS_GATE = 0.0
ONSET_SHARPNESS_WINDOW_MS = 20
MIN_INTER_ONSET_MS = 30

# Onset detection core
ONSET_DELTA = 0.10
ONSET_HOP_LENGTH = 256
ONSET_BACKTRACK = False

# Sample-level refinement
ONSET_REFINE_ENABLED = True
ONSET_REFINE_WINDOW_MS = 10
ONSET_REFINE_ENERGY_GATE = 0.0

# Onset method selection
ONSET_METHOD = "adaptive_hp"

# HP filter parameters
HP_SMOOTH_LAMBDA = 50
HP_THRESHOLD_LAMBDA = 5e7
HP_ENVELOPE_WINDOW_MS = 10
HP_ENVELOPE_HOP_MS = 1

# Moving-median parameters
MEDIAN_WINDOW_MS = 200
MEDIAN_THRESHOLD_SCALE = 1.5

# Superflux parameters
SUPERFLUX_LAG = 2
SUPERFLUX_MAX_SIZE = 3

# CFAR parameters
CFAR_GUARD_MS = 20
CFAR_TRAINING_MS = 200
CFAR_THRESHOLD_FACTOR = 4.0

# Per-band parameters
PER_BAND_N_BANDS = 6
PER_BAND_FREQ_MIN = 200
PER_BAND_FREQ_MAX = None
PER_BAND_MEDIAN_MS = 200
PER_BAND_THRESHOLD_SCALE = 1.5
PER_BAND_MIN_BANDS = 2

# Syllable nuclei parameters (used when ONSET_METHOD == "syllable_nuclei")
SYLLABLE_INTENSITY_THRESHOLD = -25.0
SYLLABLE_MIN_DIP_DB = 2.0
SYLLABLE_MIN_PAUSE_MS = 30.0
SYLLABLE_VOICING_THRESHOLD = 0.3
SYLLABLE_TIME_STEP = 0.01

# Whisper word-onset parameters (used when ONSET_METHOD == "whisper_words")
WHISPER_MODEL_SIZE = "base"
WHISPER_LANGUAGE = None
WHISPER_WORD_TIMESTAMPS = True

# WhisperX phoneme alignment parameters (used when ONSET_METHOD == "whisperx_phonemes")
WHISPERX_MODEL_SIZE = "base"
WHISPERX_LANGUAGE = None
WHISPERX_DEVICE = "cpu"

# madmom beat tracker parameters (ONSET_METHOD == "madmom_beats")
MADMOM_MIN_BPM = 40
MADMOM_MAX_BPM = 240
MADMOM_FPS = 100
MADMOM_DOWNBEATS = False
MADMOM_TRANSITION_LAMBDA = 100

# Standalone pitch tracker (runs after onset detection, any method)
PITCH_TRACKER = "none"               # "none", "pyin", "crepe", "praat"
PITCH_FMIN = 65.0                    # Min expected F0 in Hz
PITCH_FMAX = 1047.0                  # Max expected F0 in Hz

# Tempo-adaptive MIN_INTER_ONSET_MS
TEMPO_ADAPTIVE_MIN_IOI = False
TEMPO_ADAPTIVE_FRACTION = 0.5

# Speech analysis options
PAUSE_THRESHOLD_MS = 250.0
EXPORT_TEXTGRID = True
EXPORT_TRANSCRIPT = True

# --- Apply preset overrides ---
if ACTIVE_PRESET is not None:
    if ACTIVE_PRESET not in PRESETS:
        raise ValueError(f"Unknown preset '{ACTIVE_PRESET}'. Choose from: {list(PRESETS.keys())}")
    _preset = PRESETS[ACTIVE_PRESET]
    print(f"Using preset '{ACTIVE_PRESET}': {_preset['description']}")
    HIGHPASS_CUTOFF_HZ = _preset["HIGHPASS_CUTOFF_HZ"]
    ONSET_AMPLITUDE_GATE = _preset["ONSET_AMPLITUDE_GATE"]
    ONSET_AMPLITUDE_WINDOW_MS = _preset["ONSET_AMPLITUDE_WINDOW_MS"]
    ONSET_SHARPNESS_GATE = _preset["ONSET_SHARPNESS_GATE"]
    ONSET_SHARPNESS_WINDOW_MS = _preset["ONSET_SHARPNESS_WINDOW_MS"]
    MIN_INTER_ONSET_MS = _preset["MIN_INTER_ONSET_MS"]
    ONSET_CLUSTER_WINDOW_MS = _preset["ONSET_CLUSTER_WINDOW_MS"]
    STABLE_RHYTHM_TOLERANCE = _preset["STABLE_RHYTHM_TOLERANCE"]
    ONSET_DELTA = _preset["ONSET_DELTA"]
    ONSET_HOP_LENGTH = _preset["ONSET_HOP_LENGTH"]
    ONSET_BACKTRACK = _preset["ONSET_BACKTRACK"]
    ONSET_REFINE_ENABLED = _preset["ONSET_REFINE_ENABLED"]
    ONSET_REFINE_WINDOW_MS = _preset["ONSET_REFINE_WINDOW_MS"]
    ONSET_REFINE_ENERGY_GATE = _preset["ONSET_REFINE_ENERGY_GATE"]
    ONSET_METHOD = _preset["ONSET_METHOD"]
    MADMOM_MIN_BPM = _preset.get("MADMOM_MIN_BPM", MADMOM_MIN_BPM)
    MADMOM_MAX_BPM = _preset.get("MADMOM_MAX_BPM", MADMOM_MAX_BPM)
    MADMOM_FPS = _preset.get("MADMOM_FPS", MADMOM_FPS)
    MADMOM_DOWNBEATS = _preset.get("MADMOM_DOWNBEATS", MADMOM_DOWNBEATS)
    MADMOM_TRANSITION_LAMBDA = _preset.get("MADMOM_TRANSITION_LAMBDA", MADMOM_TRANSITION_LAMBDA)
    PITCH_TRACKER = _preset.get("PITCH_TRACKER", PITCH_TRACKER)
    PITCH_FMIN = _preset.get("PITCH_FMIN", PITCH_FMIN)
    PITCH_FMAX = _preset.get("PITCH_FMAX", PITCH_FMAX)
    TEMPO_ADAPTIVE_MIN_IOI = _preset.get("TEMPO_ADAPTIVE_MIN_IOI", TEMPO_ADAPTIVE_MIN_IOI)
    TEMPO_ADAPTIVE_FRACTION = _preset.get("TEMPO_ADAPTIVE_FRACTION", TEMPO_ADAPTIVE_FRACTION)

_PROJECT_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
audio_folder = os.path.join(_PROJECT_DIR, "audioFiles_muted_clean")
output_excel_path = os.path.join(_PROJECT_DIR, "Cross_Species_Rhythm_Data.xlsx")
EXTRACTOR_SPECIFY_FILES = False
EXTRACTOR_SELECTED_FILES = []

# ------------------------------------------------------------------
# GUI CONFIG OVERRIDE — Read settings from pipeline_config.json if present.
# ------------------------------------------------------------------
_config_path = os.path.join(_PROJECT_DIR, "pipeline_config.json")
if os.path.isfile(_config_path):
    import json as _json
    with open(_config_path) as _f:
        _cfg = _json.load(_f).get("extractor", {})
    audio_folder = _cfg.get("audio_folder", audio_folder)
    output_excel_path = _cfg.get("output_excel_path", output_excel_path)
    EXTRACTOR_SPECIFY_FILES = _cfg.get("EXTRACTOR_SPECIFY_FILES", EXTRACTOR_SPECIFY_FILES)
    EXTRACTOR_SELECTED_FILES = _cfg.get("EXTRACTOR_SELECTED_FILES", EXTRACTOR_SELECTED_FILES)
    ACTIVE_PRESET = _cfg.get("ACTIVE_PRESET", ACTIVE_PRESET)
    CREATE_SPECTROGRAMS = _cfg.get("CREATE_SPECTROGRAMS", CREATE_SPECTROGRAMS)
    SPECTROGRAM_CHUNK_ENABLED = _cfg.get("SPECTROGRAM_CHUNK_ENABLED", SPECTROGRAM_CHUNK_ENABLED)
    SPECTROGRAM_CHUNK_SECONDS = _cfg.get("SPECTROGRAM_CHUNK_SECONDS", SPECTROGRAM_CHUNK_SECONDS)
    CREATE_AUDACITY_LABELS = _cfg.get("CREATE_AUDACITY_LABELS", CREATE_AUDACITY_LABELS)
    ADD_COLUMN_COMMENTS = _cfg.get("ADD_COLUMN_COMMENTS", True)
    CLUSTER_OVERLAPPING_ONSETS = _cfg.get("CLUSTER_OVERLAPPING_ONSETS", CLUSTER_OVERLAPPING_ONSETS)
    ONSET_CLUSTER_WINDOW_MS = _cfg.get("ONSET_CLUSTER_WINDOW_MS", ONSET_CLUSTER_WINDOW_MS)
    FILTER_STABLE_RHYTHMS = _cfg.get("FILTER_STABLE_RHYTHMS", FILTER_STABLE_RHYTHMS)
    STABLE_RHYTHM_TOLERANCE = _cfg.get("STABLE_RHYTHM_TOLERANCE", STABLE_RHYTHM_TOLERANCE)
    APPLY_HIGHPASS_FILTER = _cfg.get("APPLY_HIGHPASS_FILTER", APPLY_HIGHPASS_FILTER)
    HIGHPASS_CUTOFF_HZ = _cfg.get("HIGHPASS_CUTOFF_HZ", HIGHPASS_CUTOFF_HZ)
    ONSET_AMPLITUDE_GATE = _cfg.get("ONSET_AMPLITUDE_GATE", ONSET_AMPLITUDE_GATE)
    ONSET_AMPLITUDE_WINDOW_MS = _cfg.get("ONSET_AMPLITUDE_WINDOW_MS", ONSET_AMPLITUDE_WINDOW_MS)
    ONSET_SHARPNESS_GATE = _cfg.get("ONSET_SHARPNESS_GATE", ONSET_SHARPNESS_GATE)
    ONSET_SHARPNESS_WINDOW_MS = _cfg.get("ONSET_SHARPNESS_WINDOW_MS", ONSET_SHARPNESS_WINDOW_MS)
    MIN_INTER_ONSET_MS = _cfg.get("MIN_INTER_ONSET_MS", MIN_INTER_ONSET_MS)
    ONSET_METHOD = _cfg.get("ONSET_METHOD", ONSET_METHOD)
    ONSET_DELTA = _cfg.get("ONSET_DELTA", ONSET_DELTA)
    ONSET_HOP_LENGTH = _cfg.get("ONSET_HOP_LENGTH", ONSET_HOP_LENGTH)
    ONSET_BACKTRACK = _cfg.get("ONSET_BACKTRACK", ONSET_BACKTRACK)
    ONSET_REFINE_ENABLED = _cfg.get("ONSET_REFINE_ENABLED", ONSET_REFINE_ENABLED)
    ONSET_REFINE_WINDOW_MS = _cfg.get("ONSET_REFINE_WINDOW_MS", ONSET_REFINE_WINDOW_MS)
    ONSET_REFINE_ENERGY_GATE = _cfg.get("ONSET_REFINE_ENERGY_GATE", ONSET_REFINE_ENERGY_GATE)
    HP_SMOOTH_LAMBDA = _cfg.get("HP_SMOOTH_LAMBDA", HP_SMOOTH_LAMBDA)
    HP_THRESHOLD_LAMBDA = _cfg.get("HP_THRESHOLD_LAMBDA", HP_THRESHOLD_LAMBDA)
    HP_ENVELOPE_WINDOW_MS = _cfg.get("HP_ENVELOPE_WINDOW_MS", HP_ENVELOPE_WINDOW_MS)
    HP_ENVELOPE_HOP_MS = _cfg.get("HP_ENVELOPE_HOP_MS", HP_ENVELOPE_HOP_MS)
    MEDIAN_WINDOW_MS = _cfg.get("MEDIAN_WINDOW_MS", MEDIAN_WINDOW_MS)
    MEDIAN_THRESHOLD_SCALE = _cfg.get("MEDIAN_THRESHOLD_SCALE", MEDIAN_THRESHOLD_SCALE)
    SUPERFLUX_LAG = _cfg.get("SUPERFLUX_LAG", SUPERFLUX_LAG)
    SUPERFLUX_MAX_SIZE = _cfg.get("SUPERFLUX_MAX_SIZE", SUPERFLUX_MAX_SIZE)
    CFAR_GUARD_MS = _cfg.get("CFAR_GUARD_MS", CFAR_GUARD_MS)
    CFAR_TRAINING_MS = _cfg.get("CFAR_TRAINING_MS", CFAR_TRAINING_MS)
    CFAR_THRESHOLD_FACTOR = _cfg.get("CFAR_THRESHOLD_FACTOR", CFAR_THRESHOLD_FACTOR)
    PER_BAND_N_BANDS = _cfg.get("PER_BAND_N_BANDS", PER_BAND_N_BANDS)
    PER_BAND_FREQ_MIN = _cfg.get("PER_BAND_FREQ_MIN", PER_BAND_FREQ_MIN)
    PER_BAND_FREQ_MAX = _cfg.get("PER_BAND_FREQ_MAX", PER_BAND_FREQ_MAX)
    PER_BAND_MEDIAN_MS = _cfg.get("PER_BAND_MEDIAN_MS", PER_BAND_MEDIAN_MS)
    PER_BAND_THRESHOLD_SCALE = _cfg.get("PER_BAND_THRESHOLD_SCALE", PER_BAND_THRESHOLD_SCALE)
    PER_BAND_MIN_BANDS = _cfg.get("PER_BAND_MIN_BANDS", PER_BAND_MIN_BANDS)
    # Syllable nuclei parameters
    SYLLABLE_INTENSITY_THRESHOLD = float(_cfg.get("SYLLABLE_INTENSITY_THRESHOLD", SYLLABLE_INTENSITY_THRESHOLD))
    SYLLABLE_MIN_DIP_DB = float(_cfg.get("SYLLABLE_MIN_DIP_DB", SYLLABLE_MIN_DIP_DB))
    SYLLABLE_MIN_PAUSE_MS = float(_cfg.get("SYLLABLE_MIN_PAUSE_MS", SYLLABLE_MIN_PAUSE_MS))
    SYLLABLE_VOICING_THRESHOLD = float(_cfg.get("SYLLABLE_VOICING_THRESHOLD", SYLLABLE_VOICING_THRESHOLD))
    SYLLABLE_TIME_STEP = float(_cfg.get("SYLLABLE_TIME_STEP", SYLLABLE_TIME_STEP))
    # Whisper word-onset parameters
    WHISPER_MODEL_SIZE = _cfg.get("WHISPER_MODEL_SIZE", WHISPER_MODEL_SIZE)
    WHISPER_LANGUAGE = _cfg.get("WHISPER_LANGUAGE", WHISPER_LANGUAGE)
    WHISPER_WORD_TIMESTAMPS = bool(_cfg.get("WHISPER_WORD_TIMESTAMPS", WHISPER_WORD_TIMESTAMPS))
    # WhisperX phoneme alignment parameters
    WHISPERX_MODEL_SIZE = _cfg.get("WHISPERX_MODEL_SIZE", WHISPERX_MODEL_SIZE)
    WHISPERX_LANGUAGE = _cfg.get("WHISPERX_LANGUAGE", WHISPERX_LANGUAGE)
    WHISPERX_DEVICE = _cfg.get("WHISPERX_DEVICE", WHISPERX_DEVICE)
    # madmom beat tracker parameters
    MADMOM_MIN_BPM = _cfg.get("MADMOM_MIN_BPM", MADMOM_MIN_BPM)
    MADMOM_MAX_BPM = _cfg.get("MADMOM_MAX_BPM", MADMOM_MAX_BPM)
    MADMOM_FPS = _cfg.get("MADMOM_FPS", MADMOM_FPS)
    MADMOM_DOWNBEATS = bool(_cfg.get("MADMOM_DOWNBEATS", MADMOM_DOWNBEATS))
    MADMOM_TRANSITION_LAMBDA = _cfg.get("MADMOM_TRANSITION_LAMBDA", MADMOM_TRANSITION_LAMBDA)
    # Standalone pitch tracker
    PITCH_TRACKER = _cfg.get("PITCH_TRACKER", PITCH_TRACKER)
    PITCH_FMIN = float(_cfg.get("PITCH_FMIN", PITCH_FMIN))
    PITCH_FMAX = float(_cfg.get("PITCH_FMAX", PITCH_FMAX))
    # Tempo-adaptive MIN_INTER_ONSET_MS
    TEMPO_ADAPTIVE_MIN_IOI = bool(_cfg.get("TEMPO_ADAPTIVE_MIN_IOI", TEMPO_ADAPTIVE_MIN_IOI))
    TEMPO_ADAPTIVE_FRACTION = float(_cfg.get("TEMPO_ADAPTIVE_FRACTION", TEMPO_ADAPTIVE_FRACTION))
    # Speech analysis options
    PAUSE_THRESHOLD_MS = float(_cfg.get("PAUSE_THRESHOLD_MS", PAUSE_THRESHOLD_MS))
    EXPORT_TEXTGRID = bool(_cfg.get("EXPORT_TEXTGRID", EXPORT_TEXTGRID))
    EXPORT_TRANSCRIPT = bool(_cfg.get("EXPORT_TRANSCRIPT", EXPORT_TRANSCRIPT))
    del _json, _f, _cfg
del _PROJECT_DIR, _config_path

# Accumulators for Excel export
file_summary_data = []
dyadic_events_data = []
stable_dyadic_events_data = []


# ==========================================
# HELPER FUNCTIONS
# ==========================================

def cluster_onsets(onset_times, cluster_window_ms):
    """Merge near-simultaneous onsets into a single onset at the cluster start."""
    if len(onset_times) <= 1:
        return np.array(onset_times, dtype=float)

    cluster_window_seconds = cluster_window_ms / 1000.0
    clustered_onsets = [float(onset_times[0])]

    for onset_time in onset_times[1:]:
        if float(onset_time) - clustered_onsets[-1] < cluster_window_seconds:
            continue
        clustered_onsets.append(float(onset_time))

    return np.array(clustered_onsets, dtype=float)


def resolve_audio_files(audio_folder_path, selected_files=None,
                        valid_extensions=(".wav", ".mp3", ".flac", ".ogg")):
    """Return supported audio filenames, optionally filtered to *selected_files*."""
    available = [
        name for name in sorted(os.listdir(audio_folder_path))
        if name.lower().endswith(valid_extensions)
    ]
    requested = list(dict.fromkeys(selected_files or []))
    if not requested:
        return available, []
    requested_set = set(requested)
    filtered = [name for name in available if name in requested_set]
    missing = [name for name in requested if name not in available]
    return filtered, missing


def apply_highpass(signal, sr, cutoff_hz):
    """Apply a 4th-order Butterworth high-pass filter to the audio signal."""
    sos = butter(4, cutoff_hz, btype="high", fs=sr, output="sos")
    return sosfilt(sos, signal).astype(signal.dtype)


def gate_onsets_by_amplitude(onset_times, signal, sr, gate_fraction, window_ms):
    """Remove onsets whose local RMS energy falls below *gate_fraction* of the peak RMS."""
    if gate_fraction <= 0 or len(onset_times) == 0:
        return onset_times

    window_samples = int(sr * window_ms / 1000.0)
    rms_values = librosa.feature.rms(y=signal, frame_length=window_samples, hop_length=window_samples // 2)[0]
    rms_times = librosa.frames_to_time(
        np.arange(len(rms_values)), sr=sr, hop_length=window_samples // 2
    )
    peak_rms = float(np.max(rms_values))
    threshold = gate_fraction * peak_rms

    kept = []
    for onset_time in onset_times:
        closest_idx = int(np.argmin(np.abs(rms_times - onset_time)))
        if rms_values[closest_idx] >= threshold:
            kept.append(onset_time)

    return np.array(kept, dtype=float)


def gate_onsets_by_sharpness(onset_times, signal, sr, gate_fraction, window_ms):
    """Remove onsets whose attack slope is below *gate_fraction* of the steepest attack.

    Attack slope is measured as the maximum positive derivative of the Hilbert
    amplitude envelope within a ±window_ms region around each onset.  Onsets
    with a slope below ``gate_fraction * peak_slope`` are discarded.
    """
    if gate_fraction <= 0 or len(onset_times) == 0:
        return onset_times

    from scipy.signal import hilbert as scipy_hilbert

    analytic = scipy_hilbert(signal)
    envelope = np.abs(analytic).astype(np.float64)
    envelope_diff = np.diff(envelope)

    window_samples = int(sr * window_ms / 1000.0)
    half_window = window_samples // 2

    slopes = []
    for onset_time in onset_times:
        center = int(onset_time * sr)
        start = max(center - half_window, 0)
        end = min(center + half_window, len(envelope_diff))
        if start < end:
            slopes.append(float(np.max(envelope_diff[start:end])))
        else:
            slopes.append(0.0)

    peak_slope = max(slopes) if slopes else 1.0
    if peak_slope <= 0:
        return onset_times

    threshold = gate_fraction * peak_slope
    kept = [t for t, s in zip(onset_times, slopes) if s >= threshold]
    return np.array(kept, dtype=float)


def enforce_min_interval(onset_times, min_interval_ms):
    """Drop onsets that follow the previous one by fewer than *min_interval_ms*."""
    if min_interval_ms <= 0 or len(onset_times) <= 1:
        return onset_times

    min_interval_s = min_interval_ms / 1000.0
    kept = [onset_times[0]]
    for t in onset_times[1:]:
        if t - kept[-1] >= min_interval_s:
            kept.append(t)
    return np.array(kept, dtype=float)


def is_stable_match(interval_a_ms, interval_b_ms, tolerance):
    """Return True when two alternating intervals differ by no more than the tolerance."""
    largest_interval = max(interval_a_ms, interval_b_ms)
    if largest_interval <= 0:
        return False
    return abs(interval_a_ms - interval_b_ms) / largest_interval <= tolerance


def get_stable_dyad_flags(intervals_ms, tolerance):
    """Flag dyads that belong to locally stable alternating interval patterns."""
    dyad_count = max(len(intervals_ms) - 1, 0)
    stable_flags = []

    for index in range(dyad_count):
        comparisons = []

        if index + 2 < len(intervals_ms):
            comparisons.append(is_stable_match(intervals_ms[index], intervals_ms[index + 2], tolerance))

        if index + 3 < len(intervals_ms):
            comparisons.append(is_stable_match(intervals_ms[index + 1], intervals_ms[index + 3], tolerance))

        stable_flags.append(bool(comparisons) and all(comparisons))

    return stable_flags


def build_dyad_records(filename, intervals_ms, stable_flags):
    """Convert an interval sequence into dyadic-event rows for export and plotting."""
    records = []

    for index in range(len(intervals_ms) - 1):
        interval_1 = float(intervals_ms[index])
        interval_2 = float(intervals_ms[index + 1])
        cycle_duration = interval_1 + interval_2
        short_interval = min(interval_1, interval_2)
        long_interval = max(interval_1, interval_2)
        rhythm_ratio = interval_1 / cycle_duration if cycle_duration > 0 else 0

        records.append(
            {
                "File Name": filename,
                "Dyad Index": index + 1,
                "Interval 1 (ms)": round(interval_1, 2),
                "Interval 2 (ms)": round(interval_2, 2),
                "Cycle Duration [cd] (ms)": round(cycle_duration, 2),
                "Short Interval [i_s] (ms)": round(short_interval, 2),
                "Long Interval [i_l] (ms)": round(long_interval, 2),
                "Rhythm Ratio [r_k]": round(rhythm_ratio, 4),
                "Stable Rhythm": stable_flags[index],
            }
        )

    return records


def calculate_rhythm_metrics_thebeat(seq, dyad_records):
    """Calculate file-level rhythm summary statistics using the thebeat Sequence.

    The thebeat Sequence object provides IOIs directly; we use these alongside
    the dyad records built from the same intervals for full metric computation.
    """
    if seq is None or len(seq.iois) < 2 or not dyad_records:
        return {
            "Average Cycle Duration (ms)": None,
            "nPVI (Isochrony)": None,
            "CV of Intervals": None,
            "r_k Std Dev": None,
            "r_k Entropy (Categorical Measure)": None,
        }

    cycle_durations = np.array([r["Cycle Duration [cd] (ms)"] for r in dyad_records], dtype=float)
    rk_values = np.array([r["Rhythm Ratio [r_k]"] for r in dyad_records], dtype=float)

    average_cycle_duration = float(np.mean(cycle_durations))
    rk_count = len(rk_values)
    npvi = (400 / rk_count) * np.sum(np.abs(rk_values - 0.5)) if rk_count > 0 else 0

    # Use thebeat Sequence's IOIs (in seconds) for interval stats
    intervals_s = seq.iois
    mean_interval = float(np.mean(intervals_s))
    interval_cv = float(np.std(intervals_s) / mean_interval) if mean_interval > 0 else 0
    rk_std_dev = float(np.std(rk_values)) if rk_count > 0 else 0

    counts, _ = np.histogram(rk_values, bins=10, range=(0, 1))
    rk_entropy = float(entropy(counts))

    return {
        "Average Cycle Duration (ms)": average_cycle_duration,
        "nPVI (Isochrony)": npvi,
        "CV of Intervals": interval_cv,
        "r_k Std Dev": rk_std_dev,
        "r_k Entropy (Categorical Measure)": rk_entropy,
    }


def calculate_stable_subset_metrics(stable_dyad_records):
    """Calculate summary statistics for the stable-rhythm dyads only."""
    if not stable_dyad_records:
        return {
            "Stable Rhythm nPVI": None,
            "Stable Rhythm CV": None,
            "Stable Rhythm r_k Std Dev": None,
            "Stable Rhythm Entropy": None,
        }

    rk_values = np.array([r["Rhythm Ratio [r_k]"] for r in stable_dyad_records], dtype=float)
    interval_values = np.array(
        [
            value
            for r in stable_dyad_records
            for value in (r["Interval 1 (ms)"], r["Interval 2 (ms)"])
        ],
        dtype=float,
    )

    mean_interval = float(np.mean(interval_values)) if len(interval_values) > 0 else 0
    stable_cv = float(np.std(interval_values) / mean_interval) if mean_interval > 0 else 0
    rk_count = len(rk_values)
    stable_npvi = (400 / rk_count) * np.sum(np.abs(rk_values - 0.5)) if rk_count > 0 else 0
    stable_std_dev = float(np.std(rk_values)) if rk_count > 0 else 0

    counts, _ = np.histogram(rk_values, bins=10, range=(0, 1))
    stable_entropy = float(entropy(counts))

    return {
        "Stable Rhythm nPVI": stable_npvi,
        "Stable Rhythm CV": stable_cv,
        "Stable Rhythm r_k Std Dev": stable_std_dev,
        "Stable Rhythm Entropy": stable_entropy,
    }


def main():
    # ==========================================
    # 2. AUDIO PROCESSING LOOP
    # ==========================================
    if not os.path.isdir(audio_folder):
        print(f"ERROR: Audio folder not found: {audio_folder}")
        print("Run the Audio Editor (Step 1) first, or update the audio_folder path.")
        sys.exit(1)

    print("Starting extraction process using 'thebeat' engine...")

    _VALID_EXTENSIONS = (".wav", ".mp3", ".flac", ".ogg")
    _selected_files = EXTRACTOR_SELECTED_FILES if EXTRACTOR_SPECIFY_FILES else []
    _audio_files, _missing_audio_files = resolve_audio_files(
        audio_folder, _selected_files, _VALID_EXTENSIONS)

    for _missing_name in _missing_audio_files:
        print(f"WARNING: Requested file not found or unsupported: {_missing_name}")

    if _selected_files:
        print(f"Selected files: {len(_audio_files)} of {len(_selected_files)} requested file(s)")

    if not _audio_files:
        print(f"WARNING: No audio files found in {audio_folder}")
        print(f"Supported formats: {', '.join(_VALID_EXTENSIONS)}")

    for filename in _audio_files:
        file_path = os.path.join(audio_folder, filename)
        print(f"Processing: {filename}...")

        try:

            y, sr = librosa.load(file_path, sr=None)

            # Apply high-pass filter to remove low-frequency rumble before onset detection.
            if APPLY_HIGHPASS_FILTER and HIGHPASS_CUTOFF_HZ > 0:
                y = apply_highpass(y, sr, HIGHPASS_CUTOFF_HZ)

            total_duration = librosa.get_duration(y=y, sr=sr)

            # Coarse acoustic summaries for the file summary sheet.
            tempo_bpm, _ = librosa.beat.beat_track(y=y, sr=sr)
            estimated_bpm = float(tempo_bpm[0]) if isinstance(tempo_bpm, np.ndarray) else float(tempo_bpm)

            rms = librosa.feature.rms(y=y)
            avg_loudness = np.mean(rms)

            centroid = librosa.feature.spectral_centroid(y=y, sr=sr)
            avg_brightness = np.mean(centroid)

            # Build keyword arguments for the selected onset method.
            if ONSET_METHOD == "adaptive_hp":
                _onset_kwargs = dict(
                    hp_smooth_lambda=HP_SMOOTH_LAMBDA,
                    hp_threshold_lambda=HP_THRESHOLD_LAMBDA,
                    envelope_window_ms=HP_ENVELOPE_WINDOW_MS,
                    envelope_hop_ms=HP_ENVELOPE_HOP_MS,
                )
            elif ONSET_METHOD == "librosa":
                _onset_kwargs = dict(
                    hop_length=ONSET_HOP_LENGTH,
                    delta=ONSET_DELTA,
                    backtrack=ONSET_BACKTRACK,
                )
            elif ONSET_METHOD == "moving_median":
                _onset_kwargs = dict(
                    envelope_window_ms=HP_ENVELOPE_WINDOW_MS,
                    envelope_hop_ms=HP_ENVELOPE_HOP_MS,
                    median_window_ms=MEDIAN_WINDOW_MS,
                    threshold_scale=MEDIAN_THRESHOLD_SCALE,
                )
            elif ONSET_METHOD == "superflux":
                _onset_kwargs = dict(
                    hop_length=ONSET_HOP_LENGTH,
                    lag=SUPERFLUX_LAG,
                    max_size=SUPERFLUX_MAX_SIZE,
                    delta=ONSET_DELTA,
                )
            elif ONSET_METHOD == "cfar":
                _onset_kwargs = dict(
                    envelope_window_ms=HP_ENVELOPE_WINDOW_MS,
                    envelope_hop_ms=HP_ENVELOPE_HOP_MS,
                    guard_ms=CFAR_GUARD_MS,
                    training_ms=CFAR_TRAINING_MS,
                    threshold_factor=CFAR_THRESHOLD_FACTOR,
                )
            elif ONSET_METHOD == "per_band":
                _onset_kwargs = dict(
                    hop_length=ONSET_HOP_LENGTH,
                    n_bands=PER_BAND_N_BANDS,
                    freq_min=PER_BAND_FREQ_MIN,
                    freq_max=PER_BAND_FREQ_MAX,
                    median_window_ms=PER_BAND_MEDIAN_MS,
                    threshold_scale=PER_BAND_THRESHOLD_SCALE,
                    min_bands=PER_BAND_MIN_BANDS,
                )
            elif ONSET_METHOD == "syllable_nuclei":
                _onset_kwargs = dict(
                    intensity_threshold=SYLLABLE_INTENSITY_THRESHOLD,
                    min_dip_db=SYLLABLE_MIN_DIP_DB,
                    min_pause_ms=SYLLABLE_MIN_PAUSE_MS,
                    voicing_threshold=SYLLABLE_VOICING_THRESHOLD,
                    time_step=SYLLABLE_TIME_STEP,
                )
            elif ONSET_METHOD == "whisper_words":
                _onset_kwargs = dict(
                    model_size=WHISPER_MODEL_SIZE,
                    language=WHISPER_LANGUAGE,
                    word_timestamps=WHISPER_WORD_TIMESTAMPS,
                )
            elif ONSET_METHOD == "whisperx_phonemes":
                _onset_kwargs = dict(
                    model_size=WHISPERX_MODEL_SIZE,
                    language=WHISPERX_LANGUAGE,
                    device=WHISPERX_DEVICE,
                )
            elif ONSET_METHOD == "madmom_beats":
                _onset_kwargs = dict(
                    min_bpm=MADMOM_MIN_BPM,
                    max_bpm=MADMOM_MAX_BPM,
                    fps=MADMOM_FPS,
                    downbeats=MADMOM_DOWNBEATS,
                    transition_lambda=MADMOM_TRANSITION_LAMBDA,
                )
            else:
                _onset_kwargs = {}

            onset_times = detect_onsets(ONSET_METHOD, y, sr, **_onset_kwargs)

            # Standalone pitch tracker (runs independently of onset method)
            if PITCH_TRACKER and PITCH_TRACKER != "none":
                try:
                    extract_pitch_metrics(
                        y, sr, method=PITCH_TRACKER,
                        fmin=PITCH_FMIN, fmax=PITCH_FMAX,
                    )
                except Exception as _pitch_err:
                    print(f"  -> Pitch tracker ({PITCH_TRACKER}) failed: {_pitch_err}")

            # Sample-level refinement
            if ONSET_REFINE_ENABLED and len(onset_times) > 0:
                onset_times = refine_onsets_to_sample(
                    onset_times, y, sr,
                    window_ms=ONSET_REFINE_WINDOW_MS,
                    energy_gate=ONSET_REFINE_ENERGY_GATE,
                )

            # Merge near-simultaneous onsets
            clustered_onset_times = (
                cluster_onsets(onset_times, ONSET_CLUSTER_WINDOW_MS)
                if CLUSTER_OVERLAPPING_ONSETS
                else np.array(onset_times, dtype=float)
            )

            # Amplitude gate
            if ONSET_AMPLITUDE_GATE > 0:
                clustered_onset_times = gate_onsets_by_amplitude(
                    clustered_onset_times, y, sr, ONSET_AMPLITUDE_GATE, ONSET_AMPLITUDE_WINDOW_MS
                )

            # Sharpness gate
            if ONSET_SHARPNESS_GATE > 0:
                clustered_onset_times = gate_onsets_by_sharpness(
                    clustered_onset_times, y, sr, ONSET_SHARPNESS_GATE, ONSET_SHARPNESS_WINDOW_MS
                )

            # Minimum inter-onset spacing
            # Tempo-adaptive override: compute MIN_INTER_ONSET_MS from BPM.
            _effective_min_ioi = MIN_INTER_ONSET_MS
            if TEMPO_ADAPTIVE_MIN_IOI:
                _est_bpm = None
                if last_madmom_tempo and last_madmom_tempo.get("BPM", 0) > 0:
                    _est_bpm = last_madmom_tempo["BPM"]
                elif len(clustered_onset_times) >= 3:
                    _iois = np.diff(clustered_onset_times)
                    _med_ioi = float(np.median(_iois))
                    if _med_ioi > 0:
                        _est_bpm = 60.0 / _med_ioi
                if _est_bpm and _est_bpm > 0:
                    _beat_ms = 60000.0 / _est_bpm
                    _effective_min_ioi = max(1, int(_beat_ms * TEMPO_ADAPTIVE_FRACTION))
                    print(f"  -> Tempo-adaptive min IOI: {_est_bpm:.1f} BPM "
                          f"× {TEMPO_ADAPTIVE_FRACTION:.0%} = {_effective_min_ioi} ms")
            if _effective_min_ioi > 0:
                clustered_onset_times = enforce_min_interval(clustered_onset_times, _effective_min_ioi)

            # --- USING 'thebeat' FOR RHYTHM ANALYSIS ---
            seq = None
            intervals_seconds = np.diff(clustered_onset_times)
            intervals_ms = intervals_seconds * 1000

            if len(clustered_onset_times) >= 3:
                # Build a thebeat Sequence from IOIs (intervals between events).
                iois_seconds = np.diff(clustered_onset_times)
                seq = Sequence(iois_seconds)
                # thebeat stores IOIs in the same units we gave it (seconds),
                # so we use seq.iois for consistency checks and metric computation.
                intervals_ms = seq.iois * 1000
                intervals_seconds = seq.iois

            stable_dyad_flags = get_stable_dyad_flags(intervals_ms, STABLE_RHYTHM_TOLERANCE)
            file_dyad_records = build_dyad_records(filename, intervals_ms, stable_dyad_flags)
            file_stable_dyad_records = [r for r in file_dyad_records if r["Stable Rhythm"]]

            dyadic_events_data.extend(file_dyad_records)
            if FILTER_STABLE_RHYTHMS:
                stable_dyadic_events_data.extend(file_stable_dyad_records)

            metrics = calculate_rhythm_metrics_thebeat(seq, file_dyad_records)
            stable_metrics = calculate_stable_subset_metrics(file_stable_dyad_records)

            # --- AUDACITY LABEL EXPORT ---
            if CREATE_AUDACITY_LABELS and len(clustered_onset_times) > 0:
                write_audacity_labels(
                    audio_folder,
                    filename,
                    clustered_onset_times,
                    refine_enabled=ONSET_REFINE_ENABLED,
                )

            # --- SPECTROGRAM VISUALIZER ---
            if CREATE_SPECTROGRAMS and len(clustered_onset_times) > 0:
                base_stem = os.path.splitext(filename)[0]
                onset_arr = np.array(clustered_onset_times)

                if SPECTROGRAM_CHUNK_ENABLED and SPECTROGRAM_CHUNK_SECONDS > 0:
                    chunk_dur = float(SPECTROGRAM_CHUNK_SECONDS)
                    n_chunks = max(1, int(np.ceil(total_duration / chunk_dur)))
                else:
                    chunk_dur = total_duration
                    n_chunks = 1

                for chunk_idx in range(n_chunks):
                    t_start = chunk_idx * chunk_dur
                    t_end = min(t_start + chunk_dur, total_duration)
                    s_start = int(t_start * sr)
                    s_end = int(t_end * sr)
                    y_chunk = y[s_start:s_end]

                    chunk_onsets = onset_arr[(onset_arr >= t_start) & (onset_arr < t_end)]

                    spectrogram = librosa.amplitude_to_db(np.abs(librosa.stft(y_chunk)), ref=np.max)
                    plt.figure(figsize=(10, 4))
                    librosa.display.specshow(spectrogram, x_axis="time", y_axis="log",
                                             sr=sr, x_coords=np.linspace(t_start, t_end,
                                             spectrogram.shape[1]))
                    plt.colorbar(format="%+2.0f dB")
                    if n_chunks == 1:
                        plt.title(f"Detected Onsets (thebeat): {filename}")
                        image_filename = f"{base_stem}_plot.png"
                    else:
                        plt.title(f"Detected Onsets (thebeat): {filename}  [{t_start:.1f}\u2013{t_end:.1f} s]")
                        image_filename = f"{base_stem}_{chunk_idx + 1}.png"
                    plt.vlines(
                        chunk_onsets,
                        ymin=0,
                        ymax=sr / 2,
                        color="r",
                        alpha=0.8,
                        linestyle="--",
                        label="Onsets",
                    )
                    image_path = os.path.join(audio_folder, image_filename)
                    save_matplotlib_figure(plt, image_path)
                    plt.close()

            # --- SAVE FILE SUMMARY DATA ---
            file_summary_data.append(
                {
                    "File Name": filename,
                    "Total Duration (s)": round(total_duration, 3),
                    "Estimated Overall BPM": round(estimated_bpm, 1),
                    "Average Cycle Duration (ms)": round(metrics["Average Cycle Duration (ms)"], 2)
                    if metrics["Average Cycle Duration (ms)"] is not None
                    else "N/A",
                    "Avg Loudness (Energy)": round(avg_loudness, 4),
                    "Avg Brightness (Centroid Hz)": round(avg_brightness, 2),
                    "Total Onsets Found (Raw)": len(onset_times),
                    "Total Onsets Used": len(clustered_onset_times),
                    "Onsets Merged by Clustering": max(len(onset_times) - len(clustered_onset_times), 0),
                    "High-Pass Filter Hz": HIGHPASS_CUTOFF_HZ if APPLY_HIGHPASS_FILTER else "Off",
                    "Amplitude Gate Threshold": ONSET_AMPLITUDE_GATE if ONSET_AMPLITUDE_GATE > 0 else "Off",
                    "Sharpness Gate Threshold": ONSET_SHARPNESS_GATE if ONSET_SHARPNESS_GATE > 0 else "Off",
                    "Min Inter-Onset (ms)": MIN_INTER_ONSET_MS if MIN_INTER_ONSET_MS > 0 else "Off",
                    "Onset Delta": ONSET_DELTA if ONSET_METHOD in ("librosa", "superflux") else "N/A",
                    "Onset Method": ONSET_METHOD,
                    "Active Preset": ACTIVE_PRESET if ACTIVE_PRESET else "None",
                    "Extractor Engine": "thebeat",
                    "Onset Refinement": f"±{ONSET_REFINE_WINDOW_MS}ms (sample-level)" if ONSET_REFINE_ENABLED else "Off",
                    "Stable Dyads Retained": len(file_stable_dyad_records),
                    "Stable Rhythm Filter Enabled": FILTER_STABLE_RHYTHMS,
                    "nPVI (Isochrony)": round(metrics["nPVI (Isochrony)"], 2)
                    if metrics["nPVI (Isochrony)"] is not None
                    else "N/A",
                    "CV of Intervals": round(metrics["CV of Intervals"], 4)
                    if metrics["CV of Intervals"] is not None
                    else "N/A",
                    "r_k Std Dev": round(metrics["r_k Std Dev"], 4)
                    if metrics["r_k Std Dev"] is not None
                    else "N/A",
                    "r_k Entropy (Categorical Measure)": round(metrics["r_k Entropy (Categorical Measure)"], 4)
                    if metrics["r_k Entropy (Categorical Measure)"] is not None
                    else "N/A",
                    "Stable Rhythm nPVI": round(stable_metrics["Stable Rhythm nPVI"], 2)
                    if stable_metrics["Stable Rhythm nPVI"] is not None
                    else "N/A",
                    "Stable Rhythm CV": round(stable_metrics["Stable Rhythm CV"], 4)
                    if stable_metrics["Stable Rhythm CV"] is not None
                    else "N/A",
                    "Stable Rhythm r_k Std Dev": round(stable_metrics["Stable Rhythm r_k Std Dev"], 4)
                    if stable_metrics["Stable Rhythm r_k Std Dev"] is not None
                    else "N/A",
                    "Stable Rhythm Entropy": round(stable_metrics["Stable Rhythm Entropy"], 4)
                    if stable_metrics["Stable Rhythm Entropy"] is not None
                    else "N/A",
                    "Exact Onset Times Used (s)": ", ".join(
                        [str(round(onset_time, 6 if ONSET_REFINE_ENABLED else 3)) for onset_time in clustered_onset_times]
                    ),
                }
            )

        except Exception as e:
            print(f"  -> ERROR processing {filename}: {e} — skipping this file")

    # ==========================================
    # 3. EXPORT EXCEL FILE WITH MULTIPLE TABS
    # ==========================================
    df_summary = pd.DataFrame(file_summary_data)
    df_dyads = pd.DataFrame(dyadic_events_data)
    df_stable_dyads = pd.DataFrame(stable_dyadic_events_data)

    # Ensure the output directory exists (handles any custom path from the GUI).
    _output_dir = os.path.dirname(output_excel_path)
    if _output_dir and not os.path.isdir(_output_dir):
        os.makedirs(_output_dir, exist_ok=True)

    workbook_sheets = {
        "File Summaries": df_summary,
        "Dyadic Events (For Plots)": df_dyads,
    }
    if FILTER_STABLE_RHYTHMS:
        workbook_sheets["Dyadic Events (Stable Rhythms)"] = df_stable_dyads
    write_workbook_sheets(output_excel_path, workbook_sheets)

    # --- Add explanatory comments to column headers ---
    if ADD_COLUMN_COMMENTS:
        from openpyxl import load_workbook
        from openpyxl.comments import Comment

        _SUMMARY_COMMENTS = {
            "File Name": "The name of the audio file that was analysed.",
            "Total Duration (s)": "Total length of the audio file in seconds.",
            "Estimated Overall BPM": "Estimated beats per minute, calculated from the median inter-onset interval across the entire file.",
            "Average Cycle Duration (ms)": "Mean duration of one full rhythmic cycle (two consecutive intervals combined) in milliseconds.",
            "Avg Loudness (Energy)": "Average RMS energy of the audio signal across all frames — a measure of overall volume/loudness.",
            "Avg Brightness (Centroid Hz)": "Average spectral centroid in Hz — higher values mean the sound energy is concentrated at higher frequencies (brighter timbre).",
            "Total Onsets Found (Raw)": "Number of onset events initially detected before any clustering or merging was applied.",
            "Total Onsets Used": "Number of onsets remaining after clustering and all gating filters. These are the onsets used for rhythm analysis.",
            "Onsets Merged by Clustering": "Number of raw onsets that were merged into neighbouring onsets because they fell within the cluster window.",
            "High-Pass Filter Hz": "High-pass filter cutoff frequency used during onset detection (if enabled). 'Off' means no high-pass was applied at this stage.",
            "Amplitude Gate Threshold": "Fraction of peak RMS energy an onset must exceed to be kept. 'Off' means no amplitude gating was applied.",
            "Sharpness Gate Threshold": "Fraction of the steepest attack slope an onset must exceed to be kept. 'Off' means no sharpness gating was applied.",
            "Min Inter-Onset (ms)": "Minimum allowed gap in milliseconds between consecutive onsets. Closer onsets are removed. 'Off' means no minimum was enforced.",
            "Onset Delta": "Peak-picking sensitivity threshold used by the librosa/superflux onset methods. 'N/A' for methods that don't use this parameter.",
            "Onset Method": "Which onset detection algorithm was used: adaptive_hp, librosa, moving_median, superflux, cfar, or per_band.",
            "Active Preset": "Name of the detection preset applied (e.g. 'birdsong', 'percussion_clean'), or 'None' if parameters were set manually.",
            "Extractor Engine": "Which extraction engine was used: 'standard' (full-featured) or 'thebeat' (thebeat package).",
            "Onset Refinement": "Whether sample-level Hilbert-envelope refinement was applied to improve onset timing precision, and the search window used.",
            "Stable Dyads Retained": "Number of dyadic rhythm events that passed the stable-rhythm consistency filter for this file.",
            "Stable Rhythm Filter Enabled": "Whether the stable-rhythm filter was turned on (True) or off (False) for this analysis run.",
            "nPVI (Isochrony)": "Normalised Pairwise Variability Index — measures how much consecutive intervals differ. 0 = perfectly isochronous (metronomic), higher = more variable timing.",
            "CV of Intervals": "Coefficient of Variation of inter-onset intervals (standard deviation / mean). Lower values indicate more regular timing.",
            "r_k Std Dev": "Standard deviation of the rhythm ratio (r_k) values across all dyads. Lower values mean more consistent rhythm ratios.",
            "r_k Entropy (Categorical Measure)": "Shannon entropy of the rhythm ratio distribution. Higher values indicate a wider spread of rhythm types; lower values indicate clustering around specific ratios.",
            "Stable Rhythm nPVI": "nPVI computed using only the stable-rhythm dyads (those passing the consistency filter). 'N/A' if no stable dyads exist.",
            "Stable Rhythm CV": "CV of intervals computed using only the stable-rhythm subset. 'N/A' if no stable dyads exist.",
            "Stable Rhythm r_k Std Dev": "Standard deviation of r_k for stable dyads only. 'N/A' if no stable dyads exist.",
            "Stable Rhythm Entropy": "Shannon entropy of r_k for stable dyads only. 'N/A' if no stable dyads exist.",
            "Exact Onset Times Used (s)": "Comma-separated list of every onset time (in seconds from file start) that was used for rhythm analysis, after all filtering and clustering.",
        }

        _DYAD_COMMENTS = {
            "File Name": "The audio file this dyadic event came from.",
            "Dyad Index": "Sequential index of this dyad within the file (1 = first pair of consecutive intervals).",
            "Interval 1 (ms)": "Duration in milliseconds of the first interval in this consecutive pair.",
            "Interval 2 (ms)": "Duration in milliseconds of the second interval in this consecutive pair.",
            "Cycle Duration [cd] (ms)": "Sum of both intervals (Interval 1 + Interval 2) — the total duration of one rhythmic cycle in milliseconds.",
            "Short Interval [i_s] (ms)": "The shorter of the two intervals in this dyad, in milliseconds.",
            "Long Interval [i_l] (ms)": "The longer of the two intervals in this dyad, in milliseconds.",
            "Rhythm Ratio [r_k]": "Interval 1 divided by the Cycle Duration (i₁ / cd). Ranges from 0 to 1. A value of 0.5 means both intervals are equal (isochronous). Values below 0.5 mean the first interval is shorter; above 0.5 means it is longer.",
            "Stable Rhythm": "Whether this dyad passed the stable-rhythm consistency filter (True/False). True means this rhythm pattern was repeated consistently across neighbouring cycles.",
        }

        _comment_author = "Bioacoustics Pipeline"
        wb = load_workbook(output_excel_path)
        for sheet_name, col_comments in [
            ("File Summaries", _SUMMARY_COMMENTS),
            ("Dyadic Events (For Plots)", _DYAD_COMMENTS),
            ("Dyadic Events (Stable Rhythms)", _DYAD_COMMENTS),
        ]:
            if sheet_name not in wb.sheetnames:
                continue
            ws = wb[sheet_name]
            for col_idx, cell in enumerate(ws[1], start=1):
                header_text = cell.value
                if header_text and header_text in col_comments:
                    cell.comment = Comment(col_comments[header_text], _comment_author)
        save_openpyxl_workbook(wb, output_excel_path)
        print("Column header comments added to Excel file.")

    print(f"\nSUCCESS! Processed {len(file_summary_data)} files using 'thebeat' engine.")
    print(f"Extracted {len(dyadic_events_data)} total dyadic rhythms.")
    if FILTER_STABLE_RHYTHMS:
        print(f"Retained {len(stable_dyadic_events_data)} dyadic rhythms in the stable-rhythm sheet.")


if __name__ == "__main__":
    main()