from __future__ import annotations

import datetime
import html
import json
import os
import sys
from copy import deepcopy

import numpy as np
from PyQt6.QtCore import Qt, QTimer
from PyQt6.QtGui import QColor
from PyQt6.QtWidgets import (
    QApplication,
    QCheckBox,
    QDialog,
    QDialogButtonBox,
    QFileDialog,
    QGroupBox,
    QHBoxLayout,
    QInputDialog,
    QLabel,
    QMessageBox,
    QProgressBar,
    QPushButton,
    QScrollArea,
    QTableWidget,
    QTableWidgetItem,
    QTextBrowser,
    QVBoxLayout,
    QWidget,
)

try:
    from form_widgets import (
        AUDIO_EXTENSIONS,
        DescriptionLabel,
        FilePicker,
        FolderPicker,
        _ALL_DESC_LABELS,
        _CheckableAudioFileCombo,
        _add_row,
        _list_audio_files_in_folder,
        get_form_widget_palette,
    )
except ImportError:
    from GUI.form_widgets import (
        AUDIO_EXTENSIONS,
        DescriptionLabel,
        FilePicker,
        FolderPicker,
        _ALL_DESC_LABELS,
        _CheckableAudioFileCombo,
        _add_row,
        _list_audio_files_in_folder,
        get_form_widget_palette,
    )

try:
    from per_file_settings_support import PerFileSettingsDialog
except ImportError:
    from GUI.per_file_settings_support import PerFileSettingsDialog

try:
    from pipeline_prep_support import (
        AnalysisHintsDialog,
        _FileMetadataSection,
        _make_title_info_button,
        _prep_apply_hint_overrides_to_muter,
        _prep_apply_hints_to_profile,
        _prep_default_hints,
        _prep_float_or_none,
        _prep_infer_signal_character,
        _prep_load_signal_profile_hint,
        _prep_recommend_onset_settings,
        _prep_sanitize_hints,
    )
except ImportError:
    from GUI.pipeline_prep_support import (
        AnalysisHintsDialog,
        _FileMetadataSection,
        _make_title_info_button,
        _prep_apply_hint_overrides_to_muter,
        _prep_apply_hints_to_profile,
        _prep_default_hints,
        _prep_float_or_none,
        _prep_infer_signal_character,
        _prep_load_signal_profile_hint,
        _prep_recommend_onset_settings,
        _prep_sanitize_hints,
    )

try:
    from onset_editor_dialogs import _ExcelColumnDialog
    from onset_editor_io import _HAS_EXCEL_IO, _eio
except ImportError:
    from GUI.onset_editor_dialogs import _ExcelColumnDialog
    from GUI.onset_editor_io import _HAS_EXCEL_IO, _eio


_DESC_LEVEL = 0

_ACCENT = "#4caf50"
_ACCENT_DIM = "#2e7d32"
_BG = "#1e1e2e"
_BG_MID = "#262636"
_BG_WIDGET = "#2c2c3c"
_BG_INPUT = "#323248"
_BORDER = "#3a3a50"
_TEXT = "#dcdcdc"
_TEXT_DIM = "#8888a0"


def _sync_theme_aliases() -> None:
    global _ACCENT, _ACCENT_DIM, _BG, _BG_MID, _BG_WIDGET
    global _BG_INPUT, _BORDER, _TEXT, _TEXT_DIM
    palette = get_form_widget_palette()
    _ACCENT = palette.accent
    _ACCENT_DIM = palette.accent_dim
    _BG = palette.bg
    _BG_MID = palette.bg_mid
    _BG_WIDGET = palette.bg_widget
    _BG_INPUT = palette.bg_input
    _BORDER = palette.border
    _TEXT = palette.text
    _TEXT_DIM = palette.text_dim


_sync_theme_aliases()


class _OutsideOnsetImportDialog(QDialog):
    """Preview matching outside-onset rows before importing them."""

    def __init__(
        self,
        *,
        source_path: str,
        target_excel_path: str,
        matches: list[dict],
        unmatched_source_names: list[str],
        duplicate_targets: list[dict],
        parent=None,
    ):
        super().__init__(parent)
        self.setWindowTitle("Import Outside Onsets")
        self.setMinimumWidth(700)
        self.resize(760, 540)
        self.setStyleSheet(
            f"QDialog {{ background: {_BG}; color: {_TEXT}; }}"
            f"QLabel {{ color: {_TEXT}; background: transparent; }}"
            f"QCheckBox {{ color: {_TEXT}; }}"
            f"QTextBrowser {{ background: {_BG_WIDGET}; color: {_TEXT}; "
            f"border: 1px solid {_BORDER}; border-radius: 6px; padding: 8px; }}"
            f"QPushButton {{ background: {_BG_WIDGET}; color: {_TEXT}; "
            f"border: 1px solid {_BORDER}; border-radius: 4px; padding: 5px 18px; }}"
            f"QPushButton:hover {{ border-color: {_ACCENT}; color: {_ACCENT}; }}"
        )

        layout = QVBoxLayout(self)
        layout.setSpacing(10)
        layout.setContentsMargins(14, 14, 14, 14)

        summary = QLabel(
            f"<b>Source:</b> {html.escape(os.path.basename(source_path))}<br>"
            f"<b>Target workbook:</b> {html.escape(os.path.basename(target_excel_path) or target_excel_path)}<br>"
            f"<b>Matched files:</b> {len(matches)}"
        )
        summary.setStyleSheet(f"color: {_TEXT_DIM}; font-size: 11px;")
        layout.addWidget(summary)

        browser = QTextBrowser()
        browser.setOpenExternalLinks(False)
        browser.setHtml(self._build_html(matches, unmatched_source_names, duplicate_targets))
        layout.addWidget(browser, stretch=1)

        self._overwrite_cb = QCheckBox("Overwrite existing onset values in the target workbook")
        self._overwrite_cb.setChecked(True)
        self._overwrite_cb.setToolTip(
            "When off, files that already have onset values in the target workbook are skipped."
        )
        layout.addWidget(self._overwrite_cb)

        buttons = QDialogButtonBox(
            QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel
        )
        buttons.button(QDialogButtonBox.StandardButton.Ok).setText("Add Onsets")
        buttons.accepted.connect(self.accept)
        buttons.rejected.connect(self.reject)
        layout.addWidget(buttons)

    @staticmethod
    def _build_html(matches: list[dict], unmatched_source_names: list[str], duplicate_targets: list[dict]) -> str:
        rows = []
        for match in matches[:30]:
            current_label = (
                f"{match['current_onset_count']} existing"
                if match.get("current_onset_count")
                else ("row exists" if match.get("target_row_exists") else "new row")
            )
            rows.append(
                "<tr>"
                f"<td>{html.escape(match['source_name'])}</td>"
                f"<td>{html.escape(match['audio_filename'])}</td>"
                f"<td>{int(match['onset_count'])}</td>"
                f"<td>{html.escape(current_label)}</td>"
                "</tr>"
            )

        more_rows = ""
        if len(matches) > 30:
            more_rows = f"<p><i>Showing the first 30 of {len(matches)} matched files.</i></p>"

        unmatched_html = ""
        if unmatched_source_names:
            preview = ", ".join(html.escape(name) for name in unmatched_source_names[:8])
            suffix = " ..." if len(unmatched_source_names) > 8 else ""
            unmatched_html = (
                f"<p><b>Unmatched source rows:</b> {len(unmatched_source_names)}"
                f"<br><span style='color:{_TEXT_DIM}'>{preview}{suffix}</span></p>"
            )

        duplicate_html = ""
        if duplicate_targets:
            preview = ", ".join(
                html.escape(entry["source_name"]) for entry in duplicate_targets[:8]
            )
            suffix = " ..." if len(duplicate_targets) > 8 else ""
            duplicate_html = (
                f"<p><b>Duplicate source rows skipped:</b> {len(duplicate_targets)}"
                f"<br><span style='color:{_TEXT_DIM}'>{preview}{suffix}</span></p>"
            )

        return (
            "<h3>Matching Files</h3>"
            "<table cellspacing='0' cellpadding='4' border='1' style='border-color:#3a3a50; border-collapse:collapse;'>"
            "<tr><th>Source Row</th><th>Current Audio File</th><th>Incoming Onsets</th><th>Current Workbook</th></tr>"
            + "".join(rows)
            + "</table>"
            + more_rows
            + unmatched_html
            + duplicate_html
        )

    def overwrite_existing(self) -> bool:
        return self._overwrite_cb.isChecked()


_METADATA_FIELDS: list[tuple[str, str]] = [
    ("Group",
     "Group label used by nearly every experimental test "
     "(e.g. species, population, condition)."),
    ("Species",
     "Species name (tree tip label) - used by Mantel and PGLS."),
    ("Latitude",
     "Geographic latitude in decimal degrees - Mantel geographic matrix."),
    ("Longitude",
     "Geographic longitude in decimal degrees - Mantel geographic matrix."),
    ("Modality",
     "Signal modality (e.g. 'vocal', 'percussive') - GLMM fixed effect."),
    ("Function",
     "Signal function (e.g. 'territorial', 'courtship') - GLMM fixed effect."),
    ("Tempo_BPM",
     "Preferred tempo in beats per minute - GLMM / PGLS predictor."),
    ("BodyMass_kg",
     "Adult body mass in kilograms - GLMM / PGLS predictor."),
]


class PipelinePrepPanel(QScrollArea):
    """Step 0: scan a folder for audio files and show what pipeline
    artefacts already exist for each file."""

    _COL_FILE = 0
    _COL_LAYERS = 1
    _COL_FOCUS = 2
    _COL_SIGNALS = 3
    _COL_EDITED = 4
    _COL_EXCEL = 5
    _COL_LABELS = 6
    _COL_PREANALYSIS = 7

    _COL_PATH_LABELS = {
        1: "Onset Layers folder pattern",
        2: "Focus Regions file pattern",
        3: "Selected Signals folder pattern",
        4: "Edited Output folder",
        5: "Excel workbook",
        6: "Labels file pattern",
        7: "Pre-Analysis JSON",
    }

    def __init__(self, parent=None):
        _sync_theme_aliases()
        super().__init__(parent)
        self._global_analysis_hints = _prep_default_hints()
        self._file_analysis_hints = {}
        self._last_analysis_rows = []
        self._inv_column_paths = {}

        self.setWidgetResizable(True)
        content = QWidget()
        lay = QVBoxLayout(content)
        lay.setSpacing(14)
        lay.setContentsMargins(16, 16, 16, 16)

        grp_io = QGroupBox("Input / Output")
        g_io = QVBoxLayout(grp_io)
        g_io.setSpacing(8)
        g_io.setContentsMargins(10, 12, 10, 10)

        self.input_folder = FolderPicker("")
        self.input_folder.line_edit.setPlaceholderText(
            "Main folder containing audio files to analyse")
        _add_row(g_io, "Audio input folder", self.input_folder,
                 "Root folder of raw audio files for the pipeline.",
                 extended_desc="The master folder that contains every audio "
                 "recording you want the pipeline to work with. Pipeline Prep "
                 "scans this folder (and optionally its 'data/' subfolder) for "
                 "audio files plus any pre-existing sidecar artefacts "
                 "(Onset Layers, Focus Regions, Selected Signals, Labels, "
                 "Excel workbooks, pre-analysis JSONs). All downstream steps "
                 "(Audio Editor, Onset Finder, etc.) auto-fill their own input "
                 "paths from this folder when their 'Auto' checkboxes are on.",
                 label_width=170)

        spec_row = QHBoxLayout()
        self.specify_files_cb = QCheckBox("Specify files:")
        self.specify_files_cb.setChecked(False)
        self.specify_files_cb.setToolTip(
            "When checked, only listed files are analysed. "
            "When unchecked, all audio files in the folder are used.")
        spec_row.addWidget(self.specify_files_cb)
        self.selected_files_combo = _CheckableAudioFileCombo()
        self.selected_files_combo.setEnabled(False)
        spec_row.addWidget(self.selected_files_combo, stretch=1)
        g_io.addLayout(spec_row)
        _spec_desc = DescriptionLabel(
            "When checked, only listed files are analysed. "
            "When unchecked, all audio files in the folder are used.",
            "Restrict the pipeline to a hand-picked subset of files in the "
            "input folder. Turn this on when you want to debug settings on one "
            "or two recordings before running the full batch, or when only "
            "some files in a shared folder belong to the current analysis.",
            "Normally the pipeline looks at every recording in the folder. "
            "Tick this box to instead pick just a few files from the list - "
            "like choosing specific photos from a camera roll instead of "
            "printing the whole roll.")
        g_io.addWidget(_spec_desc)
        _ALL_DESC_LABELS.append(_spec_desc)
        self.specify_files_cb.stateChanged.connect(self._on_specify_toggled)

        self.excel_path = FilePicker("", "Excel files (*.xlsx *.xls)")
        self.excel_path.line_edit.setPlaceholderText(
            "Excel workbook from Onset Finder (optional)")
        _add_row(g_io, "Excel file", self.excel_path,
                 "Onset Finder output Excel. Used to check which audio files "
                 "already have rhythm data.",
                 extended_desc="Optional - point this at an existing "
                 "Cross_Species_Rhythm_Data.xlsx produced by a previous "
                 "Onset Finder run. Pipeline Prep uses it to flag which audio "
                 "files have already been analysed (so you can skip or re-run "
                 "them) and to auto-fill the 'Input Excel' field on later steps "
                 "(Flower Raster Plots, Histograms, nPVI-by-Group). Leave "
                 "blank the very first time you run a folder.",
                 label_width=170)

        outside_row = QHBoxLayout()
        outside_row.addSpacing(170)
        self.load_outside_onsets_btn = QPushButton("Load Outside Onsets")
        self.load_outside_onsets_btn.setToolTip(
            "Import onset times from another CSV/Excel file into the current workbook."
        )
        self.load_outside_onsets_btn.setEnabled(_HAS_EXCEL_IO)
        self.load_outside_onsets_btn.clicked.connect(self._load_outside_onsets)
        outside_row.addWidget(self.load_outside_onsets_btn, alignment=Qt.AlignmentFlag.AlignLeft)
        outside_row.addStretch(1)
        g_io.addLayout(outside_row)
        if not _HAS_EXCEL_IO:
            outside_desc = DescriptionLabel(
                "External onset import is unavailable because spreadsheet support could not be loaded.",
                "This button depends on the shared Excel/CSV onset I/O helpers used by the Onset Editor.",
                "If the button is disabled, the app could not import the spreadsheet libraries it needs.",
            )
        else:
            outside_desc = DescriptionLabel(
                "Bring onset timestamps from another CSV/Excel into the workbook above.",
                "Use this when another detector already produced onset times. Pipeline Prep will preview which source rows match audio files in the current folder, then write those onsets into the current workbook and recompute the rhythm-summary, raw-dyad, and stable-dyad sheets for each matched file.",
                "This lets you keep the pipeline workbook format even when the onset times came from another program.",
            )
        g_io.addWidget(outside_desc)
        _ALL_DESC_LABELS.append(outside_desc)

        lay.addWidget(grp_io)

        grp_inv = QGroupBox("File Inventory")
        g_inv = QVBoxLayout(grp_inv)
        g_inv.setSpacing(8)
        g_inv.setContentsMargins(10, 12, 10, 10)

        self._inv_help_btn = _make_title_info_button(
            grp_inv,
            tooltip=("Explains what the File Inventory shows, how each column "
                     "is detected, and why it matters for Pre-Analysis."),
            callback=self._show_file_inventory_help,
        )

        scan_row = QHBoxLayout()
        self._scan_btn = QPushButton("  Scan Folder  ")
        self._scan_btn.setToolTip(
            "Scan the input folder for audio files and check what pipeline "
            "outputs already exist for each one.")
        self._scan_btn.setStyleSheet(
            f"QPushButton {{ background-color: {_ACCENT_DIM}; color: white; "
            f"border-radius: 5px; padding: 5px 14px; font-size: 12px; border: none; }} "
            f"QPushButton:hover {{ background-color: {_ACCENT}; }}"
        )
        self._scan_btn.clicked.connect(self._scan_folder)
        scan_row.addWidget(self._scan_btn)
        self._scan_status = QLabel("")
        self._scan_status.setStyleSheet(
            f"color: {_TEXT_DIM}; font-size: 11px; background: transparent;")
        scan_row.addWidget(self._scan_status, stretch=1)
        g_inv.addLayout(scan_row)

        self._inv_table = QTableWidget(0, 8)
        self._inv_table.setHorizontalHeaderLabels([
            "Audio File", "Onset Layers 📁", "Focus Regions 📁",
            "Sel. Signals 📁", "Edited Output 📁", "In Excel 📁",
            "Labels 📁", "Pre-Analysis 📁",
        ])
        self._inv_table.horizontalHeader().setStretchLastSection(True)
        self._inv_table.horizontalHeader().setSectionResizeMode(
            0, self._inv_table.horizontalHeader().ResizeMode.Stretch)
        for col in range(1, 8):
            self._inv_table.horizontalHeader().setSectionResizeMode(
                col, self._inv_table.horizontalHeader().ResizeMode.ResizeToContents)
        self._inv_table.setEditTriggers(
            QTableWidget.EditTrigger.NoEditTriggers)
        self._inv_table.setSelectionBehavior(
            QTableWidget.SelectionBehavior.SelectRows)
        self._inv_table.setAlternatingRowColors(True)
        self._inv_table.setMinimumHeight(220)
        self._inv_table.setStyleSheet(
            f"QTableWidget {{ background-color: {_BG_WIDGET}; color: {_TEXT}; "
            f"gridline-color: {_BORDER}; font-size: 12px; "
            f"alternate-background-color: {_BG_INPUT}; }}"
            f"QHeaderView::section {{ background-color: {_BG_MID}; color: {_TEXT}; "
            f"border: 1px solid {_BORDER}; padding: 4px; font-weight: bold; "
            f"font-size: 11px; }}"
        )
        g_inv.addWidget(self._inv_table)

        self._inv_table.horizontalHeader().sectionClicked.connect(
            self._on_inv_header_clicked)
        for col, tip in {
            1: "Click to view/change where Onset Layers are looked up",
            2: "Click to view/change where Focus Regions are looked up",
            3: "Click to view/change where Selected Signals are looked up",
            4: "Click to view/change where Edited Output is looked up",
            5: "Click to view/change the Excel file used for lookups",
            6: "Click to view/change where Labels are looked up",
            7: "Click to view/change the Pre-Analysis JSON path",
        }.items():
            item = self._inv_table.horizontalHeaderItem(col)
            if item:
                item.setToolTip(tip)

        lay.addWidget(grp_inv)

        self._metadata_section = _FileMetadataSection(
            self,
            metadata_fields=_METADATA_FIELDS,
            desc_level=_DESC_LEVEL,
        )
        lay.addWidget(self._metadata_section)

        grp_an = QGroupBox("Pre-Analyze Audio Files")
        g_an = QVBoxLayout(grp_an)
        g_an.setSpacing(6)
        g_an.setContentsMargins(10, 12, 10, 10)

        self._an_help_btn = _make_title_info_button(
            grp_an,
            tooltip=("Explains exactly what data Pre-Analyze reads, how it "
                     "converts that into Audio Editor and Onset Finder "
                     "recommendations, and why it is worth running."),
            callback=self._show_pre_analyze_help,
        )

        hint_row = QHBoxLayout()
        hint_row.setSpacing(8)
        self._edit_all_hints_btn = QPushButton("Edit All-File Hints")
        self._edit_all_hints_btn.setToolTip(
            "Provide optional hints that should apply to all files during pre-analysis.")
        self._edit_all_hints_btn.clicked.connect(self._edit_all_hints)
        hint_row.addWidget(self._edit_all_hints_btn)

        self._edit_file_hints_btn = QPushButton("Edit One-File Hints")
        self._edit_file_hints_btn.setToolTip(
            "Provide optional hints for one specific file (overrides all-file hints).")
        self._edit_file_hints_btn.clicked.connect(self._edit_one_file_hints)
        hint_row.addWidget(self._edit_file_hints_btn)
        hint_row.addStretch(1)
        g_an.addLayout(hint_row)

        self._hints_summary = QLabel("")
        self._hints_summary.setStyleSheet(
            f"color: {_TEXT_DIM}; font-size: 11px; background: transparent;")
        g_an.addWidget(self._hints_summary)
        g_an.addSpacing(4)

        run_row = QHBoxLayout()
        run_row.setSpacing(6)
        self._pre_analyze_btn = QPushButton("  Pre-Analyze Files  ")
        self._pre_analyze_btn.setToolTip(
            "Analyze each audio file and recommend Audio Editor and Onset Finder settings.")
        self._pre_analyze_btn.setStyleSheet(
            f"QPushButton {{ background-color: {_ACCENT_DIM}; color: white; "
            f"border-radius: 5px; padding: 5px 14px; font-size: 12px; border: none; }} "
            f"QPushButton:hover {{ background-color: {_ACCENT}; }}"
        )
        self._pre_analyze_btn.clicked.connect(self._run_pre_analysis)
        run_row.addWidget(self._pre_analyze_btn)
        run_row.addSpacing(12)

        self._use_signal_hints_cb = QCheckBox("Use signal hints")
        self._use_signal_hints_cb.setChecked(True)
        self._use_signal_hints_cb.setToolTip(
            "Use Selected Signals / Focus Regions when available.")
        run_row.addWidget(self._use_signal_hints_cb)
        run_row.addSpacing(10)

        self._apply_muter_cb = QCheckBox("Apply to Audio Editor")
        self._apply_muter_cb.setChecked(True)
        self._apply_muter_cb.setToolTip(
            "Apply representative recommendations to the Audio Editor panel.")
        run_row.addWidget(self._apply_muter_cb)
        run_row.addSpacing(10)

        self._apply_extractor_cb = QCheckBox("Apply to Onset Finder")
        self._apply_extractor_cb.setChecked(True)
        self._apply_extractor_cb.setToolTip(
            "Apply representative recommendations to the Onset Finder panel.")
        run_row.addWidget(self._apply_extractor_cb)
        run_row.addStretch(1)

        self._view_per_file_btn = QPushButton("  View / Edit Per-File Settings  ")
        self._view_per_file_btn.setToolTip(
            "Open a table of every per-file Audio Editor and Onset Finder "
            "setting, with per-setting toggles to use the General (panel) "
            "value or the Specific (per-file) value that Pre-Analyze chose.")
        self._view_per_file_btn.setStyleSheet(
            f"QPushButton {{ background-color: {_BG_WIDGET}; color: {_TEXT}; "
            f"border: 1px solid {_BORDER}; border-radius: 5px; "
            f"padding: 5px 12px; font-size: 12px; }} "
            f"QPushButton:hover {{ border-color: {_ACCENT}; color: {_ACCENT}; }}"
        )
        self._view_per_file_btn.clicked.connect(self._open_per_file_settings_dialog)
        run_row.addWidget(self._view_per_file_btn)
        g_an.addLayout(run_row)

        for _tip, _det, _nov in [
            ("Use Selected Signals / Focus Regions when available.",
             "When turned on, pre-analysis uses any Selected Signals or Focus "
             "Regions you drew in the Onset Editor as strong cues for "
             "recommending filter bands, amplitude gates, and preset choices. "
             "If no signal/focus data exists for a file, analysis falls back "
             "to purely acoustic metrics.",
             "If you've already marked 'good sounds' or 'zones to focus on' "
             "in the Onset Editor, this tells the pre-analysis to take those "
             "hints into account when picking settings - like giving a chef "
             "your taste preferences before they cook."),
            ("Apply representative recommendations to the Audio Editor panel.",
             "After pre-analysis finishes, copy the aggregated (corpus-median) "
             "recommendations for high-pass cutoff, amplitude threshold, "
             "denoising strength, etc. into the Step 1 Audio Editor panel. "
             "Per-file recommendations are always saved to the JSON regardless; "
             "this switch only controls whether the GUI panel itself is "
             "updated with representative defaults.",
             "When pre-analysis finishes, auto-fill the Audio Editor panel "
             "(Step 1) with the 'average' recommended settings for your "
             "recordings - so you don't have to copy every number yourself."),
            ("Apply representative recommendations to the Onset Finder panel.",
             "Same as 'Apply to Audio Editor', but for the Step 2 Onset Finder "
             "panel. The aggregated recommended onset method, minimum "
             "inter-onset interval, amplitude gate, etc. are written into the "
             "Onset Finder panel so you can review/tweak before running.",
             "When pre-analysis finishes, auto-fill the Onset Finder panel "
             "(Step 2) with the 'average' recommended beat-detection settings "
             "for your recordings."),
        ]:
            _lbl = DescriptionLabel(_tip, _det, _nov)
            g_an.addWidget(_lbl)
            _ALL_DESC_LABELS.append(_lbl)

        import_row = QHBoxLayout()
        self._import_preanalysis_btn = QPushButton("Import Pre-Analysis JSON")
        self._import_preanalysis_btn.setToolTip(
            "Load a previously saved AudioEditor_PerFile_PreAnalysis.json "
            "instead of re-running pre-analysis.")
        self._import_preanalysis_btn.clicked.connect(self._import_pre_analysis)
        import_row.addWidget(self._import_preanalysis_btn)

        self._import_report_btn = QPushButton("Import Analysis Report")
        self._import_report_btn.setToolTip(
            "Load a previously saved Audio_PreAnalysis_Report.json (full report "
            "with metrics, recommendations, and per-file settings).")
        self._import_report_btn.clicked.connect(self._import_analysis_report)
        import_row.addWidget(self._import_report_btn)

        import_row.addStretch(1)
        g_an.addLayout(import_row)
        g_an.addSpacing(6)

        self.per_file_settings_path = FilePicker("", "JSON files (*.json)")
        self.per_file_settings_path.line_edit.setPlaceholderText(
            "Path to per-file Audio Editor settings JSON")
        _add_row(
            g_an,
            "Per-file audio settings",
            self.per_file_settings_path,
            "This JSON is consumed by scripts/main.py for per-file Audio Editor settings.",
            extended_desc="Sidecar JSON (AudioEditor_PerFile_PreAnalysis.json by "
            "default) that maps each audio filename to its own recommended "
            "Audio Editor settings (high-pass cutoff, denoising, amplitude "
            "threshold, etc.). When this path is present, scripts/main.py "
            "applies per-file overrides on top of the global Audio Editor "
            "panel values - so each recording gets parameters tuned to its "
            "own noise floor and spectral character.",
            label_width=170,
        )

        self.analysis_report_path = FilePicker("", "JSON files (*.json)")
        self.analysis_report_path.line_edit.setPlaceholderText(
            "Path to full pre-analysis report JSON")
        _add_row(
            g_an,
            "Analysis report",
            self.analysis_report_path,
            "Full report with analysis metrics, recommendations, and reasoning.",
            extended_desc="The complete pre-analysis report (Audio_PreAnalysis_"
            "Report.json by default). Unlike the lean per-file settings JSON, "
            "this report also records the raw acoustic metrics (SNR, noise "
            "floor, dominant frequencies, onset density, etc.) and the "
            "human-readable rationale for each recommendation. Useful for "
            "auditing why the pipeline chose specific parameters and for "
            "re-loading a prior analysis without re-running it.",
            label_width=170,
        )

        self._analysis_status = QLabel("")
        self._analysis_status.setStyleSheet(
            f"color: {_TEXT_DIM}; font-size: 11px; background: transparent;")
        g_an.addWidget(self._analysis_status)

        self._analysis_progress = QProgressBar()
        self._analysis_progress.setMinimum(0)
        self._analysis_progress.setMaximum(100)
        self._analysis_progress.setValue(0)
        self._analysis_progress.setTextVisible(True)
        self._analysis_progress.setFormat("%v / %m files")
        self._analysis_progress.setFixedHeight(18)
        self._analysis_progress.setStyleSheet(
            f"QProgressBar {{ border: 1px solid {_BORDER}; border-radius: 4px; "
            f"background-color: {_BG_WIDGET}; color: {_TEXT}; font-size: 11px; "
            f"text-align: center; }}"
            f"QProgressBar::chunk {{ background-color: {_ACCENT_DIM}; border-radius: 3px; }}"
        )
        self._analysis_progress.hide()
        g_an.addWidget(self._analysis_progress)

        self._analysis_table = QTableWidget(0, 9)
        self._analysis_table.setHorizontalHeaderLabels([
            "Audio File", "Profile", "SNR (dB)", "Noise (dB)",
            "Character", "Onset Method", "Min IOI (ms)", "HP / LP (Hz)", "Notes",
        ])
        self._analysis_table.horizontalHeader().setStretchLastSection(True)
        self._analysis_table.horizontalHeader().setSectionResizeMode(
            0, self._analysis_table.horizontalHeader().ResizeMode.Stretch)
        for col in range(1, 9):
            self._analysis_table.horizontalHeader().setSectionResizeMode(
                col, self._analysis_table.horizontalHeader().ResizeMode.ResizeToContents)
        self._analysis_table.setEditTriggers(
            QTableWidget.EditTrigger.NoEditTriggers)
        self._analysis_table.setSelectionBehavior(
            QTableWidget.SelectionBehavior.SelectRows)
        self._analysis_table.setAlternatingRowColors(True)
        self._analysis_table.setMinimumHeight(220)
        self._analysis_table.setStyleSheet(
            f"QTableWidget {{ background-color: {_BG_WIDGET}; color: {_TEXT}; "
            f"gridline-color: {_BORDER}; font-size: 12px; "
            f"alternate-background-color: {_BG_INPUT}; }}"
            f"QHeaderView::section {{ background-color: {_BG_MID}; color: {_TEXT}; "
            f"border: 1px solid {_BORDER}; padding: 4px; font-weight: bold; "
            f"font-size: 11px; }}"
        )
        g_an.addWidget(self._analysis_table)

        lay.addWidget(grp_an)

        self.input_folder.textChanged.connect(self._on_input_changed)

        self._update_hints_summary()
        lay.addStretch()
        self.setWidget(content)

    def _on_specify_toggled(self, state):
        on = self.specify_files_cb.isChecked()
        self.selected_files_combo.setEnabled(on)
        if on:
            folder = self.input_folder.text().strip()
            if folder:
                self.selected_files_combo.set_folder(folder)
                self.selected_files_combo.ensure_first_selected()

    def _on_input_changed(self, text):
        folder = (text or "").strip()
        if os.path.isdir(folder):
            self.selected_files_combo.set_folder(folder)
            self._sync_default_analysis_paths(folder, overwrite=True)
            self._auto_detect_excel(folder)
            self._auto_detect_column_paths(folder)
            QTimer.singleShot(200, self._scan_folder)
        else:
            self._analysis_table.setRowCount(0)
            self._analysis_status.setText("")
            self._inv_column_paths.clear()

    def _auto_detect_excel(self, folder: str):
        candidates = [
            os.path.join(folder, "Cross_Species_Rhythm_Data.xlsx"),
            os.path.join(folder, "data", "Cross_Species_Rhythm_Data.xlsx"),
        ]
        candidates += [
            os.path.join(folder, "AudioData_OnsetFinder.xlsx"),
            os.path.join(folder, "data", "AudioData_OnsetFinder.xlsx"),
        ]
        for candidate in candidates:
            if os.path.isfile(candidate):
                self.excel_path.setText(candidate)
                return
        for directory in [folder, os.path.join(folder, "data")]:
            if os.path.isdir(directory):
                for filename in sorted(os.listdir(directory)):
                    if filename.lower().endswith(".xlsx"):
                        self.excel_path.setText(os.path.join(directory, filename))
                        return

    def _auto_detect_column_paths(self, folder: str):
        self._inv_column_paths.clear()
        data_dir = os.path.join(folder, "data")
        out_dir = data_dir if os.path.isdir(data_dir) else folder

        self._inv_column_paths[self._COL_LAYERS] = folder
        self._inv_column_paths[self._COL_FOCUS] = folder
        self._inv_column_paths[self._COL_SIGNALS] = folder
        self._inv_column_paths[self._COL_LABELS] = folder
        self._inv_column_paths[self._COL_EDITED] = os.path.dirname(folder) if folder else ""

        excel = self.excel_path.text().strip()
        if excel:
            self._inv_column_paths[self._COL_EXCEL] = excel

        settings_path = os.path.join(out_dir, "AudioEditor_PerFile_PreAnalysis.json")
        self._inv_column_paths[self._COL_PREANALYSIS] = settings_path

    def _on_inv_header_clicked(self, logical_index: int):
        if logical_index == self._COL_FILE:
            return
        label = self._COL_PATH_LABELS.get(logical_index, f"Column {logical_index}")
        current = self._inv_column_paths.get(logical_index, "")
        is_file = logical_index in (self._COL_EXCEL, self._COL_PREANALYSIS)

        if is_file:
            filt = "Excel files (*.xlsx *.xls)" if logical_index == self._COL_EXCEL else "JSON files (*.json)"
            path, _ = QFileDialog.getOpenFileName(self, label, current or "", filt)
        else:
            path = QFileDialog.getExistingDirectory(self, label, current or "")

        if path:
            self._inv_column_paths[logical_index] = path
            if logical_index == self._COL_EXCEL:
                self.excel_path.setText(path)
            self._scan_folder()

    def _sync_default_analysis_paths(self, folder: str, overwrite=False):
        if not folder or not os.path.isdir(folder):
            return
        settings_path, report_path = self._default_analysis_paths(folder)
        if not os.path.isfile(settings_path):
            alt = os.path.join(folder, "AudioEditor_PerFile_PreAnalysis.json")
            if os.path.isfile(alt):
                settings_path = alt
        if not os.path.isfile(report_path):
            alt = os.path.join(folder, "Audio_PreAnalysis_Report.json")
            if os.path.isfile(alt):
                report_path = alt
        if overwrite or not self.per_file_settings_path.text().strip():
            self.per_file_settings_path.setText(settings_path)
        if overwrite or not self.analysis_report_path.text().strip():
            self.analysis_report_path.setText(report_path)

    def _default_analysis_paths(self, folder: str):
        data_dir = os.path.join(folder, "data")
        out_dir = data_dir if os.path.isdir(data_dir) else folder
        settings_path = os.path.join(out_dir, "AudioEditor_PerFile_PreAnalysis.json")
        report_path = os.path.join(out_dir, "Audio_PreAnalysis_Report.json")
        return settings_path, report_path

    def _analysis_target_files(self) -> list[str]:
        folder = self.input_folder.text().strip()
        all_files = _list_audio_files_in_folder(folder)
        if not self.specify_files_cb.isChecked():
            return all_files
        selected = self.selected_files_combo.selected_files()
        return [f for f in selected if f in all_files]

    def _current_stable_tolerance(self) -> float:
        extractor_panel = self._find_sibling_panel("extractor_panel")
        if extractor_panel is not None:
            try:
                values = extractor_panel.get_values()
                return float(values.get("STABLE_RHYTHM_TOLERANCE", 0.25))
            except Exception:
                pass
            try:
                return float(extractor_panel.stable_tolerance.text())
            except Exception:
                pass
        return 0.25

    def _preview_outside_onset_matches(
        self,
        source_path: str,
        *,
        source_sheet: str | int,
        source_filename_col: str,
        source_onset_col: str,
    ) -> dict:
        folder = self.input_folder.text().strip()
        available_audio_filenames = _list_audio_files_in_folder(folder)
        if not available_audio_filenames:
            raise ValueError("Select an input folder with audio files before importing outside onsets.")

        preview = _eio.scan_matching_onset_rows(
            source_path,
            available_audio_filenames,
            source_filename_col,
            source_onset_col,
            source_sheet,
        )

        target_excel_path = self.excel_path.text().strip()
        existing_names: set[str] = set()
        existing_onsets: dict[str, list[float]] = {}
        if target_excel_path and os.path.isfile(target_excel_path):
            try:
                existing_names = {
                    str(name).strip().lower()
                    for name in _eio.get_filenames(target_excel_path, "File Name", "File Summaries")
                }
            except Exception:
                existing_names = set()
            try:
                existing_onsets = {
                    str(name).strip().lower(): onset_times
                    for name, onset_times in _eio.load_all_onsets(
                        target_excel_path,
                        "File Name",
                        "Exact Onset Times Used (s)",
                        "File Summaries",
                    ).items()
                }
            except Exception:
                existing_onsets = {}

        enriched_matches = []
        for match in preview["matches"]:
            lookup_name = match["audio_filename"].strip().lower()
            current_onsets = existing_onsets.get(lookup_name, [])
            enriched = dict(match)
            enriched["target_row_exists"] = lookup_name in existing_names
            enriched["current_onset_count"] = len(current_onsets)
            enriched_matches.append(enriched)

        preview["matches"] = enriched_matches
        preview["target_excel_path"] = target_excel_path
        preview["available_audio_filenames"] = available_audio_filenames
        return preview

    def _import_outside_onsets(
        self,
        source_path: str,
        *,
        source_sheet: str | int,
        source_filename_col: str,
        source_onset_col: str,
        overwrite_existing: bool,
    ) -> dict:
        target_excel_path = self.excel_path.text().strip()
        if not target_excel_path:
            raise ValueError("Choose the target Excel workbook before importing outside onsets.")
        if target_excel_path.lower().endswith(".xls") and not target_excel_path.lower().endswith(".xlsx"):
            raise ValueError("Please choose an .xlsx workbook for outside-onset import.")

        folder = self.input_folder.text().strip()
        available_audio_filenames = _list_audio_files_in_folder(folder)
        if not available_audio_filenames:
            raise ValueError("No audio files were found in the selected input folder.")

        return _eio.import_matching_onsets_to_workbook(
            source_path,
            target_excel_path,
            available_audio_filenames,
            source_filename_col=source_filename_col,
            source_onset_col=source_onset_col,
            source_sheet=source_sheet,
            overwrite_existing=overwrite_existing,
            stable_tolerance=self._current_stable_tolerance(),
        )

    def _load_outside_onsets(self):
        if not _HAS_EXCEL_IO:
            QMessageBox.warning(
                self,
                "Spreadsheet Support Unavailable",
                "The shared Excel/CSV onset import helpers could not be loaded.",
            )
            return

        folder = self.input_folder.text().strip()
        if not folder or not os.path.isdir(folder):
            QMessageBox.information(
                self,
                "Choose Input Folder",
                "Select the current audio input folder before importing outside onsets.",
            )
            return

        target_excel_path = self.excel_path.text().strip()
        if not target_excel_path:
            suggested = os.path.join(folder, "Cross_Species_Rhythm_Data.xlsx")
            chosen_path, _ = QFileDialog.getSaveFileName(
                self,
                "Choose Target Workbook",
                suggested,
                "Excel Files (*.xlsx)",
            )
            if not chosen_path:
                return
            self.excel_path.setText(chosen_path)
            target_excel_path = chosen_path

        if target_excel_path.lower().endswith(".xls") and not target_excel_path.lower().endswith(".xlsx"):
            QMessageBox.warning(
                self,
                "Unsupported Workbook Format",
                "Outside-onset import writes .xlsx workbooks. Please choose an .xlsx target file.",
            )
            return

        source_path, _ = QFileDialog.getOpenFileName(
            self,
            "Choose Outside Onset Spreadsheet",
            folder,
            "Spreadsheet Files (*.csv *.xlsx *.xls)",
        )
        if not source_path:
            return

        try:
            col_dlg = _ExcelColumnDialog(source_path, parent=self)
        except Exception as exc:
            QMessageBox.critical(
                self,
                "Could Not Read Spreadsheet",
                f"Failed to inspect the selected spreadsheet:\n{exc}",
            )
            return

        if col_dlg.exec() != int(QDialog.DialogCode.Accepted):
            return

        try:
            preview = self._preview_outside_onset_matches(
                source_path,
                source_sheet=col_dlg.sheet_name(),
                source_filename_col=col_dlg.filename_column(),
                source_onset_col=col_dlg.onset_column(),
            )
        except Exception as exc:
            QMessageBox.critical(
                self,
                "Outside Onset Preview Failed",
                str(exc),
            )
            return

        if not preview["matches"]:
            QMessageBox.information(
                self,
                "No Matching Files",
                "No outside-onset rows matched audio files in the current input folder.",
            )
            return

        preview_dlg = _OutsideOnsetImportDialog(
            source_path=source_path,
            target_excel_path=target_excel_path,
            matches=preview["matches"],
            unmatched_source_names=preview["unmatched_source_names"],
            duplicate_targets=preview["duplicate_targets"],
            parent=self,
        )
        if preview_dlg.exec() != int(QDialog.DialogCode.Accepted):
            return

        try:
            result = self._import_outside_onsets(
                source_path,
                source_sheet=col_dlg.sheet_name(),
                source_filename_col=col_dlg.filename_column(),
                source_onset_col=col_dlg.onset_column(),
                overwrite_existing=preview_dlg.overwrite_existing(),
            )
        except Exception as exc:
            QMessageBox.critical(
                self,
                "Outside Onset Import Failed",
                str(exc),
            )
            return

        self._scan_folder()
        details = [
            f"Imported onset data for {result['imported']} file(s).",
        ]
        if result["overwritten_files"]:
            details.append(f"Overwrote existing onset rows for {len(result['overwritten_files'])} file(s).")
        if result["skipped_existing"]:
            details.append(f"Skipped {len(result['skipped_existing'])} file(s) that already had onset values.")
        if result["unmatched_source_names"]:
            details.append(f"{len(result['unmatched_source_names'])} source row(s) did not match the current folder.")
        QMessageBox.information(self, "Outside Onsets Imported", "\n".join(details))

    def _edit_all_hints(self):
        dlg = AnalysisHintsDialog(
            hints=self._global_analysis_hints,
            title="Pre-Analysis Hints (All Files)",
            parent=self,
        )
        if dlg.exec() == int(QDialog.DialogCode.Accepted):
            self._global_analysis_hints = dlg.values()
            self._update_hints_summary()

    def _edit_one_file_hints(self):
        files = self._analysis_target_files()
        if not files:
            folder = self.input_folder.text().strip()
            files = _list_audio_files_in_folder(folder)
        if not files:
            QMessageBox.information(
                self,
                "No Audio Files",
                "Select an input folder with audio files before editing per-file hints.",
            )
            return

        current = None
        if self._analysis_table.currentRow() >= 0:
            item = self._analysis_table.item(self._analysis_table.currentRow(), 0)
            if item:
                current = item.text()
        if current not in files:
            current = files[0]

        chosen, ok = QInputDialog.getItem(
            self,
            "Choose File",
            "Select the file to attach hints to:",
            files,
            files.index(current) if current in files else 0,
            False,
        )
        if not ok or not chosen:
            return

        base = _prep_sanitize_hints(self._global_analysis_hints)
        if chosen in self._file_analysis_hints:
            base.update(_prep_sanitize_hints(self._file_analysis_hints[chosen]))
        dlg = AnalysisHintsDialog(
            hints=base,
            title=f"Pre-Analysis Hints ({chosen})",
            parent=self,
        )
        if dlg.exec() == int(QDialog.DialogCode.Accepted):
            self._file_analysis_hints[chosen] = dlg.values()
            self._update_hints_summary()

    def _update_hints_summary(self):
        n_files = len(self._file_analysis_hints)
        global_bits = []
        if self._global_analysis_hints.get("target_character", "auto") != "auto":
            global_bits.append(self._global_analysis_hints["target_character"])
        if self._global_analysis_hints.get("recording_quality", "auto") != "auto":
            global_bits.append(self._global_analysis_hints["recording_quality"])
        if self._global_analysis_hints.get("speech_like", False):
            global_bits.append("speech-like")
        if self._global_analysis_hints.get("expected_bpm") is not None:
            global_bits.append(f"{self._global_analysis_hints['expected_bpm']:.0f} BPM")

        global_text = ", ".join(global_bits) if global_bits else "auto"
        self._hints_summary.setText(
            f"All-files hints: {global_text}  |  Per-file overrides: {n_files}")

    def _import_pre_analysis(self):
        start_dir = self.input_folder.text().strip() or ""
        path, _ = QFileDialog.getOpenFileName(
            self,
            "Import Pre-Analysis JSON",
            start_dir,
            "JSON files (*.json);;All files (*)",
        )
        if not path:
            return

        try:
            with open(path, encoding="utf-8") as f:
                data = json.load(f)
        except Exception as exc:
            QMessageBox.critical(
                self,
                "Import Failed",
                f"Could not read the JSON file:\n{exc}",
            )
            return

        if not isinstance(data, dict) or not data:
            QMessageBox.warning(
                self,
                "Invalid Format",
                "The JSON file does not appear to contain per-file pre-analysis data.",
            )
            return

        if "files" in data and isinstance(data["files"], dict):
            per_file_payload = data["files"]
        else:
            per_file_payload = data

        sample = next(iter(per_file_payload.values()), None)
        if not isinstance(sample, dict) or not any(
            key in sample for key in ("settings", "analysis", "onset_recommendations")
        ):
            QMessageBox.warning(
                self,
                "Invalid Format",
                "The JSON file does not appear to contain per-file pre-analysis data.\n\n"
                "Expected a JSON object mapping filenames to analysis objects with "
                "'settings', 'analysis', or 'onset_recommendations' keys.",
            )
            return

        self.per_file_settings_path.setText(path)

        if "files" in data and "representative_file" in data:
            self.analysis_report_path.setText(path)

        rows = []
        for fname, entry in per_file_payload.items():
            if not isinstance(entry, dict):
                continue
            analysis = entry.get("analysis", {})
            settings = entry.get("settings", {})
            onset = entry.get("onset_recommendations", {})
            reasoning = entry.get("reasoning", {})
            onset_reasoning = entry.get("onset_reasoning", {})
            profile_source = entry.get("profile_source", "imported")
            hints = entry.get("hints", {})

            h = float(analysis.get("harmonic_ratio", 0.5) or 0.5)
            char = _prep_infer_signal_character(h)

            rows.append({
                "filename": fname,
                "profile_source": profile_source,
                "analysis": analysis,
                "settings": settings,
                "reasoning": reasoning,
                "onset_settings": onset,
                "onset_reasoning": onset_reasoning,
                "character": char,
                "notes": ["imported"],
                "hints": hints,
            })

        self._last_analysis_rows = rows
        self._analysis_table.setRowCount(len(rows))
        for r, row in enumerate(rows):
            analysis = row.get("analysis", {})
            settings = row.get("settings", {})
            onset = row.get("onset_settings", {})

            snr = _prep_float_or_none(analysis.get("snr_db"))
            noise = _prep_float_or_none(analysis.get("noise_floor_db"))
            hp = int(settings.get("MUTER_HIGHPASS_HZ", 0) or 0)
            lp = int(settings.get("MUTER_LOWPASS_HZ", 0) or 0)

            cells = [
                str(row.get("filename", "")),
                str(row.get("profile_source", "imported")),
                "" if snr is None else f"{snr:.1f}",
                "" if noise is None else f"{noise:.1f}",
                str(row.get("character", "mixed")),
                str(onset.get("ONSET_METHOD", "")),
                str(onset.get("MIN_INTER_ONSET_MS", "")),
                f"{hp} / {lp}",
                ", ".join(row.get("notes", [])),
            ]
            for c, text in enumerate(cells):
                self._analysis_table.setItem(r, c, QTableWidgetItem(text))

        if rows:
            snrs = [float(r["analysis"].get("snr_db", 20.0) or 20.0) for r in rows]
            median_snr = float(np.median(snrs))
            representative = min(
                rows,
                key=lambda rr: abs(float(rr["analysis"].get("snr_db", 20.0) or 20.0) - median_snr),
            )
            self._apply_recommendations_to_panels(
                representative.get("settings", {}),
                representative.get("onset_settings", {}),
            )

        self._analysis_status.setText(
            f"Imported pre-analysis for {len(rows)} file(s) from JSON.")

    def _import_analysis_report(self):
        start_dir = self.input_folder.text().strip() or ""
        path, _ = QFileDialog.getOpenFileName(
            self,
            "Import Analysis Report JSON",
            start_dir,
            "JSON files (*.json);;All files (*)",
        )
        if not path:
            return

        try:
            with open(path, encoding="utf-8") as f:
                data = json.load(f)
        except Exception as exc:
            QMessageBox.critical(
                self, "Import Failed",
                f"Could not read the JSON file:\n{exc}",
            )
            return

        if not isinstance(data, dict) or "files" not in data or not isinstance(data["files"], dict):
            QMessageBox.warning(
                self, "Invalid Format",
                "The JSON file does not appear to be a valid analysis report "
                "(should have a 'files' key).",
            )
            return

        self.analysis_report_path.setText(path)
        self.per_file_settings_path.setText(path)

        per_file_payload = data["files"]
        rows = []
        for fname, entry in per_file_payload.items():
            if not isinstance(entry, dict):
                continue
            analysis = entry.get("analysis", {})
            settings = entry.get("settings", {})
            onset = entry.get("onset_recommendations", {})
            reasoning = entry.get("reasoning", {})
            onset_reasoning = entry.get("onset_reasoning", {})
            profile_source = entry.get("profile_source", "imported")
            hints = entry.get("hints", {})

            h = float(analysis.get("harmonic_ratio", 0.5) or 0.5)
            char = _prep_infer_signal_character(h)

            rows.append({
                "filename": fname,
                "profile_source": profile_source,
                "analysis": analysis,
                "settings": settings,
                "reasoning": reasoning,
                "onset_settings": onset,
                "onset_reasoning": onset_reasoning,
                "character": char,
                "notes": ["imported"],
                "hints": hints,
            })

        self._last_analysis_rows = rows
        self._analysis_table.setRowCount(len(rows))
        for r, row in enumerate(rows):
            analysis = row.get("analysis", {})
            settings = row.get("settings", {})
            onset = row.get("onset_settings", {})

            snr = _prep_float_or_none(analysis.get("snr_db"))
            noise = _prep_float_or_none(analysis.get("noise_floor_db"))
            hp = int(settings.get("MUTER_HIGHPASS_HZ", 0) or 0)
            lp = int(settings.get("MUTER_LOWPASS_HZ", 0) or 0)

            cells = [
                str(row.get("filename", "")),
                str(row.get("profile_source", "imported")),
                "" if snr is None else f"{snr:.1f}",
                "" if noise is None else f"{noise:.1f}",
                str(row.get("character", "mixed")),
                str(onset.get("ONSET_METHOD", "")),
                str(onset.get("MIN_INTER_ONSET_MS", "")),
                f"{hp} / {lp}",
                ", ".join(row.get("notes", [])),
            ]
            for c, text in enumerate(cells):
                self._analysis_table.setItem(r, c, QTableWidgetItem(text))

        if rows:
            snrs = [float(r["analysis"].get("snr_db", 20.0) or 20.0) for r in rows]
            median_snr = float(np.median(snrs))
            representative = min(
                rows,
                key=lambda rr: abs(float(rr["analysis"].get("snr_db", 20.0) or 20.0) - median_snr),
            )
            self._apply_recommendations_to_panels(
                representative.get("settings", {}),
                representative.get("onset_settings", {}),
            )

        self._analysis_status.setText(
            f"Imported analysis report for {len(rows)} file(s) from JSON.")

    def _apply_recommendations_to_panels(self, audio_settings: dict,
                                         onset_settings: dict):
        main = self.window()
        if self._apply_muter_cb.isChecked() and hasattr(main, "muter_panel"):
            cur = main.muter_panel.get_values()
            cur.update(audio_settings)
            cur["MUTER_PRESET"] = None
            main.muter_panel.set_values(cur)

        if self._apply_extractor_cb.isChecked() and hasattr(main, "extractor_panel"):
            cur = main.extractor_panel.get_values()
            cur.update(onset_settings)
            cur["ACTIVE_PRESET"] = None
            main.extractor_panel.set_values(cur)

    def _run_pre_analysis(self):
        folder = self.input_folder.text().strip()
        if not os.path.isdir(folder):
            QMessageBox.warning(self, "Folder Not Found",
                                "Choose a valid audio input folder first.")
            return

        files = self._analysis_target_files()
        if not files:
            QMessageBox.information(
                self,
                "No Files Selected",
                "No audio files are selected for pre-analysis.",
            )
            return

        self._sync_default_analysis_paths(folder, overwrite=False)

        try:
            from analysis.audio_recommendations import (
                analyze_audio,
                harmonic_percussive_ratio as _aa_hp_ratio,
                spectral_profile as _aa_spectral_profile,
            )
        except Exception:
            project_root = os.path.dirname(
                os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
            )
            if project_root not in sys.path:
                sys.path.insert(0, project_root)
            try:
                from analysis.audio_recommendations import (
                    analyze_audio,
                    harmonic_percussive_ratio as _aa_hp_ratio,
                    spectral_profile as _aa_spectral_profile,
                )
            except Exception as exc:
                QMessageBox.critical(
                    self,
                    "Audio Analyzer Unavailable",
                    f"Failed to import analysis/audio_recommendations.py:\n{exc}",
                )
                return

        self._pre_analyze_btn.setEnabled(False)
        total = len(files)
        self._analysis_progress.setMaximum(total)
        self._analysis_progress.setValue(0)
        self._analysis_progress.setFormat("%v / %m files")
        self._analysis_progress.show()
        self._analysis_status.setText(f"Analyzing {total} file(s)...")
        QApplication.setOverrideCursor(Qt.CursorShape.WaitCursor)

        rows = []
        per_file_payload = {}
        failures = []

        try:
            for i, fname in enumerate(files, start=1):
                self._analysis_status.setText(f"Analyzing {i}/{total}: {fname}")
                self._analysis_progress.setValue(i)
                QApplication.processEvents()

                file_path = os.path.join(folder, fname)
                hints = _prep_sanitize_hints(self._global_analysis_hints)
                if fname in self._file_analysis_hints:
                    hints.update(_prep_sanitize_hints(self._file_analysis_hints[fname]))
                hints = _prep_sanitize_hints(hints)

                try:
                    profile = None
                    profile_source = "none"
                    if self._use_signal_hints_cb.isChecked() and hints.get("use_signal_hints", True):
                        profile, profile_source = _prep_load_signal_profile_hint(
                            folder, fname, _aa_spectral_profile, _aa_hp_ratio)
                    profile = _prep_apply_hints_to_profile(profile, hints)

                    result = analyze_audio(file_path, signal_profile=profile)
                    analysis = dict(result.get("analysis", {}))
                    settings = dict(result.get("settings", {}))
                    reasoning = dict(result.get("reasoning", {}))

                    notes = _prep_apply_hint_overrides_to_muter(settings, analysis, hints)
                    onset_settings, onset_reasoning = _prep_recommend_onset_settings(
                        analysis, settings, hints)

                    char = "mixed"
                    if profile and isinstance(profile.get("summary"), dict):
                        char = str(profile["summary"].get("signal_character", "mixed"))
                    else:
                        h = float(analysis.get("harmonic_ratio", 0.5) or 0.5)
                        char = _prep_infer_signal_character(h)

                    row = {
                        "filename": fname,
                        "profile_source": profile_source,
                        "analysis": analysis,
                        "settings": settings,
                        "reasoning": reasoning,
                        "onset_settings": onset_settings,
                        "onset_reasoning": onset_reasoning,
                        "character": char,
                        "notes": notes,
                        "hints": hints,
                    }
                    rows.append(row)

                    per_file_payload[fname] = {
                        "settings": settings,
                        "analysis": analysis,
                        "reasoning": reasoning,
                        "onset_recommendations": onset_settings,
                        "onset_reasoning": onset_reasoning,
                        "profile_source": profile_source,
                        "hints": hints,
                    }
                except Exception as exc:
                    failures.append(f"{fname}: {exc}")

            self._last_analysis_rows = rows
            self._analysis_table.setRowCount(len(rows))
            for r, row in enumerate(rows):
                analysis = row.get("analysis", {})
                settings = row.get("settings", {})
                onset = row.get("onset_settings", {})

                snr = _prep_float_or_none(analysis.get("snr_db"))
                noise = _prep_float_or_none(analysis.get("noise_floor_db"))
                hp = int(settings.get("MUTER_HIGHPASS_HZ", 0) or 0)
                lp = int(settings.get("MUTER_LOWPASS_HZ", 0) or 0)

                cells = [
                    str(row.get("filename", "")),
                    str(row.get("profile_source", "none")),
                    "" if snr is None else f"{snr:.1f}",
                    "" if noise is None else f"{noise:.1f}",
                    str(row.get("character", "mixed")),
                    str(onset.get("ONSET_METHOD", "")),
                    str(onset.get("MIN_INTER_ONSET_MS", "")),
                    f"{hp} / {lp}",
                    ", ".join(row.get("notes", [])),
                ]
                for c, text in enumerate(cells):
                    self._analysis_table.setItem(r, c, QTableWidgetItem(text))

            if not rows:
                self._analysis_status.setText("Pre-analysis failed for all selected files.")
                if failures:
                    QMessageBox.warning(
                        self,
                        "Pre-Analysis Failed",
                        "No files were successfully analyzed.\n\n"
                        + "\n".join(failures[:8]),
                    )
                return

            snrs = [float(r["analysis"].get("snr_db", 20.0) or 20.0) for r in rows]
            median_snr = float(np.median(snrs))
            representative = min(
                rows,
                key=lambda rr: abs(float(rr["analysis"].get("snr_db", 20.0) or 20.0) - median_snr),
            )

            settings_path = self.per_file_settings_path.text().strip()
            report_path = self.analysis_report_path.text().strip()
            default_settings, default_report = self._default_analysis_paths(folder)
            if not settings_path:
                settings_path = default_settings
            if not report_path:
                report_path = default_report

            settings_dir = os.path.dirname(settings_path)
            report_dir = os.path.dirname(report_path)
            if settings_dir:
                os.makedirs(settings_dir, exist_ok=True)
            if report_dir:
                os.makedirs(report_dir, exist_ok=True)

            with open(settings_path, "w", encoding="utf-8") as f:
                per_file_on_disk = dict(per_file_payload)
                per_file_on_disk["__pre_analysis_original__"] = deepcopy(
                    per_file_payload)
                json.dump(per_file_on_disk, f, indent=2)

            report_obj = {
                "created_at": datetime.datetime.now().isoformat(timespec="seconds"),
                "input_folder": folder,
                "n_files_requested": len(files),
                "n_files_analyzed": len(rows),
                "failures": failures,
                "representative_file": representative.get("filename"),
                "representative_audio_settings": representative.get("settings", {}),
                "representative_onset_settings": representative.get("onset_settings", {}),
                "files": per_file_payload,
            }
            with open(report_path, "w", encoding="utf-8") as f:
                json.dump(report_obj, f, indent=2)

            self.per_file_settings_path.setText(settings_path)
            self.analysis_report_path.setText(report_path)

            self._apply_recommendations_to_panels(
                representative.get("settings", {}),
                representative.get("onset_settings", {}),
            )

            fail_txt = f"  ({len(failures)} failed)" if failures else ""
            self._analysis_status.setText(
                f"Pre-analysis complete: {len(rows)} file(s) analyzed{fail_txt}.")
        finally:
            QApplication.restoreOverrideCursor()
            self._pre_analyze_btn.setEnabled(True)
            self._analysis_progress.setValue(self._analysis_progress.maximum())

    def _scan_folder(self):
        folder = self.input_folder.text().strip()
        if not os.path.isdir(folder):
            self._scan_status.setText("⚠ Folder not found.")
            self._inv_table.setRowCount(0)
            if hasattr(self, "_metadata_section"):
                self._metadata_section.set_filenames([])
            return

        audio_files = _list_audio_files_in_folder(folder)
        if not audio_files:
            self._scan_status.setText("No audio files found.")
            self._inv_table.setRowCount(0)
            if hasattr(self, "_metadata_section"):
                self._metadata_section.set_filenames([])
            return

        try:
            all_entries = set(os.listdir(folder))
        except OSError:
            all_entries = set()

        dir_entries = {entry for entry in all_entries
                       if os.path.isdir(os.path.join(folder, entry))}

        excel_files_in_sheet = set()
        excel_path = self.excel_path.text().strip()
        if excel_path and os.path.isfile(excel_path):
            try:
                import openpyxl
                wb = openpyxl.load_workbook(excel_path, read_only=True,
                                            data_only=True)
                if "File Summaries" in wb.sheetnames:
                    ws = wb["File Summaries"]
                    for row in ws.iter_rows(min_row=2, max_col=1,
                                            values_only=True):
                        if row[0]:
                            excel_files_in_sheet.add(str(row[0]))
                wb.close()
            except Exception:
                pass

        preanalysis_files = set()
        for candidate in [
            os.path.join(folder, "data", "AudioEditor_PerFile_PreAnalysis.json"),
            os.path.join(folder, "AudioEditor_PerFile_PreAnalysis.json"),
        ]:
            if os.path.isfile(candidate):
                try:
                    with open(candidate, encoding="utf-8") as f:
                        pa_data = json.load(f)
                    if isinstance(pa_data, dict):
                        pa_entries = pa_data.get("files", pa_data)
                        if isinstance(pa_entries, dict):
                            preanalysis_files.update(pa_entries.keys())
                except Exception:
                    pass
                break
        user_pa_path = self.per_file_settings_path.text().strip()
        if user_pa_path and os.path.isfile(user_pa_path):
            try:
                with open(user_pa_path, encoding="utf-8") as f:
                    pa_data = json.load(f)
                if isinstance(pa_data, dict):
                    pa_entries = pa_data.get("files", pa_data)
                    if isinstance(pa_entries, dict):
                        preanalysis_files.update(pa_entries.keys())
            except Exception:
                pass

        self._inv_table.setRowCount(len(audio_files))

        for r, fname in enumerate(audio_files):
            stem = os.path.splitext(fname)[0]
            self._inv_table.setItem(r, 0, QTableWidgetItem(fname))

            layers_dir = f"{stem}_OnsetLayers"
            layer_count = 0
            layers_path = ""
            if layers_dir in dir_entries:
                layers_path = os.path.join(folder, layers_dir)
            elif os.path.isdir(os.path.join(folder, "data", layers_dir)):
                layers_path = os.path.join(folder, "data", layers_dir)
            else:
                custom = self._inv_column_paths.get(self._COL_LAYERS, "")
                if custom and custom != folder:
                    cand = os.path.join(custom, layers_dir)
                    if os.path.isdir(cand):
                        layers_path = cand
            if layers_path:
                try:
                    layer_count = len([f for f in os.listdir(layers_path)
                                       if f.endswith(".json")])
                except OSError:
                    pass
            self._set_status_cell(r, 1, layer_count > 0,
                                  f"{layer_count}" if layer_count else "")

            fr_file = f"{stem}_focus_regions.json"
            has_fr = fr_file in all_entries
            fr_path = os.path.join(folder, fr_file) if has_fr else ""
            if not has_fr:
                data_fr = os.path.join(folder, "data", fr_file)
                if os.path.isfile(data_fr):
                    has_fr = True
                    fr_path = data_fr
            fr_detail = ""
            if has_fr:
                try:
                    with open(fr_path, encoding="utf-8") as f:
                        fr_data = json.load(f)
                    if isinstance(fr_data, dict):
                        regions = fr_data.get(fname, [])
                        n_pos = sum(1 for rr in regions
                                    if rr.get("polarity") != "negative")
                        n_neg = sum(1 for rr in regions
                                    if rr.get("polarity") == "negative")
                        parts = []
                        if n_pos:
                            parts.append(f"{n_pos}P")
                        if n_neg:
                            parts.append(f"{n_neg}N")
                        fr_detail = " ".join(parts)
                except Exception:
                    pass
            self._set_status_cell(r, 2, has_fr, fr_detail)

            ss_dir = f"{stem}_SelectedSignals"
            has_ss = ss_dir in dir_entries
            ss_path = os.path.join(folder, ss_dir) if has_ss else ""
            if not has_ss:
                data_ss = os.path.join(folder, "data", ss_dir)
                if os.path.isdir(data_ss):
                    has_ss = True
                    ss_path = data_ss
            if not has_ss:
                custom_ss_dir = self._inv_column_paths.get(self._COL_SIGNALS, "")
                if custom_ss_dir and custom_ss_dir != folder:
                    cand = os.path.join(custom_ss_dir, ss_dir)
                    if os.path.isdir(cand):
                        has_ss = True
                        ss_path = cand
            ss_detail = ""
            if has_ss:
                n_pos = n_neg = 0
                pos_dir = os.path.join(ss_path, "signalPositive")
                neg_dir = os.path.join(ss_path, "signalNegative")
                try:
                    if os.path.isdir(pos_dir):
                        n_pos = len([f for f in os.listdir(pos_dir)
                                     if os.path.splitext(f)[1].lower()
                                     in AUDIO_EXTENSIONS])
                    if os.path.isdir(neg_dir):
                        n_neg = len([f for f in os.listdir(neg_dir)
                                     if os.path.splitext(f)[1].lower()
                                     in AUDIO_EXTENSIONS])
                except OSError:
                    pass
                parts = []
                if n_pos:
                    parts.append(f"{n_pos}P")
                if n_neg:
                    parts.append(f"{n_neg}N")
                ss_detail = " ".join(parts)
            self._set_status_cell(r, 3, has_ss, ss_detail)

            suffixes = ["_muted_clean", "_demucs_stems",
                        "_rejected_noise", "_hpss_percussive",
                        "_hpss_harmonic"]
            parent = os.path.dirname(folder)
            folder_base = os.path.basename(folder.rstrip("/\\"))
            found_outputs = []
            try:
                parent_entries = set(os.listdir(parent)) if parent else set()
            except OSError:
                parent_entries = set()
            for suffix in suffixes:
                out_dir = f"{folder_base}{suffix}"
                if out_dir in parent_entries:
                    out_path = os.path.join(parent, out_dir, fname)
                    if os.path.isfile(out_path):
                        found_outputs.append(suffix.lstrip("_"))
            has_edited = bool(found_outputs)
            edit_detail = ", ".join(found_outputs) if found_outputs else ""
            self._set_status_cell(r, 4, has_edited, edit_detail)

            in_excel = fname in excel_files_in_sheet or stem in excel_files_in_sheet
            self._set_status_cell(r, 5, in_excel)

            label_file = f"{stem}_labels.txt"
            has_labels = label_file in all_entries
            if not has_labels:
                if os.path.isfile(os.path.join(folder, "data", label_file)):
                    has_labels = True
            self._set_status_cell(r, 6, has_labels)

            has_pa = fname in preanalysis_files
            self._set_status_cell(r, 7, has_pa)

        self._scan_status.setText(
            f"Found {len(audio_files)} audio file(s).")

        if hasattr(self, "_metadata_section"):
            self._metadata_section.set_filenames(audio_files)

    def _set_status_cell(self, row, col, present, detail=""):
        if present:
            text = f"✓ {detail}".strip() if detail else "✓"
            item = QTableWidgetItem(text)
            item.setForeground(QColor(_ACCENT))
        else:
            item = QTableWidgetItem("✗")
            item.setForeground(QColor(_TEXT_DIM))
        item.setTextAlignment(
            Qt.AlignmentFlag.AlignCenter | Qt.AlignmentFlag.AlignVCenter)
        self._inv_table.setItem(row, col, item)

    def _show_help_dialog(self, title: str, html: str):
        dlg = QDialog(self)
        dlg.setWindowTitle(title)
        dlg.setMinimumSize(760, 620)
        dlg.setStyleSheet(
            f"QDialog {{ background-color: {_BG_MID}; color: {_TEXT}; }}"
        )

        v = QVBoxLayout(dlg)
        v.setContentsMargins(12, 12, 12, 12)
        v.setSpacing(10)

        browser = QTextBrowser()
        browser.setOpenExternalLinks(False)
        browser.setStyleSheet(
            f"QTextBrowser {{ background-color: {_BG_WIDGET}; color: {_TEXT}; "
            f"border: 1px solid {_BORDER}; border-radius: 4px; "
            f"padding: 10px; font-size: 12px; }}"
        )
        styled = (
            "<html><head><style>"
            f"body {{ color: {_TEXT}; font-family: -apple-system, Segoe UI, sans-serif; "
            "font-size: 12px; line-height: 1.45; }}"
            f"h2 {{ color: {_ACCENT}; margin-top: 2px; margin-bottom: 6px; "
            "font-size: 15px; }}"
            f"h3 {{ color: {_ACCENT}; margin-top: 14px; margin-bottom: 4px; "
            "font-size: 13px; }}"
            f"code {{ background-color: {_BG_MID}; color: {_TEXT}; "
            "padding: 1px 4px; border-radius: 3px; font-size: 11px; }}"
            "ul { margin-top: 2px; margin-bottom: 6px; }"
            "li { margin-bottom: 3px; }"
            f"b {{ color: {_TEXT}; }}"
            f".muted {{ color: {_TEXT_DIM}; }}"
            "</style></head><body>"
            + html
            + "</body></html>"
        )
        browser.setHtml(styled)
        v.addWidget(browser, stretch=1)

        btn_row = QHBoxLayout()
        btn_row.addStretch(1)
        close_btn = QPushButton("  Close  ")
        close_btn.setStyleSheet(
            f"QPushButton {{ background-color: {_ACCENT_DIM}; color: white; "
            f"border-radius: 5px; padding: 5px 14px; font-size: 12px; "
            f"border: none; }} "
            f"QPushButton:hover {{ background-color: {_ACCENT}; }}"
        )
        close_btn.clicked.connect(dlg.accept)
        btn_row.addWidget(close_btn)
        v.addLayout(btn_row)

        dlg.exec()

    def _show_file_inventory_help(self):
        html = """
<h2>File Inventory — How this feature works</h2>

<h3>What it does</h3>
<p>The <b>Scan Folder</b> button looks inside your chosen input folder
(plus its <code>data/</code> subfolder as a fallback) and builds one row
per audio file. Each row shows, at a glance, whether seven kinds of
supporting artifacts already exist for that file.</p>

<h3>How each column is detected</h3>
<p>For an audio file named <code>example.wav</code> (stem =
<code>example</code>), each column is populated by looking for the
following on disk:</p>
<ul>
<li><b>Onset Layers 📁</b> — directory <code>example_OnsetLayers/</code>;
the cell shows the number of <code>.json</code> layer files inside.</li>
<li><b>Focus Regions 📁</b> — file <code>example_focus_regions.json</code>;
the cell shows <code>NP</code> / <code>NN</code> counts
(positive / negative polarity regions).</li>
<li><b>Sel. Signals 📁</b> — directory
<code>example_SelectedSignals/</code> containing
<code>signalPositive/</code> and <code>signalNegative/</code> audio
clips; counts are shown as <code>NP</code> / <code>NN</code>.</li>
<li><b>Edited Output 📁</b> — sibling folders named after the input
folder with the suffixes <code>_muted_clean</code>,
<code>_demucs_stems</code>, <code>_rejected_noise</code>,
<code>_hpss_percussive</code>, <code>_hpss_harmonic</code>;
a ✓ means the file exists in at least one of them.</li>
<li><b>In Excel 📁</b> — opens the workbook at "Output Excel Path",
reads the <code>File Summaries</code> sheet (column A), and marks ✓ if
the filename or its stem appears there.</li>
<li><b>Labels 📁</b> — file <code>example_labels.txt</code>
(Audacity-style label track).</li>
<li><b>Pre-Analysis 📁</b> — JSON file
<code>AudioEditor_PerFile_PreAnalysis.json</code> (in the folder or its
<code>data/</code> subfolder, or at the custom path you supplied);
✓ if this file appears as a key inside it.</li>
</ul>
<p class="muted">Tip: clicking a column header (📁) lets you redirect
the lookup to a different folder or file.</p>

<h3>Why this matters</h3>
<p>The inventory is <i>descriptive</i>: it does not change anything on
disk. It exists for two reasons:</p>
<ul>
<li><b>Situational awareness</b> — you can see which files already have
selected signals, focus regions, edits, onset layers, etc. before you
commit to running any pipeline step.</li>
<li><b>Pre-Analysis inputs</b> — three of the columns (Sel. Signals,
Focus Regions, and any explicit <code>*_signal_profile.json</code>) are
the exact hints that the <b>Pre-Analyze Files</b> button will consume
per file when "Use signal hints" is checked. A ✓ in those columns means
Pre-Analysis will get a more targeted per-file recommendation rather
than a generic one.</li>
</ul>
<p>The remaining columns (Onset Layers, Edited Output, In Excel,
Labels, Pre-Analysis) are informational only — they tell you which
downstream artifacts already exist, but they are <b>not</b> used as
inputs by Pre-Analyze itself.</p>
"""
        self._show_help_dialog("File Inventory — Help", html)

    def _show_pre_analyze_help(self):
        html = """
<h2>Pre-Analyze Audio Files — How this feature works</h2>

<h3>What it does</h3>
<p>For each selected audio file, Pre-Analyze calls
<code>audio_analyzer.analyze_audio()</code> to measure objective
acoustic properties, then derives recommended
<b>Audio Editor</b> (<code>MUTER_*</code>) and <b>Onset Finder</b>
settings from those measurements. Two JSON files are written:</p>
<ul>
<li><code>AudioEditor_PerFile_PreAnalysis.json</code> — the per-file
payload that the Onset Finder reads at run time (via
<code>per_file_settings_path</code>) so each file uses its own tuned
onset method, <code>MIN_INTER_ONSET_MS</code>,
<code>ONSET_DELTA</code>, and high-pass settings.</li>
<li><code>AudioEditor_PreAnalysis_Report.json</code> — a
human-readable report with measurements, reasoning strings, and the
"representative" file chosen for the GUI panels.</li>
</ul>

<h3>What data is actually used — per file</h3>
<p><b>From the audio itself (always):</b></p>
<ul>
<li>Noise floor (dB), SNR, peak level, dynamic range</li>
<li>Spectral profile: centroid, bandwidth, energy below 200 Hz, energy
above 8 kHz</li>
<li>Harmonic-vs-percussive ratio (HPSS decomposition)</li>
<li>Tonal noise frequencies (fixed-pitch interferers like hum / cicadas)</li>
<li>Noise stationarity (how steady the background is)</li>
</ul>

<p><b>From signal hints (only when "Use signal hints" is checked)</b> —
loaded in this priority order, first match wins:</p>
<ol>
<li><code>{stem}_signal_profile.json</code> (explicit profile file) →
source = <code>signal_profile</code></li>
<li><code>{stem}_SelectedSignals/signalPositive</code> +
<code>signalNegative/</code> — up to 12 clips are loaded and summarized
for frequency range, harmonicity, and attack sharpness →
source = <code>selected_signals</code></li>
<li><code>{stem}_focus_regions.json</code> —
<code>f_low</code>, <code>f_high</code>, <code>t_start</code>,
<code>t_end</code>, and <code>polarity</code> of each region →
source = <code>focus_regions</code></li>
</ol>
<p>A profile supplies: <code>freq_range_hz</code>,
<code>harmonicity</code>, <code>attack_sharpness</code>,
<code>signal_character</code> (harmonic / percussive / mixed),
<code>spectral_bandwidth_hz</code>, and a separate negative-region
summary for unwanted sounds.</p>

<p><b>From hints you entered</b> (All-File + One-File Hints buttons;
One-File overrides All-File):</p>
<ul>
<li><code>signal_min_hz</code> / <code>signal_max_hz</code> — clamp the
band around the signal</li>
<li><code>target_character</code> — force harmonic / percussive / mixed</li>
<li><code>speech_like</code> — route to a speech-friendly onset method</li>
<li><code>expected_bpm</code> — directly sets
<code>MIN_INTER_ONSET_MS</code> via
<code>60000 / BPM × 0.45</code></li>
<li><code>target_density</code> — dense / moderate / sparse fallback
for <code>MIN_INTER_ONSET_MS</code></li>
<li><code>prefer_onset_method</code> — override the auto-picked
<code>ONSET_METHOD</code></li>
</ul>

<p class="muted"><b>File Inventory columns that are NOT read by
Pre-Analyze:</b> Onset Layers, Edited Output, In Excel, Labels, and the
Pre-Analysis column itself. Those are shown for awareness only.</p>

<h3>How the recommendations are derived</h3>
<p><b>Audio Editor (<code>MUTER_*</code>):</b></p>
<ul>
<li><code>MUTER_HIGHPASS_HZ</code> — set ~30% below the signal's lowest
frequency; raised if a negative region sits below the signal band.
Falls back to an <code>energy_below_200hz</code> heuristic when no
profile is available.</li>
<li><code>MUTER_LOWPASS_HZ</code> — set ~30% above the signal's highest
frequency; tightened if a negative region sits above the signal band.</li>
<li><code>MUTER_HPSS_ENABLED</code> / <code>_TARGET</code> /
<code>_MARGIN</code> — enabled when <code>|H − P| > 0.2</code> or when
the negative profile has the opposite character from the signal.</li>
<li>Spectral denoise, bandpass boost, notch filters, transient
sharpening — chosen from SNR, stationarity, and tonal-noise results.</li>
</ul>

<p><b>Onset Finder:</b></p>
<ul>
<li><code>ONSET_METHOD</code> — <code>syllable_nuclei</code> if
<code>speech_like</code>; <code>madmom_beats</code> or
<code>superflux</code> if percussive (P ≥ 0.62);
<code>adaptive_hp</code> if harmonic (H ≥ 0.62); otherwise
<code>librosa</code> or <code>superflux</code> based on SNR.</li>
<li><code>MIN_INTER_ONSET_MS</code> — from <code>expected_bpm</code> if
given; else from <code>target_density</code>; else from a heuristic on
H/P ratio with a +6 ms bump when SNR &lt; 12 dB.</li>
<li><code>ONSET_DELTA</code> — sensitivity scaled from SNR:
0.13 (SNR&lt;10), 0.10, 0.08, 0.06 (SNR≥28).</li>
<li><code>APPLY_HIGHPASS_FILTER</code> /
<code>HIGHPASS_CUTOFF_HZ</code> — mirrors the recommended muter
high-pass so the Onset Finder sees pre-filtered energy.</li>
<li><code>PITCH_TRACKER</code> — <code>pyin</code> when the target is
harmonic (and not speech); otherwise <code>none</code>.</li>
</ul>

<p><b>Representative file.</b> After all files are analyzed, the one
with the <i>median SNR</i> is chosen as representative. Its
recommendations are pushed into the Audio Editor and Onset Finder
panels (gated by the two "Apply to…" checkboxes). The per-file JSON
is always saved in full, so the Onset Finder still applies each file's
own tuned settings when it actually runs.</p>

<h3>Why run it</h3>
<p>Recordings in a real corpus vary enormously in SNR, frequency
content, and signal character. Hand-tuning ~30 parameters per file is
impractical, so Pre-Analyze measures each file objectively and picks
starting values that are usually close to optimal — leaving you to
tweak from a reasonable baseline rather than guess.</p>
"""
        self._show_help_dialog("Pre-Analyze Audio Files — Help", html)

    def _find_sibling_panel(self, attr_name: str):
        w = self.parent()
        while w is not None:
            if hasattr(w, attr_name):
                return getattr(w, attr_name)
            w = w.parent()
        return None

    def _open_per_file_settings_dialog(self):
        path = self.per_file_settings_path.text().strip()
        if not path:
            folder = ""
            try:
                folder = self.input_folder.text().strip()
            except Exception:
                pass
            if folder and os.path.isdir(folder):
                settings_path, _ = self._default_analysis_paths(folder)
                path = settings_path
                self.per_file_settings_path.setText(path)

        muter_panel = self._find_sibling_panel("muter_panel")
        extractor_panel = self._find_sibling_panel("extractor_panel")
        manager = self._find_sibling_panel("per_file_mgr")
        if manager is not None and path and manager.path() != path:
            manager.load(path)
        elif manager is not None and not manager.path() and path:
            manager.load(path)

        try:
            dlg = PerFileSettingsDialog(
                json_path=path,
                muter_panel=muter_panel,
                extractor_panel=extractor_panel,
                manager=manager,
                parent=self,
            )
        except Exception as exc:
            QMessageBox.critical(
                self,
                "Could Not Open Per-File Settings",
                f"Failed to open per-file settings dialog:\n{exc}",
            )
            return

        dlg.exec()

    def get_audio_analyzer_values(self):
        return {
            "RUN_AUDIO_ANALYZER": bool(self.per_file_settings_path.text().strip()),
            "per_file_settings_path": self.per_file_settings_path.text().strip(),
            "analysis_report_path": self.analysis_report_path.text().strip(),
        }

    def set_audio_analyzer_values(self, d):
        if not isinstance(d, dict):
            return
        p = d.get("per_file_settings_path")
        if p:
            self.per_file_settings_path.setText(str(p))
        rp = d.get("analysis_report_path")
        if rp:
            self.analysis_report_path.setText(str(rp))

    def get_values(self):
        return {
            "PREP_INPUT_FOLDER": self.input_folder.text(),
            "PREP_SPECIFY_FILES": self.specify_files_cb.isChecked(),
            "PREP_SELECTED_FILES": self.selected_files_combo.selected_files(),
            "PREP_EXCEL_PATH": self.excel_path.text(),
            "PREP_USE_SIGNAL_HINTS": self._use_signal_hints_cb.isChecked(),
            "PREP_APPLY_TO_MUTER": self._apply_muter_cb.isChecked(),
            "PREP_APPLY_TO_EXTRACTOR": self._apply_extractor_cb.isChecked(),
            "PREP_PER_FILE_AUDIO_SETTINGS_PATH": self.per_file_settings_path.text(),
            "PREP_ANALYSIS_REPORT_PATH": self.analysis_report_path.text(),
            "PREP_ANALYSIS_HINTS_GLOBAL": self._global_analysis_hints,
            "PREP_ANALYSIS_HINTS_BY_FILE": self._file_analysis_hints,
        }

    def set_values(self, d):
        self.input_folder.setText(
            str(d.get("PREP_INPUT_FOLDER", self.input_folder.text())))
        self.specify_files_cb.setChecked(
            bool(d.get("PREP_SPECIFY_FILES", False)))
        self.selected_files_combo.set_selected_files(
            d.get("PREP_SELECTED_FILES", []))
        if self.specify_files_cb.isChecked():
            self.selected_files_combo.ensure_first_selected()
        self.excel_path.setText(
            str(d.get("PREP_EXCEL_PATH", self.excel_path.text())))
        self._use_signal_hints_cb.setChecked(bool(d.get("PREP_USE_SIGNAL_HINTS", True)))
        self._apply_muter_cb.setChecked(bool(d.get("PREP_APPLY_TO_MUTER", True)))
        self._apply_extractor_cb.setChecked(bool(d.get("PREP_APPLY_TO_EXTRACTOR", True)))

        self.per_file_settings_path.setText(
            str(d.get("PREP_PER_FILE_AUDIO_SETTINGS_PATH",
                      self.per_file_settings_path.text())))
        self.analysis_report_path.setText(
            str(d.get("PREP_ANALYSIS_REPORT_PATH",
                      self.analysis_report_path.text())))

        self._global_analysis_hints = _prep_sanitize_hints(
            d.get("PREP_ANALYSIS_HINTS_GLOBAL", self._global_analysis_hints))
        raw_by_file = d.get("PREP_ANALYSIS_HINTS_BY_FILE", {})
        cleaned = {}
        if isinstance(raw_by_file, dict):
            for fn, hints in raw_by_file.items():
                cleaned[str(fn)] = _prep_sanitize_hints(hints)
        self._file_analysis_hints = cleaned
        self._update_hints_summary()

        folder = self.input_folder.text().strip()
        if os.path.isdir(folder):
            self._sync_default_analysis_paths(folder, overwrite=False)


__all__ = ["PipelinePrepPanel"]