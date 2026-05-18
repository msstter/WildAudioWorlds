"""Tests for the shared beat/tempo engine and compatibility wrapper."""

from pathlib import Path

import numpy as np
import openpyxl

from scripts import beat_tempo_step
from scripts.beat_tempo_engine import BeatTempoConfig, find_audio_files, plp_local_maxima, run_beat_tempo_step


def test_find_audio_files_filters_and_sorts_supported_extensions(tmp_path):
    audio_dir = tmp_path / "audio"
    audio_dir.mkdir()
    (audio_dir / "b.wav").write_bytes(b"")
    (audio_dir / "a.mp3").write_bytes(b"")
    (audio_dir / "notes.txt").write_text("ignore", encoding="utf-8")

    assert find_audio_files(str(audio_dir)) == ["a.mp3", "b.wav"]


def test_plp_local_maxima_keeps_strongest_peak_within_spacing_window():
    pulse_curve = np.array([0.0, 0.5, 0.2, 0.9, 0.1, 0.2, 0.0, 0.8, 0.0], dtype=float)
    pulse_times = np.arange(len(pulse_curve), dtype=float) * 0.05

    maxima = plp_local_maxima(pulse_curve, pulse_times, min_spacing_ms=130.0)

    assert np.allclose(maxima, np.array([0.15, 0.35], dtype=float))


def test_run_beat_tempo_step_writes_summary_and_detail_rows(monkeypatch, tmp_path):
    audio_dir = tmp_path / "audio"
    audio_dir.mkdir()
    (audio_dir / "demo.wav").write_bytes(b"")
    output_excel = tmp_path / "beat_tempo.xlsx"

    config = BeatTempoConfig(
        audio_folder=str(audio_dir),
        output_excel_path=str(output_excel),
        export_pulse_detail=True,
    )

    monkeypatch.setattr(
        "scripts.beat_tempo_engine.process_file",
        lambda file_path, config: {
            "beat_tempo_bpm": 123.456,
            "beat_times": np.array([0.1, 0.6], dtype=float),
            "pulse_curve": np.array([0.1, 0.9, 0.2], dtype=float),
            "pulse_times": np.array([0.0, 0.5, 1.0], dtype=float),
            "plp_peak_times": np.array([0.5], dtype=float),
        },
    )

    exit_code = run_beat_tempo_step(config, print_fn=lambda *args, **kwargs: None)

    assert exit_code == 0

    workbook = openpyxl.load_workbook(output_excel)
    ws_summary = workbook[config.output_sheet_name]
    headers = {cell.value: idx for idx, cell in enumerate(ws_summary[1], start=1) if cell.value}

    assert ws_summary.cell(row=2, column=headers["File Name"]).value == "demo.wav"
    assert ws_summary.cell(row=2, column=headers["Beat Count"]).value == 2
    assert ws_summary.cell(row=2, column=headers["PLP Peak Count"]).value == 1

    ws_detail = workbook[f"{config.output_sheet_name} (Detail)"]
    assert ws_detail.max_row == 4
    assert ws_detail.cell(row=3, column=1).value == "demo.wav"
    assert ws_detail.cell(row=3, column=5).value == 1


def test_beat_tempo_step_main_delegates_to_engine_config(monkeypatch, tmp_path):
    captured = {}

    monkeypatch.setattr(beat_tempo_step, "audio_folder", str(tmp_path / "audio"))
    monkeypatch.setattr(beat_tempo_step, "output_excel_path", str(tmp_path / "out.xlsx"))
    monkeypatch.setattr(beat_tempo_step, "BEAT_HOP_LENGTH", 256)
    monkeypatch.setattr(beat_tempo_step, "PLP_PEAK_MIN_SPACING_MS", 180)

    def fake_run(config):
        captured["config"] = config
        return 0

    monkeypatch.setattr(beat_tempo_step, "run_beat_tempo_step", fake_run)

    assert beat_tempo_step.main() == 0
    assert captured["config"].audio_folder == str(tmp_path / "audio")
    assert captured["config"].output_excel_path == str(tmp_path / "out.xlsx")
    assert captured["config"].beat_hop_length == 256
    assert captured["config"].plp_peak_min_spacing_ms == 180


def test_resolve_output_excel_path_falls_back_to_extractor_workbook(tmp_path):
    extractor_path = str(tmp_path / "extractor.xlsx")

    assert beat_tempo_step._resolve_output_excel_path(
        {"output_excel_path": "   "},
        {"output_excel_path": extractor_path},
        "",
    ) == extractor_path

    explicit_path = str(tmp_path / "beat_tempo.xlsx")
    assert beat_tempo_step._resolve_output_excel_path(
        {"output_excel_path": explicit_path},
        {"output_excel_path": extractor_path},
        "",
    ) == explicit_path