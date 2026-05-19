"""Tests for the shared standalone script output bridge."""

from __future__ import annotations

import sys
from pathlib import Path

import pandas as pd


ROOT = Path(__file__).resolve().parents[1]
SCRIPTS_DIR = ROOT / "scripts"
if str(SCRIPTS_DIR) not in sys.path:
    sys.path.insert(0, str(SCRIPTS_DIR))


from shared_output_writers import (  # noqa: E402
    save_matplotlib_figure,
    save_openpyxl_workbook,
    write_csv_dataframe,
    write_text_output,
    write_workbook_sheets,
)


def test_shared_output_writers_publish_text_csv_workbook_and_figure(tmp_path):
    import matplotlib.pyplot as plt
    from openpyxl import load_workbook

    text_path = tmp_path / "reports" / "summary.txt"
    csv_path = tmp_path / "reports" / "summary.csv"
    workbook_path = tmp_path / "reports" / "summary.xlsx"
    figure_path = tmp_path / "reports" / "plot.png"

    write_text_output(text_path, "alpha\nbeta\n")
    write_csv_dataframe(csv_path, pd.DataFrame({"value": [1, 2]}), index=False)

    def _postprocess_workbook(temp_workbook_path: Path) -> None:
        workbook = load_workbook(temp_workbook_path)
        worksheet = workbook["Summary"]
        worksheet["B2"] = "annotated"
        workbook.save(temp_workbook_path)

    write_workbook_sheets(
        workbook_path,
        {"Summary": pd.DataFrame({"name": ["one"], "value": [1]})},
        postprocess_workbook=_postprocess_workbook,
    )

    figure, axis = plt.subplots()
    axis.plot([0, 1], [1, 0])
    save_matplotlib_figure(figure, figure_path, dpi=80)
    plt.close(figure)

    assert text_path.read_text(encoding="utf-8") == "alpha\nbeta\n"
    assert csv_path.read_text(encoding="utf-8").splitlines() == ["value", "1", "2"]

    workbook = load_workbook(workbook_path)
    worksheet = workbook["Summary"]
    assert worksheet["A2"].value == "one"
    assert worksheet["B2"].value == "annotated"

    worksheet["C2"] = "republished"
    save_openpyxl_workbook(workbook, workbook_path)
    workbook = load_workbook(workbook_path)
    assert workbook["Summary"]["C2"].value == "republished"

    assert figure_path.exists()
    assert figure_path.stat().st_size > 0